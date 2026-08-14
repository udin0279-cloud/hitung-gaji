from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import re
import logging
import uuid
import asyncio
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta, time as dtime, date
from typing import List, Optional, Dict, Any, Tuple

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File, Form, Body
from fastapi.responses import StreamingResponse, HTMLResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
import csv
import io
import base64

# ---------------- DB ----------------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# ---------------- App ----------------
app = FastAPI(title="HRIS API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger("payroll")

JWT_ALGORITHM = "HS256"


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


# ---------------- Auth helpers ----------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def create_portal_token(employee_id: str, email: str) -> str:
    payload = {
        "sub": employee_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(hours=24),
        "type": "portal",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_employee(request: Request) -> Dict[str, Any]:
    token = request.cookies.get("portal_token")
    if not token:
        raise HTTPException(status_code=401, detail="Portal: not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "portal":
            raise HTTPException(status_code=401, detail="Invalid token type")
        emp = await db.employees.find_one({"id": payload["sub"]}, {"_id": 0})
        if not emp:
            raise HTTPException(status_code=401, detail="Karyawan tidak ditemukan")
        return emp
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sesi berakhir")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token tidak valid")


def set_auth_cookies(response: Response, access: str, refresh: str) -> None:
    response.set_cookie("access_token", access, httponly=True, secure=True, samesite="lax", max_age=43200, path="/")
    response.set_cookie("refresh_token", refresh, httponly=True, secure=True, samesite="lax", max_age=604800, path="/")


async def get_current_user(request: Request) -> Dict[str, Any]:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user.pop("_id", None)
        user.pop("password_hash", None)
        # Backwards compatibility: legacy "admin" role acts as super_admin
        if user.get("role") == "admin":
            user["role"] = "super_admin"
        # Legacy hr_leave role → treat as admin_privileged with izin_cuti perm
        if user.get("role") == "hr_leave":
            user["role"] = "admin_privileged"
            user["permissions"] = list(set((user.get("permissions") or []) + ["izin_cuti"]))
        # Ensure permissions field exists
        if user.get("role") == "admin_privileged" and not isinstance(user.get("permissions"), list):
            user["permissions"] = []
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# Role-based access helpers
ROLE_SUPER_ADMIN = "super_admin"
ROLE_ADMIN_PRIVILEGED = "admin_privileged"
ROLE_HR_LEAVE = "hr_leave"  # legacy, migrated to admin_privileged on startup
VALID_ROLES = {ROLE_SUPER_ADMIN, ROLE_ADMIN_PRIVILEGED}

# All menu keys that can be granted to an "Admin dengan Privilege"
MENU_KEYS = [
    "karyawan", "payroll", "inventory", "pembelian", "penjualan",
    "laporan_penjualan", "kas_operasional", "laba_rugi", "master_kategori",
    "thr", "izin_cuti", "kelola_user", "konfigurasi",
]

# Ordered rules: (path_prefix, menu_key). First matching prefix wins.
# Path is the full request path INCLUDING /api prefix.
PATH_MENU_RULES = [
    # ---- Payroll & Slips ----
    ("/api/payroll/thr", "thr"),
    ("/api/payroll/bukti-potong", "payroll"),
    ("/api/payroll", "payroll"),
    ("/api/payslip", "payroll"),
    ("/api/attendance", "payroll"),
    # ---- Employees ----
    ("/api/employees", "karyawan"),
    ("/api/contracts", "karyawan"),
    # ---- Sales sub-routes (most specific first) ----
    ("/api/sales/report", "laporan_penjualan"),
    ("/api/sales/stats", "penjualan"),
    ("/api/sales", "penjualan"),
    ("/api/products", "penjualan"),
    ("/api/inventory/customers", "penjualan"),
    # ---- Inventory / Purchasing ----
    ("/api/inventory", "inventory"),
    ("/api/purchasing", "pembelian"),
    # ---- Cashbook ----
    ("/api/cashbook", "kas_operasional"),
    # ---- Reports (Laba Rugi) ----
    ("/api/reports", "laba_rugi"),
    # ---- Categories ----
    ("/api/categories", "master_kategori"),
    # ---- Leave ----
    ("/api/leave", "izin_cuti"),
    # ---- User Mgmt ----
    ("/api/users", "kelola_user"),
    # ---- Config / Konfigurasi ----
    ("/api/config", "konfigurasi"),
    ("/api/admin/whatsapp", "konfigurasi"),
    ("/api/admin/export-database", "konfigurasi"),
    ("/api/admin/import-database", "konfigurasi"),
]

# Paths that do NOT require menu-permission (still may require login separately).
# Auth endpoints, portal endpoints, root health, dashboard, and public misc.
RBAC_BYPASS_PREFIXES = (
    "/api/auth/",
    "/api/portal/",
    "/api/dashboard",  # any authenticated admin sees a filtered dashboard
    "/api/employees-template.csv",  # only used from imports page (karyawan menu)
    "/api/employees-import",
)


def _menu_for_path(path: str) -> Optional[str]:
    for prefix, menu in PATH_MENU_RULES:
        if path.startswith(prefix):
            return menu
    return None


async def require_super_admin(user: dict = Depends(get_current_user)) -> dict:
    """Backward-compat: allow both super_admin AND admin_privileged.
    Fine-grained menu access is enforced by the RBAC middleware based on path."""
    role = user.get("role")
    if role not in VALID_ROLES:
        raise HTTPException(status_code=403, detail="Akses ditolak")
    return user


async def require_super_admin_strict(user: dict = Depends(get_current_user)) -> dict:
    """Strict: only super_admin. Reserved for meta/system routes."""
    if user.get("role") != ROLE_SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="Akses ditolak: Super Admin only")
    return user


async def require_leave_access(user: dict = Depends(get_current_user)) -> dict:
    role = user.get("role")
    if role == ROLE_SUPER_ADMIN:
        return user
    if role == ROLE_ADMIN_PRIVILEGED and "izin_cuti" in (user.get("permissions") or []):
        return user
    raise HTTPException(status_code=403, detail="Akses ditolak")


# ---------------- Models ----------------
class RegisterIn(BaseModel):
    email: EmailStr
    password: str
    name: str


class LoginIn(BaseModel):
    email: EmailStr
    password: str


class UserOut(BaseModel):
    id: str
    email: str
    name: str
    role: str


# Employee
class EmployeeIn(BaseModel):
    nik: str  # employee internal id (NIP)
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None  # WhatsApp number (Indonesian local 08... or 62...)
    position: str
    department: str
    join_date: str  # ISO date
    basic_salary: float
    fixed_allowance: float = 0  # tunjangan tetap (dianggap sebagai tunjangan taxable / legacy)
    tunjangan_jabatan: float = 0  # taxable, masuk base BPJS & PPh21
    tunjangan_transport: float = 0  # non-taxable benefit
    tunjangan_lainnya: float = 0  # non-taxable benefit
    insentif_individu: float = 0  # taxable untuk PPh21, TIDAK masuk base BPJS
    tunjangan_tidak_tetap: float = 0  # taxable PPh21, TIDAK masuk base BPJS
    tunjangan_wfh: float = 0  # non-taxable benefit
    insentif_kolektif: float = 0  # taxable PPh21, TIDAK masuk base BPJS
    insentif_lain: float = 0  # taxable PPh21, TIDAK masuk base BPJS
    potongan_terlambat: float = 0  # dipotong dari gross
    potongan_pulang_cepat: float = 0  # dipotong dari gross
    loan_installment: float = 0  # angsuran pinjaman bulanan
    loan_total_amount: float = 0  # total nilai pinjaman
    loan_tenor_total: int = 0  # total bulan tenor pinjaman
    loan_tenor_paid: int = 0  # sudah dibayar berapa bulan
    ptkp_status: str = "TK/0"  # TK/0, K/0, K/1, K/2, K/3
    npwp: Optional[str] = None
    has_npwp: bool = True
    bpjs_kesehatan: bool = True
    bpjs_ketenagakerjaan: bool = True
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_account_holder: Optional[str] = None
    employment_status: str = "tetap"  # ojt | kontrak_6 | kontrak_12 | kontrak_24 | tetap
    status_start_date: Optional[str] = None  # ISO YYYY-MM-DD — tanggal mulai OJT/Kontrak
    status_end_date: Optional[str] = None    # ISO YYYY-MM-DD — tanggal berakhir
    active: bool = True


class EmployeeOut(EmployeeIn):
    id: str
    created_at: str


# Payroll
class PayrollRunIn(BaseModel):
    period: str  # YYYY-MM
    attendance: Dict[str, Dict[str, float]] = Field(default_factory=dict)
    # attendance[employee_id] = {"days_worked": 22, "overtime_hours": 0, "bonus": 0, "deduction": 0}
    overrides: Dict[str, Dict[str, float]] = Field(default_factory=dict)
    # overrides[employee_id] = {"transport": Rp, "inc_individu": Rp, "inc_kolektif": Rp,
    #                            "inc_lain": Rp, "thr": Rp, "pinjaman": Rp}
    # Nilai override menggantikan angka default dari master karyawan / attendance.


# ---------------- Indonesian Payroll Configuration (overridable via /api/config) ----------------
CONFIG: Dict[str, Any] = {
    "ptkp_table": {
        "TK/0": 54_000_000,
        "TK/1": 58_500_000,
        "TK/2": 63_000_000,
        "TK/3": 67_500_000,
        "K/0": 58_500_000,
        "K/1": 63_000_000,
        "K/2": 67_500_000,
        "K/3": 72_000_000,
    },
    # PPh 21 brackets (UU HPP 2022) — list of [limit, rate]; last limit uses None for infinity
    "pph21_brackets": [
        [60_000_000, 0.05],
        [250_000_000, 0.15],
        [500_000_000, 0.25],
        [5_000_000_000, 0.30],
        [None, 0.35],
    ],
    "bpjs_kesehatan_employee": 0.01,
    "bpjs_kesehatan_employer": 0.04,
    "bpjs_kesehatan_max_base": 12_000_000,
    "jht_employee": 0.02,
    "jht_employer": 0.037,
    "jp_employee": 0.01,
    "jp_employer": 0.02,
    "jp_max_base": 10_042_300,
    "jkk_employer": 0.0024,
    "jkm_employer": 0.003,
    "biaya_jabatan_rate": 0.05,
    "biaya_jabatan_max_year": 6_000_000,
    "standard_workdays": 22,
    "overtime_multiplier": 1.5,
    # Tarif lembur per jam (Rp). Jika > 0, pakai nilai ini secara langsung (tanpa formula 1/173).
    # Jika 0 (default), pakai formula standar: (basic / 173) * overtime_multiplier
    "overtime_hourly_rate": 0,
}


def _pph21_brackets_normalized():
    out = []
    for b in CONFIG["pph21_brackets"]:
        limit = b[0] if b[0] is not None else float("inf")
        out.append((float(limit), float(b[1])))
    return out


def compute_pph21_annual(pkp: float) -> float:
    if pkp <= 0:
        return 0.0
    tax = 0.0
    prev_limit = 0.0
    for limit, rate in _pph21_brackets_normalized():
        taxable_in_bracket = min(pkp, limit) - prev_limit
        if taxable_in_bracket <= 0:
            break
        tax += taxable_in_bracket * rate
        prev_limit = limit
        if pkp <= limit:
            break
    return tax


def calculate_payslip(employee: Dict[str, Any], attendance: Dict[str, float], overrides: Optional[Dict[str, float]] = None) -> Dict[str, Any]:
    """
    Returns full payslip breakdown for one employee for one month.
    `overrides` (opsional) — dict berisi angka override yg dikirim dari UI Jalankan Payroll
    (Transport, Inc.Individu/Kolektif/Lain, THR, Pinjaman). Jika key ada dan bernilai numeric,
    nilainya menggantikan default dari master karyawan.
    """
    o = overrides or {}
    def _ov(key: str, default_val: float) -> float:
        if key in o and o[key] is not None:
            try:
                return float(o[key])
            except (TypeError, ValueError):
                return default_val
        return default_val
    basic = float(employee.get("basic_salary", 0))
    fixed_allowance = float(employee.get("fixed_allowance", 0))
    tj_jabatan = float(employee.get("tunjangan_jabatan", 0))
    tj_transport = _ov("transport", float(employee.get("tunjangan_transport", 0)))
    tj_lainnya = float(employee.get("tunjangan_lainnya", 0))
    insentif_individu = _ov("inc_individu", float(employee.get("insentif_individu", 0)))
    tj_tidak_tetap = float(employee.get("tunjangan_tidak_tetap", 0))
    tj_wfh = float(employee.get("tunjangan_wfh", 0))
    insentif_kolektif = _ov("inc_kolektif", float(employee.get("insentif_kolektif", 0)))
    insentif_lain = _ov("inc_lain", float(employee.get("insentif_lain", 0)))
    thr_override = _ov("thr", 0.0)  # THR baru ditambah via override UI (default 0)
    potongan_terlambat = float(employee.get("potongan_terlambat", 0))
    potongan_pulang_cepat = float(employee.get("potongan_pulang_cepat", 0))
    loan_installment = _ov("pinjaman", float(employee.get("loan_installment", 0)))
    loan_tenor_total = int(employee.get("loan_tenor_total", 0) or 0)
    loan_tenor_paid = int(employee.get("loan_tenor_paid", 0) or 0)
    overtime_hours = float(attendance.get("overtime_hours", 0) or 0)
    # THR override ditambahkan ke bonus (taxable earning yg masuk gross tanpa base BPJS)
    bonus = float(attendance.get("bonus", 0) or 0) + thr_override
    other_deduction = float(attendance.get("deduction", 0) or 0)
    standard_days = float(CONFIG["standard_workdays"]) or 22.0
    days_worked = float(attendance.get("days_worked", standard_days) or standard_days)
    late_penalty_minutes = float(attendance.get("late_penalty_minutes", 0) or 0)

    # Overtime pay — RUMUS BARU (per-menit berdasarkan gaji pokok):
    #   Upah/Hari  = Gaji Pokok / Jumlah Hari Kerja per Bulan (CONFIG.standard_workdays)
    #   Upah/Jam   = Upah/Hari / 7   (7 jam kerja normal per hari)
    #   Upah/Menit = Upah/Jam / 60
    #   Total Lembur = Total Menit Overtime × Upah/Menit
    # `overtime_hours` di attendance sudah dalam satuan JAM (dgn jeda 30m grace period),
    # convert ke menit dulu.
    WORK_HOURS_PER_DAY = 7
    _wage_per_day = (basic / standard_days) if standard_days > 0 else 0
    _wage_per_hour = _wage_per_day / WORK_HOURS_PER_DAY if _wage_per_day else 0
    _wage_per_minute = _wage_per_hour / 60 if _wage_per_hour else 0
    overtime_minutes = float(overtime_hours) * 60.0
    overtime_rate_per_hour = round(_wage_per_hour, 2)
    overtime_pay = round(overtime_minutes * _wage_per_minute, 2)
    _ot_source = "auto_pro_rata_daily"  # source untuk display di slip

    # === Potongan Otomatis Terlambat > 4 Jam ===
    # Menit telat > 240 (4 jam) dijumlah dari attendance_daily, dikalikan wage_per_minute.
    # AUTO OVERRIDE MANUAL: bila auto > 0, override field manual `potongan_terlambat` di master employee.
    auto_late_penalty = round(late_penalty_minutes * _wage_per_minute, 2) if late_penalty_minutes > 0 else 0.0
    if auto_late_penalty > 0:
        potongan_terlambat = auto_late_penalty  # override manual master value
        _late_penalty_source = "auto_from_attendance"
    else:
        _late_penalty_source = "manual_employee_master" if potongan_terlambat > 0 else "none"

    # Pro-rate basic if days_worked < standard
    prorate_factor = min(days_worked / standard_days, 1.0) if standard_days > 0 else 1.0
    basic_paid = basic * prorate_factor

    # Taxable = tunjangan yang masuk BPJS & PPh21 base
    taxable_allowance = fixed_allowance + tj_jabatan
    # Non-taxable = benefit (transport, lain-lain, WFH) — masuk gross tapi tidak masuk base
    non_taxable_allowance = tj_transport + tj_lainnya + tj_wfh
    # Taxable-only allowance (PPh21 kena, BPJS tidak) — insentif + tunjangan tidak tetap
    taxable_only_earnings = insentif_individu + insentif_kolektif + insentif_lain + tj_tidak_tetap

    gross = basic_paid + taxable_allowance + non_taxable_allowance + taxable_only_earnings + overtime_pay + bonus

    # BPJS Kesehatan (capped) — base = basic + taxable allowance only
    bpjs_kes_base = min(basic_paid + taxable_allowance, CONFIG["bpjs_kesehatan_max_base"]) if employee.get("bpjs_kesehatan") else 0
    bpjs_kes_employee = bpjs_kes_base * CONFIG["bpjs_kesehatan_employee"]
    bpjs_kes_employer = bpjs_kes_base * CONFIG["bpjs_kesehatan_employer"]

    # BPJS Ketenagakerjaan
    has_btk = employee.get("bpjs_ketenagakerjaan", True)
    jht_base = basic_paid + taxable_allowance if has_btk else 0
    jp_base = min(basic_paid + taxable_allowance, CONFIG["jp_max_base"]) if has_btk else 0

    jht_employee = jht_base * CONFIG["jht_employee"]
    jht_employer = jht_base * CONFIG["jht_employer"]
    jp_employee = jp_base * CONFIG["jp_employee"]
    jp_employer = jp_base * CONFIG["jp_employer"]
    jkk_employer = jht_base * CONFIG["jkk_employer"]
    jkm_employer = jht_base * CONFIG["jkm_employer"]

    # Annual PPh21 calculation — only taxable earnings contribute
    taxable_gross_monthly = basic_paid + taxable_allowance + taxable_only_earnings + overtime_pay + bonus
    bruto_monthly = taxable_gross_monthly + bpjs_kes_employer + jkk_employer + jkm_employer
    bruto_yearly = bruto_monthly * 12

    biaya_jabatan_yearly = min(bruto_yearly * CONFIG["biaya_jabatan_rate"], CONFIG["biaya_jabatan_max_year"])
    iuran_pengurang_yearly = (jht_employee + jp_employee) * 12

    netto_yearly = bruto_yearly - biaya_jabatan_yearly - iuran_pengurang_yearly
    ptkp = CONFIG["ptkp_table"].get(employee.get("ptkp_status", "TK/0"), 54_000_000)
    pkp = max(0, netto_yearly - ptkp)
    # Round down PKP to thousands per UU HPP
    pkp = (pkp // 1000) * 1000

    pph21_yearly = compute_pph21_annual(pkp)
    # Non-NPWP surcharge: +20%
    if not employee.get("has_npwp", True):
        pph21_yearly *= 1.2
    pph21_monthly = pph21_yearly / 12

    # Loan deduction: only if tenor still active
    loan_active = loan_installment > 0 and (loan_tenor_total == 0 or loan_tenor_paid < loan_tenor_total)
    loan_deduction = loan_installment if loan_active else 0

    # Sisa Pinjaman (Rp) — nilai total pinjaman yang belum diangsur.
    # Formula robust: coba loan_total_amount, fallback ke (installment × tenor_total)
    # jika loan_total_amount belum diset di DB. Setelah slip ini, tenor_paid akan +1
    # → sisa = total - (installment × tenor_paid_after).
    loan_total_amount = float(employee.get("loan_total_amount", 0) or 0)
    if loan_active:
        tenor_paid_after = loan_tenor_paid + 1
        # Fallback: jika loan_total_amount belum diset, hitung dari installment × tenor_total
        if loan_total_amount <= 0 and loan_tenor_total > 0:
            loan_total_amount = loan_installment * loan_tenor_total
        # Final sisa (setelah angsuran slip ini)
        if loan_total_amount > 0:
            loan_remaining_amount = max(0.0, loan_total_amount - loan_installment * tenor_paid_after)
        elif loan_tenor_total > 0:
            # Tanpa loan_total_amount + tanpa tenor → tidak bisa hitung sisa
            loan_remaining_amount = loan_installment * max(0, loan_tenor_total - tenor_paid_after)
        else:
            # Loan ongoing tanpa tenor tetap (open-ended) → sisa tidak diketahui
            loan_remaining_amount = 0.0
    else:
        loan_remaining_amount = 0.0

    # === PERUBAHAN 2026-08-05: HANYA 2 POTONGAN DI SLIP ===
    # Sesuai permintaan user (Opsi B): Total Potongan & Gaji Bersih HANYA dihitung dari
    # (1) Angsuran Pinjaman + (2) Potongan Lain-lain. BPJS/JHT/JP/PPh21 tetap DIHITUNG untuk
    # keperluan laporan tahunan/bukti potong, TAPI TIDAK dipotong dari take-home pay.
    # potongan_terlambat & potongan_pulang_cepat digabung ke Potongan Lain-lain (visible).
    other_deduction_combined = other_deduction + potongan_terlambat + potongan_pulang_cepat
    total_deductions_visible = loan_deduction + other_deduction_combined
    net_salary = gross - total_deductions_visible

    return {
        "earnings": {
            "basic_salary": round(basic_paid, 2),
            "fixed_allowance": round(fixed_allowance, 2),
            "tunjangan_jabatan": round(tj_jabatan, 2),
            "tunjangan_transport": round(tj_transport, 2),
            "tunjangan_lainnya": round(tj_lainnya, 2),
            "insentif_individu": round(insentif_individu, 2),
            "tunjangan_tidak_tetap": round(tj_tidak_tetap, 2),
            "tunjangan_wfh": round(tj_wfh, 2),
            "insentif_kolektif": round(insentif_kolektif, 2),
            "insentif_lain": round(insentif_lain, 2),
            "overtime": round(overtime_pay, 2),
            "bonus": round(bonus, 2),
            "gross": round(gross, 2),
        },
        "deductions": {
            "bpjs_kesehatan_employee": round(bpjs_kes_employee, 2),
            "jht_employee": round(jht_employee, 2),
            "jp_employee": round(jp_employee, 2),
            "pph21": round(pph21_monthly, 2),
            "loan": round(loan_deduction, 2),
            "other_deduction": round(other_deduction_combined, 2),
            # Simpan komponen mentah untuk audit; TIDAK ditampilkan di slip
            "other_deduction_raw": round(other_deduction, 2),
            "potongan_terlambat": round(potongan_terlambat, 2),
            "potongan_pulang_cepat": round(potongan_pulang_cepat, 2),
            "total": round(total_deductions_visible, 2),
        },
        "loan_info": {
            "active": loan_active,
            "installment": round(loan_installment, 2),
            "tenor_total": loan_tenor_total,
            "tenor_paid_before": loan_tenor_paid,
            "tenor_paid_after": loan_tenor_paid + 1 if loan_active else loan_tenor_paid,
            "remaining_after": max(0, loan_tenor_total - (loan_tenor_paid + 1)) if loan_active and loan_tenor_total else 0,
            "total_amount": round(loan_total_amount, 2),
            "remaining_amount": round(loan_remaining_amount, 2),
        },
        "employer_contributions": {
            "bpjs_kesehatan_employer": round(bpjs_kes_employer, 2),
            "jht_employer": round(jht_employer, 2),
            "jp_employer": round(jp_employer, 2),
            "jkk_employer": round(jkk_employer, 2),
            "jkm_employer": round(jkm_employer, 2),
        },
        "tax_detail": {
            "bruto_yearly": round(bruto_yearly, 2),
            "biaya_jabatan_yearly": round(biaya_jabatan_yearly, 2),
            "netto_yearly": round(netto_yearly, 2),
            "ptkp": ptkp,
            "pkp": round(pkp, 2),
            "pph21_yearly": round(pph21_yearly, 2),
        },
        "attendance": {
            "days_worked": days_worked,
            "overtime_hours": overtime_hours,
            "overtime_minutes": round(overtime_minutes, 2),
            "overtime_rate_per_hour": round(overtime_rate_per_hour, 2),
            "overtime_rate_per_minute": round(_wage_per_minute, 2),
            "wage_per_day": round(_wage_per_day, 2),
            "work_hours_per_day": WORK_HOURS_PER_DAY,
            "standard_workdays": standard_days,
            "overtime_rate_source": _ot_source,  # "auto_pro_rata_daily"
            "overtime_multiplier": float(CONFIG["overtime_multiplier"]),
            "late_penalty_minutes": round(late_penalty_minutes, 2),
            "late_penalty_amount": round(auto_late_penalty, 2),
            "late_penalty_source": _late_penalty_source,
        },
        "net_salary": round(net_salary, 2),
    }


# ---------------- Auth Endpoints ----------------
@api_router.post("/auth/register")
async def register(payload: RegisterIn, response: Response):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = str(uuid.uuid4())
    doc = {
        "id": user_id,
        "email": email,
        "name": payload.name,
        "role": "admin",
        "password_hash": hash_password(payload.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    access = create_access_token(user_id, email)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {"id": user_id, "email": email, "name": payload.name, "role": "admin"}


@api_router.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    email = payload.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    access = create_access_token(user["id"], email)
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return _user_view(user)


@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return _user_view(user)


# ---------------- Employee Portal (Self-service) ----------------
class PortalLoginIn(BaseModel):
    email: EmailStr
    nik: str


# [MOVED to routers/portal.py]
#   POST /portal/login
#   POST /portal/logout
#   GET  /portal/me
#   GET  /portal/payslips
#   GET  /portal/payslip/{slip_id}
#   GET  /portal/payslip/{slip_id}/pdf
#   GET  /portal/thr


# ---------------- Annual Summary & Bukti Potong 1721-A1 ----------------
async def _build_annual_summary(employee_id: str, year: int) -> Dict[str, Any]:
    year_prefix = f"{year}-"
    slips = await db.payslips.find(
        {"employee_id": employee_id, "period": {"$regex": f"^{year_prefix}"}},
        {"_id": 0},
    ).sort("period", 1).to_list(length=24)
    thrs = await db.thr_slips.find(
        {"employee_id": employee_id, "period": {"$regex": f"^{year_prefix}"}},
        {"_id": 0},
    ).to_list(length=12)

    months = {}
    for s in slips:
        e, d = s["earnings"], s["deductions"]
        months[s["period"]] = {
            "period": s["period"],
            "gross": e["gross"],
            "basic": e["basic_salary"],
            "allowance": e["fixed_allowance"],
            "overtime": e["overtime"],
            "bonus": e["bonus"],
            "pph21": d["pph21"],
            "bpjs_employee": d["bpjs_kesehatan_employee"] + d["jht_employee"] + d["jp_employee"],
            "net": s["net_salary"],
        }

    total_gross = sum(m["gross"] for m in months.values())
    total_basic = sum(m["basic"] for m in months.values())
    total_allowance = sum(m["allowance"] for m in months.values())
    total_overtime = sum(m["overtime"] for m in months.values())
    total_bonus = sum(m["bonus"] for m in months.values())
    total_pph21 = sum(m["pph21"] for m in months.values())
    total_bpjs = sum(m["bpjs_employee"] for m in months.values())
    total_net = sum(m["net"] for m in months.values())

    total_thr_gross = sum(t["thr_gross"] for t in thrs)
    total_thr_pph = sum(t["pph21_thr"] for t in thrs)

    return {
        "year": year,
        "months": sorted(months.values(), key=lambda x: x["period"]),
        "totals": {
            "gross": round(total_gross, 2),
            "basic": round(total_basic, 2),
            "allowance": round(total_allowance, 2),
            "overtime": round(total_overtime, 2),
            "bonus": round(total_bonus, 2),
            "pph21": round(total_pph21, 2),
            "bpjs_employee": round(total_bpjs, 2),
            "net": round(total_net, 2),
            "thr_gross": round(total_thr_gross, 2),
            "thr_pph21": round(total_thr_pph, 2),
        },
        "months_count": len(months),
    }


def _build_bukti_potong_pdf(employee: Dict[str, Any], summary: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5, leading=11)
    bold_style = ParagraphStyle("bold", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=8.5, leading=11)
    tiny = ParagraphStyle("tiny", parent=styles["Normal"], fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#71717a"))

    company = os.environ.get("COMPANY_NAME", "PLAZAKREASI DIGITAL PRINTING")
    year = summary["year"]
    totals = summary["totals"]

    story = []
    # Header
    header = Table([
        [Paragraph("<b>FORMULIR 1721-A1</b><br/><font size=7>BUKTI PEMOTONGAN PAJAK PENGHASILAN PASAL 21</font>", bold_style),
         Paragraph(f"<para alignment='right'><font size=7>MASA PEROLEHAN PENGHASILAN<br/></font><b>JANUARI &nbsp;–&nbsp; DESEMBER {year}</b></para>", body_style)],
    ], colWidths=[110 * mm, 70 * mm])
    header.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 1.5, colors.HexColor("#18181b")),
        ("LINEBELOW", (0, -1), (-1, -1), 1.5, colors.HexColor("#18181b")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f4f4f5")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)
    story.append(Spacer(1, 6))

    # Employer info
    story.append(Paragraph("<b>A. IDENTITAS PEMOTONG</b>", bold_style))
    employer = Table([
        ["Nama Pemotong", ":", company],
        ["NPWP Pemotong", ":", os.environ.get("COMPANY_NPWP", "00.000.000.0-000.000")],
    ], colWidths=[40 * mm, 6 * mm, 130 * mm])
    employer.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(employer)
    story.append(Spacer(1, 6))

    # Employee info
    story.append(Paragraph("<b>B. IDENTITAS PENERIMA PENGHASILAN</b>", bold_style))
    emp_table = Table([
        ["Nama", ":", employee.get("name", ""), "NIK/NIP", ":", employee.get("nik", "")],
        ["NPWP", ":", employee.get("npwp") or "—", "Status PTKP", ":", employee.get("ptkp_status", "")],
        ["Jabatan", ":", employee.get("position", ""), "Departemen", ":", employee.get("department", "")],
    ], colWidths=[22 * mm, 5 * mm, 60 * mm, 25 * mm, 5 * mm, 63 * mm])
    emp_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(emp_table)
    story.append(Spacer(1, 10))

    # Bagian C: Rincian Penghasilan
    story.append(Paragraph("<b>C. RINCIAN PENGHASILAN DAN PENGHITUNGAN PPh PASAL 21</b>", bold_style))

    biaya_jabatan = min(totals["gross"] * CONFIG["biaya_jabatan_rate"], CONFIG["biaya_jabatan_max_year"])
    ptkp = CONFIG["ptkp_table"].get(employee.get("ptkp_status", "TK/0"), 54_000_000)
    # Use the per-month progressive PPh already computed in slips (includes THR PPh separately)
    pph_total = totals["pph21"] + totals["thr_pph21"]

    rows = [
        ["NO", "PENGHASILAN", "JUMLAH (Rp)"],
        ["1", "Gaji/Tunjangan Tetap", _format_idr(totals["basic"] + totals["allowance"])],
        ["2", "Uang Lembur", _format_idr(totals["overtime"])],
        ["3", "Bonus / Tantiem", _format_idr(totals["bonus"])],
        ["4", "Tunjangan Hari Raya (THR)", _format_idr(totals["thr_gross"])],
        ["5", "Jumlah Penghasilan Bruto (1+2+3+4)", _format_idr(totals["gross"] + totals["thr_gross"])],
        ["", "PENGURANGAN", ""],
        ["6", "Biaya Jabatan", _format_idr(biaya_jabatan)],
        ["7", "Iuran BPJS Karyawan (Kes + JHT + JP)", _format_idr(totals["bpjs_employee"])],
        ["8", "Jumlah Pengurangan (6+7)", _format_idr(biaya_jabatan + totals["bpjs_employee"])],
        ["", "PERHITUNGAN PPh 21", ""],
        ["9", "Penghasilan Neto (5−8)", _format_idr(totals["gross"] + totals["thr_gross"] - biaya_jabatan - totals["bpjs_employee"])],
        ["10", f"PTKP Setahun ({employee.get('ptkp_status', 'TK/0')})", _format_idr(ptkp)],
        ["11", "PKP (9−10)", _format_idr(max(0, totals["gross"] + totals["thr_gross"] - biaya_jabatan - totals["bpjs_employee"] - ptkp))],
        ["12", "PPh 21 Terutang Setahun", _format_idr(pph_total)],
        ["13", "PPh 21 yang sudah dipotong", _format_idr(pph_total)],
        ["14", "PPh 21 kurang/(lebih) dipotong", _format_idr(0)],
    ]
    t = Table(rows, colWidths=[10 * mm, 110 * mm, 60 * mm])
    style = TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f4f4f5")),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("FONTNAME", (2, 1), (2, -1), "Courier"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d4d4d8")),
        ("BACKGROUND", (0, 6), (-1, 6), colors.HexColor("#f4f4f5")),
        ("BACKGROUND", (0, 10), (-1, 10), colors.HexColor("#f4f4f5")),
        ("FONTNAME", (0, 5), (-1, 5), "Helvetica-Bold"),
        ("FONTNAME", (0, 9), (-1, 9), "Helvetica-Bold"),
        ("FONTNAME", (0, 11), (-1, 11), "Helvetica-Bold"),
        ("FONTNAME", (0, 14), (-1, 14), "Helvetica-Bold"),
        ("LINEABOVE", (0, 14), (-1, 14), 1.2, colors.HexColor("#18181b")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    t.setStyle(style)
    story.append(t)
    story.append(Spacer(1, 12))

    # Monthly breakdown
    story.append(Paragraph("<b>D. RINCIAN BULANAN</b>", bold_style))
    m_rows = [["BULAN", "BRUTO", "PPh 21", "BPJS", "TAKE HOME"]]
    for m in summary["months"]:
        m_rows.append([m["period"], _format_idr(m["gross"]), _format_idr(m["pph21"]), _format_idr(m["bpjs_employee"]), _format_idr(m["net"])])
    m_rows.append(["TOTAL", _format_idr(totals["gross"]), _format_idr(totals["pph21"]), _format_idr(totals["bpjs_employee"]), _format_idr(totals["net"])])
    mt = Table(m_rows, colWidths=[30 * mm, 35 * mm, 35 * mm, 35 * mm, 45 * mm])
    mt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f4f4f5")),
        ("FONTNAME", (1, 1), (-1, -1), "Courier"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, -1), (-1, -1), "Courier-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fafafa")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#d4d4d8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(mt)
    story.append(Spacer(1, 12))

    # Footer signature
    story.append(Paragraph(
        "<i>Catatan: Formulir ini merupakan ringkasan dari rincian slip gaji bulanan. "
        "Gunakan sebagai referensi untuk SPT Tahunan PPh Orang Pribadi.</i>",
        tiny,
    ))
    story.append(Spacer(1, 14))
    sign_table = Table([
        ["", ""],
        ["Diterima oleh,", "Diterbitkan oleh Pemotong,"],
        ["", ""],
        [Paragraph(f"<b>{employee.get('name','')}</b><br/><font size=7>NIK: {employee.get('nik','')}</font>", body_style),
         Paragraph(f"<b>{company}</b><br/><font size=7>HR / Pajak</font>", body_style)],
    ], colWidths=[90 * mm, 90 * mm])
    sign_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TOPPADDING", (0, 2), (-1, 2), 24),
    ]))
    story.append(sign_table)

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


# [MOVED to routers/portal.py]
#   GET  /portal/annual/{year}
#   GET  /portal/bukti-potong/{year}/pdf


# Admin can also download bukti potong for any employee
# [MOVED to routers/payroll.py]  GET /payroll/bukti-potong/{employee_id}/{year}/pdf


# ---------------- Portal Magic Link (Forgot NIK) ----------------
class ForgotPortalIn(BaseModel):
    email: EmailStr


# [MOVED to routers/portal.py]
#   POST /portal/forgot
#   POST /portal/magic-login


# ---------------- Employee Endpoints (Admin) ----------------
# [MOVED to routers/employees.py]
#   GET    /employees
#   POST   /employees
#   GET    /employees/{employee_id}
#   PUT    /employees/{employee_id}
#   DELETE /employees/{employee_id}


# ---------------- Payroll Endpoints ----------------
# [MOVED to routers/payroll.py]
#   POST   /payroll/preview
#   POST   /payroll/run
#   GET    /payroll/runs
#   GET    /payroll/runs/{period}/slips
#   GET    /payroll/payslip/{slip_id}
#   DELETE /payroll/runs/{period}


# ---------------- PDF Export ----------------
def _format_idr(n: float) -> str:
    try:
        return "Rp " + f"{int(round(n)):,}".replace(",", ".")
    except Exception:
        return "Rp 0"


def _build_payslip_pdf(slip: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    label_style = ParagraphStyle("lbl", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7, textColor=colors.HexColor("#71717a"))
    val_style = ParagraphStyle("val", parent=styles["Normal"], fontName="Helvetica", fontSize=10, textColor=colors.HexColor("#18181b"))
    section_style = ParagraphStyle("sec", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor("#18181b"), spaceAfter=4, spaceBefore=6)

    story = []
    # Header
    header = Table(
        [[
            Paragraph("<b>PAYROLL.ID</b><br/><font size=7 color='#71717a'>HR · TAX · BPJS</font>", val_style),
            Paragraph(f"<para alignment='right'><font size=7 color='#71717a'>SLIP GAJI</font><br/><b><font size=14>Periode {slip['period']}</font></b><br/><font size=7 color='#71717a' face='Courier'>No: {slip['id'][:8].upper()}</font></para>", val_style),
        ]],
        colWidths=[90 * mm, 90 * mm],
    )
    header.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 1.5, colors.HexColor("#18181b")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(header)
    story.append(Spacer(1, 8))

    # Employee
    emp = Table(
        [
            [Paragraph("NAMA KARYAWAN", label_style), Paragraph(slip["name"], val_style),
             Paragraph("NIK", label_style), Paragraph(f"<font face='Courier'>{slip['nik']}</font>", val_style)],
            [Paragraph("JABATAN / DEPT", label_style), Paragraph(f"{slip['position']} · {slip['department']}", val_style),
             Paragraph("PTKP / NPWP", label_style), Paragraph(f"<font face='Courier'>{slip['ptkp_status']} · {'Ya' if slip.get('has_npwp', True) else 'Tidak'}</font>", val_style)],
        ],
        colWidths=[35 * mm, 55 * mm, 30 * mm, 60 * mm],
    )
    emp.setStyle(TableStyle([
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("LINEBELOW", (0, -1), (-1, -1), 0.4, colors.HexColor("#e4e4e7")),
    ]))
    story.append(emp)
    story.append(Spacer(1, 10))

    # Earnings & Deductions side by side
    e = slip["earnings"]
    d = slip["deductions"]
    earn_rows = [
        ["PENDAPATAN", ""],
        ["Gaji Pokok", _format_idr(e["basic_salary"])],
    ]
    if e.get("fixed_allowance", 0):
        earn_rows.append(["Tunjangan Tetap", _format_idr(e["fixed_allowance"])])
    if e.get("tunjangan_jabatan", 0):
        earn_rows.append(["Tj. Jabatan", _format_idr(e["tunjangan_jabatan"])])
    if e.get("tunjangan_transport", 0):
        earn_rows.append(["Tj. Transport", _format_idr(e["tunjangan_transport"])])
    if e.get("tunjangan_lainnya", 0):
        earn_rows.append(["Tj. Lain-lain", _format_idr(e["tunjangan_lainnya"])])
    if e.get("tunjangan_tidak_tetap", 0):
        earn_rows.append(["Tj. Tidak Tetap", _format_idr(e["tunjangan_tidak_tetap"])])
    if e.get("tunjangan_wfh", 0):
        earn_rows.append(["Insentif WFH", _format_idr(e["tunjangan_wfh"])])
    if e.get("insentif_individu", 0):
        earn_rows.append(["Insentif Individu", _format_idr(e["insentif_individu"])])
    if e.get("insentif_kolektif", 0):
        earn_rows.append(["Insentif Kolektif", _format_idr(e["insentif_kolektif"])])
    if e.get("insentif_lain", 0):
        earn_rows.append(["Insentif Lain-lain", _format_idr(e["insentif_lain"])])
    if e.get("overtime", 0):
        _att = slip.get("attendance", {})
        _ot_hours = _att.get("overtime_hours", 0)
        _ot_rate = _att.get("overtime_rate_per_hour", 0)
        _label = f"Lembur ({_ot_hours} jam × {_format_idr(_ot_rate)}/jam)" if _ot_rate else "Lembur"
        earn_rows.append([_label, _format_idr(e["overtime"])])
    if e.get("bonus", 0):
        earn_rows.append(["Bonus", _format_idr(e["bonus"])])
    earn_rows.append(["Total Bruto", _format_idr(e["gross"])])

    # === PERUBAHAN 2026-08-05: HANYA 2 POTONGAN DI SLIP (Angsuran + Lain-lain) ===
    # BPJS/JHT/JP/PPh21 DISEMBUNYIKAN sepenuhnya dari slip sesuai permintaan user.
    # Kolom "other_deduction" di backend sudah digabung dengan potongan_terlambat & pulang_cepat.
    deduct_rows = [
        ["POTONGAN", ""],
    ]
    if d.get("loan", 0):
        deduct_rows.append(["Angsuran Pinjaman", _format_idr(d["loan"])])
    if d.get("other_deduction", 0):
        deduct_rows.append(["Potongan Lain-lain", _format_idr(d["other_deduction"])])
    # Sisa Pinjaman — info transparansi untuk karyawan (tidak dipotong lagi, sekedar info)
    _li = slip.get("loan_info", {}) if isinstance(slip, dict) else {}
    if _li.get("active") and _li.get("remaining_amount"):
        _tenor = _li.get("tenor_total", 0)
        _paid_after = _li.get("tenor_paid_after", 0)
        _tenor_info = f" (tenor {_paid_after}/{_tenor})" if _tenor else ""
        deduct_rows.append([f"Sisa Pinjaman{_tenor_info}", _format_idr(_li.get("remaining_amount", 0))])
    deduct_rows.append(["Total Potongan", _format_idr(d["total"])])
    # Pad to same length
    max_len = max(len(earn_rows), len(deduct_rows))
    while len(earn_rows) < max_len:
        earn_rows.insert(-1, ["", ""])

    def style_box(rows, header_color):
        t = Table(rows, colWidths=[48 * mm, 40 * mm])
        s = TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("TEXTCOLOR", (0, 0), (-1, 0), header_color),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#a1a1aa")),
            ("FONTNAME", (1, 1), (1, -1), "Courier"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, -1), (1, -1), "Courier-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.4, colors.HexColor("#71717a")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ])
        t.setStyle(s)
        return t

    two_col = Table(
        [[style_box(earn_rows, colors.HexColor("#008A00")), style_box(deduct_rows, colors.HexColor("#E81123"))]],
        colWidths=[90 * mm, 90 * mm],
    )
    two_col.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    story.append(two_col)
    story.append(Spacer(1, 12))

    # Take home (dark block)
    th = Table(
        [[
            Paragraph(f"<font size=7 color='#a1a1aa'>TAKE HOME PAY</font><br/><font size=7 color='#a1a1aa'>Hari kerja: {slip['attendance']['days_worked']} · Lembur: {slip['attendance']['overtime_hours']} jam</font>", val_style),
            Paragraph(f"<para alignment='right'><b><font face='Courier-Bold' size=18 color='white'>{_format_idr(slip['net_salary'])}</font></b></para>", val_style),
        ]],
        colWidths=[90 * mm, 90 * mm],
    )
    th.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#18181b")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(th)
    story.append(Spacer(1, 12))

    # Rincian PPh 21 disembunyikan atas permintaan user

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


# [MOVED to routers/payroll.py]  GET /payroll/payslip/{slip_id}/pdf


# ---------------- Employee CSV Import ----------------
EMPLOYEE_CSV_HEADERS = [
    "nik", "name", "email", "phone", "position", "department", "join_date",
    "basic_salary", "fixed_allowance", "ptkp_status", "npwp", "has_npwp",
    "bpjs_kesehatan", "bpjs_ketenagakerjaan", "bank_name", "bank_account",
]


def _parse_bool(v: str, default: bool = True) -> bool:
    if v is None or v == "":
        return default
    return str(v).strip().lower() in ("1", "true", "yes", "ya", "y")


# [MOVED to routers/employees.py]
#   GET  /employees-template.csv
#   POST /employees-import


# ---------------- Attendance Router (extracted 2026-08-01) ----------------
from routers.attendance import make_router as _make_attendance_router
api_router.include_router(_make_attendance_router(db, require_super_admin, logger))


# ---------------- Dashboard ----------------
@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(require_super_admin)):
    total_employees = await db.employees.count_documents({"active": True})
    runs = await db.payroll_runs.find({}, {"_id": 0}).sort("period", -1).to_list(length=12)
    latest = runs[0] if runs else None
    trend = list(reversed([
        {"period": r["period"], "total_net": r["total_net"], "total_gross": r["total_gross"]}
        for r in runs
    ]))
    # Kontrak/OJT akan habis dalam 90 hari
    expiring = await _find_expiring_contracts(90)
    # Inventory summary utk widget dashboard
    inv_summary = {
        "total_materials": 0,
        "total_stock_value": 0.0,
        "low_stock_count": 0,
        "total_waste_this_month": 0.0,
        "waste_records_this_month": 0,
        "top_waste": [],
    }
    try:
        mats = await db.materials.find({}, {"_id": 0}).to_list(length=5000)
        total_stock_value = sum(float(m.get("current_stock", 0)) * float(m.get("purchase_price", 0)) for m in mats)
        low_stock = [m for m in mats if m.get("active", True) and float(m.get("min_stock", 0)) > 0 and float(m.get("current_stock", 0)) <= float(m.get("min_stock", 0))]
        today = datetime.now(timezone.utc).date()
        month_start = today.replace(day=1).isoformat()
        waste_month = await db.waste.find({"date": {"$gte": month_start}}, {"_id": 0}).to_list(length=5000)
        total_waste_month = sum(float(w.get("estimated_loss", 0)) for w in waste_month)
        # top waste
        top_agg: Dict[str, Dict[str, Any]] = {}
        mat_by_id = {m["id"]: m for m in mats}
        for w in waste_month:
            mid = w.get("material_id")
            if not mid:
                continue
            mat = mat_by_id.get(mid, {})
            row = top_agg.setdefault(mid, {"material_name": mat.get("name") or "-", "material_unit": mat.get("unit") or "", "qty": 0.0, "loss": 0.0})
            row["qty"] += float(w.get("quantity", 0))
            row["loss"] += float(w.get("estimated_loss", 0))
        top_waste = sorted(top_agg.values(), key=lambda r: r["loss"], reverse=True)[:3]
        for r in top_waste:
            r["qty"] = round(r["qty"], 4)
            r["loss"] = round(r["loss"], 2)
        inv_summary = {
            "total_materials": sum(1 for m in mats if m.get("active", True)),
            "total_stock_value": round(total_stock_value, 2),
            "low_stock_count": len(low_stock),
            "total_waste_this_month": round(total_waste_month, 2),
            "waste_records_this_month": len(waste_month),
            "top_waste": top_waste,
        }
    except Exception as ex:
        logger.warning(f"Inventory summary failed: {ex}")
    return {
        "total_employees": total_employees,
        "latest_run": latest,
        "trend": trend,
        "total_runs": await db.payroll_runs.count_documents({}),
        "contract_expiring": expiring[:5],  # top 5 untuk widget
        "contract_expiring_count": len(expiring),
        "inventory": inv_summary,
        "late_offenders": await _top_late_offenders(month=None, limit=5),
    }


async def _top_late_offenders(month: Optional[str] = None, limit: int = 5) -> Dict[str, Any]:
    """Top N karyawan dengan total menit telat > 4 jam TERBANYAK.

    Args:
      month: filter "YYYY-MM" (default: bulan berjalan)
      limit: berapa top karyawan (default 5)

    Return: {month, items:[{employee_id, nik, name, total_late_minutes, occurrences, estimated_penalty}], total_offenders, no_data}
    """
    if not month:
        today = datetime.now(timezone.utc).date()
        month = today.strftime("%Y-%m")
    # Rentang tanggal
    start = f"{month}-01"
    # akhir bulan
    try:
        y, m = int(month[:4]), int(month[5:7])
    except Exception:
        raise HTTPException(status_code=400, detail="Format bulan harus YYYY-MM")
    if m == 12:
        next_first = date(y + 1, 1, 1)
    else:
        next_first = date(y, m + 1, 1)
    end = (next_first - timedelta(days=1)).isoformat()

    items = await db.attendance_daily.find(
        {"date": {"$gte": start, "$lte": end}, "late_penalty_minutes": {"$gt": 0}},
        {"_id": 0},
    ).to_list(length=50000)

    if not items:
        return {"month": month, "items": [], "total_offenders": 0, "no_data": True}

    # Aggregate per employee_id
    agg: Dict[str, Dict[str, Any]] = {}
    unmatched: Dict[str, Dict[str, Any]] = {}
    for it in items:
        emp_id = it.get("employee_id")
        lpm = float(it.get("late_penalty_minutes") or 0)
        if emp_id:
            entry = agg.setdefault(emp_id, {
                "employee_id": emp_id,
                "nik": it.get("employee_nik"),
                "name": it.get("employee_name"),
                "total_late_minutes": 0.0,
                "occurrences": 0,
            })
            entry["total_late_minutes"] += lpm
            entry["occurrences"] += 1
        else:
            pin = it.get("pin") or "-"
            entry = unmatched.setdefault(pin, {
                "employee_id": None,
                "nik": None,
                "pin": pin,
                "name": it.get("employee_name") or f"PIN {pin}",
                "total_late_minutes": 0.0,
                "occurrences": 0,
            })
            entry["total_late_minutes"] += lpm
            entry["occurrences"] += 1

    # Fetch employee basic_salary utk estimasi penalty nominal
    all_entries = list(agg.values()) + list(unmatched.values())
    emp_ids = [e["employee_id"] for e in all_entries if e.get("employee_id")]
    emp_map: Dict[str, Dict[str, Any]] = {}
    if emp_ids:
        cursor = db.employees.find({"id": {"$in": emp_ids}}, {"_id": 0, "id": 1, "basic_salary": 1, "position": 1, "department": 1})
        for e in await cursor.to_list(length=len(emp_ids)):
            emp_map[e["id"]] = e

    standard_days = float(CONFIG.get("standard_workdays") or 26.0) or 26.0
    for e in all_entries:
        basic = float((emp_map.get(e.get("employee_id"), {}) or {}).get("basic_salary") or 0)
        wpm = ((basic / standard_days) / 7.0) / 60.0 if basic > 0 else 0.0
        e["total_late_minutes"] = round(e["total_late_minutes"], 2)
        e["estimated_penalty"] = round(e["total_late_minutes"] * wpm, 2)
        e["position"] = (emp_map.get(e.get("employee_id"), {}) or {}).get("position")
        e["department"] = (emp_map.get(e.get("employee_id"), {}) or {}).get("department")

    all_entries.sort(key=lambda x: (x["total_late_minutes"], x["occurrences"]), reverse=True)
    return {
        "month": month,
        "items": all_entries[:limit],
        "total_offenders": len(all_entries),
        "no_data": False,
    }


@api_router.get("/dashboard/late-offenders")
async def dashboard_late_offenders(
    user: dict = Depends(require_super_admin),
    month: Optional[str] = None,
    limit: int = 5,
):
    """Top N karyawan dengan total menit telat > 4 jam terbanyak dalam sebulan."""
    if limit < 1 or limit > 50:
        limit = 5
    return await _top_late_offenders(month=month, limit=limit)


async def _find_expiring_contracts(days_ahead: int = 60) -> List[Dict[str, Any]]:
    """Return employees dgn status OJT/Kontrak yg status_end_date jatuh dalam N hari (termasuk sudah lewat)."""
    today = datetime.now(timezone.utc).date()
    cutoff = today + timedelta(days=days_ahead)
    cursor = db.employees.find({
        "active": True,
        "employment_status": {"$in": ["ojt", "kontrak_6", "kontrak_12", "kontrak_24"]},
        "status_end_date": {"$ne": None, "$exists": True},
    }, {"_id": 0}).sort("status_end_date", 1)
    items = await cursor.to_list(length=500)
    result = []
    for e in items:
        end_str = e.get("status_end_date")
        if not end_str:
            continue
        try:
            end_date = datetime.strptime(end_str[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if end_date > cutoff:
            continue
        days_left = (end_date - today).days
        result.append({
            "id": e.get("id"),
            "nik": e.get("nik"),
            "name": e.get("name"),
            "position": e.get("position"),
            "department": e.get("department"),
            "employment_status": e.get("employment_status"),
            "status_start_date": e.get("status_start_date"),
            "status_end_date": end_str,
            "days_left": days_left,
            "expired": days_left < 0,
        })
    return result


@api_router.get("/contracts/expiring")
async def contracts_expiring(days: int = 60, user: dict = Depends(require_super_admin)):
    """List karyawan OJT/Kontrak yang berakhir dalam N hari (default 60)."""
    items = await _find_expiring_contracts(days)
    return {"days": days, "count": len(items), "items": items}


# ---------------- Config ----------------
@api_router.get("/config/constants")
async def config_constants(user: dict = Depends(require_super_admin)):
    return {
        "ptkp_table": CONFIG["ptkp_table"],
        "pph21_brackets": [{"limit": b[0], "rate": b[1]} for b in CONFIG["pph21_brackets"]],
        "bpjs": {
            "kesehatan_employee": CONFIG["bpjs_kesehatan_employee"],
            "kesehatan_employer": CONFIG["bpjs_kesehatan_employer"],
            "kesehatan_max_base": CONFIG["bpjs_kesehatan_max_base"],
            "jht_employee": CONFIG["jht_employee"],
            "jht_employer": CONFIG["jht_employer"],
            "jp_employee": CONFIG["jp_employee"],
            "jp_employer": CONFIG["jp_employer"],
            "jp_max_base": CONFIG["jp_max_base"],
            "jkk_employer": CONFIG["jkk_employer"],
            "jkm_employer": CONFIG["jkm_employer"],
        },
        "biaya_jabatan_rate": CONFIG["biaya_jabatan_rate"],
        "biaya_jabatan_max_year": CONFIG["biaya_jabatan_max_year"],
        "standard_workdays": CONFIG["standard_workdays"],
        "overtime_multiplier": CONFIG["overtime_multiplier"],
        "overtime_hourly_rate": CONFIG.get("overtime_hourly_rate", 0),
    }


class ConfigUpdateIn(BaseModel):
    ptkp_table: Optional[Dict[str, float]] = None
    pph21_brackets: Optional[List[List[Any]]] = None  # [[limit_or_null, rate], ...]
    bpjs_kesehatan_employee: Optional[float] = None
    bpjs_kesehatan_employer: Optional[float] = None
    bpjs_kesehatan_max_base: Optional[float] = None
    jht_employee: Optional[float] = None
    jht_employer: Optional[float] = None
    jp_employee: Optional[float] = None
    jp_employer: Optional[float] = None
    jp_max_base: Optional[float] = None
    jkk_employer: Optional[float] = None
    jkm_employer: Optional[float] = None
    biaya_jabatan_rate: Optional[float] = None
    biaya_jabatan_max_year: Optional[float] = None
    standard_workdays: Optional[float] = None
    overtime_multiplier: Optional[float] = None
    overtime_hourly_rate: Optional[float] = None


@api_router.put("/config/constants")
async def update_config(payload: ConfigUpdateIn, user: dict = Depends(require_super_admin)):
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        return {"updated": 0}
    # Normalize pph21_brackets nulls
    if "pph21_brackets" in update:
        normalized = []
        for row in update["pph21_brackets"]:
            limit = row[0] if (row[0] is not None and row[0] != "") else None
            rate = float(row[1])
            normalized.append([float(limit) if limit is not None else None, rate])
        update["pph21_brackets"] = normalized
    CONFIG.update(update)
    await db.app_config.update_one(
        {"id": "payroll_config"},
        {"$set": {**update, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"updated": len(update), "config": update}


async def _load_config_from_db():
    rec = await db.app_config.find_one({"id": "payroll_config"}, {"_id": 0})
    if rec:
        for k in list(CONFIG.keys()):
            if k in rec and rec[k] is not None:
                CONFIG[k] = rec[k]
        logger.info("Loaded payroll config overrides from DB")


# ---------------- THR (Tunjangan Hari Raya) ----------------
class THRRunIn(BaseModel):
    period: str  # YYYY-MM (month THR is paid)


def _months_between(start_iso: str, end_dt: datetime) -> float:
    try:
        sd = datetime.fromisoformat(start_iso)
    except Exception:
        try:
            sd = datetime.strptime(start_iso, "%Y-%m-%d")
        except Exception:
            return 12.0
    delta = (end_dt.year - sd.year) * 12 + (end_dt.month - sd.month)
    if end_dt.day < sd.day:
        delta -= 1
    return max(0.0, float(delta))


def _calculate_thr(employee: Dict[str, Any], reference_dt: datetime) -> Dict[str, Any]:
    """1x (basic + fixed_allowance) for tenure >= 12 months; proportional for < 12 months (min 1 month)."""
    basic = float(employee.get("basic_salary", 0))
    allowance = float(employee.get("fixed_allowance", 0))
    monthly_base = basic + allowance
    months = _months_between(employee.get("join_date", "2020-01-01"), reference_dt)
    if months < 1:
        thr_gross = 0.0
        formula = "Belum berhak (masa kerja < 1 bulan)"
    elif months >= 12:
        thr_gross = monthly_base
        formula = "1x Gaji + Tunjangan Tetap"
    else:
        thr_gross = monthly_base * (months / 12.0)
        formula = f"({months:.0f}/12) x Gaji + Tunjangan Tetap"
    # PPh 21 on THR using progressive bracket on isolated annual amount per Indonesian practice (simplified)
    # Treat THR as standalone annual income for tax (jumlah neto disetahunkan)
    # For simplicity use 5% bracket; for higher PKP brackets, fall back to compute_pph21_annual on (annual_gross + thr) - (annual_gross alone)
    annual_no_thr = monthly_base * 12
    ptkp = CONFIG["ptkp_table"].get(employee.get("ptkp_status", "TK/0"), 54_000_000)
    pkp_no_thr = max(0, annual_no_thr - ptkp)
    pkp_with_thr = max(0, annual_no_thr + thr_gross - ptkp)
    pph_no_thr = compute_pph21_annual(pkp_no_thr)
    pph_with_thr = compute_pph21_annual(pkp_with_thr)
    pph21_thr = max(0.0, pph_with_thr - pph_no_thr)
    if not employee.get("has_npwp", True):
        pph21_thr *= 1.2
    net = thr_gross - pph21_thr
    return {
        "months_of_service": months,
        "monthly_base": round(monthly_base, 2),
        "thr_gross": round(thr_gross, 2),
        "pph21_thr": round(pph21_thr, 2),
        "thr_net": round(net, 2),
        "formula": formula,
    }


# [MOVED to routers/payroll.py]
#   POST /payroll/thr/preview
#   POST /payroll/thr/run
#   GET  /payroll/thr/runs
#   GET  /payroll/thr/{period}/slips


# ---------------- Email Payslip ----------------
def _payslip_html(slip: Dict[str, Any]) -> str:
    e, d = slip["earnings"], slip["deductions"]
    li = slip.get("loan_info", {}) or {}
    company = os.environ.get("COMPANY_NAME", "PLAZAKREASI DIGITAL PRINTING")
    # Baris potongan dinamis — HANYA Angsuran Pinjaman + Potongan Lain-lain (per Opsi B, 2026-08-05)
    deduction_rows_html = ""
    if d.get("loan", 0):
        deduction_rows_html += f'<tr><td>Angsuran Pinjaman</td><td align="right" style="font-family:monospace;">{_format_idr(d["loan"])}</td></tr>'
    if d.get("other_deduction", 0):
        deduction_rows_html += f'<tr><td>Potongan Lain-lain</td><td align="right" style="font-family:monospace;">{_format_idr(d["other_deduction"])}</td></tr>'
    # Info Sisa Pinjaman (tidak dipotong, sekedar info karyawan)
    sisa_row = ""
    if li.get("active") and li.get("remaining_amount"):
        _tenor = li.get("tenor_total", 0)
        _paid_after = li.get("tenor_paid_after", 0)
        _tenor_info = f" · tenor {_paid_after}/{_tenor}" if _tenor else ""
        sisa_row = f'<tr style="color:#71717a;font-size:12px;"><td>Sisa Pinjaman{_tenor_info}</td><td align="right" style="font-family:monospace;">{_format_idr(li.get("remaining_amount", 0))}</td></tr>'
    return f"""
    <table width="100%" cellpadding="0" cellspacing="0" style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; color: #18181b;">
      <tr><td style="padding: 24px 0; border-bottom: 2px solid #18181b;">
        <h2 style="margin:0; font-size:20px; letter-spacing:-0.5px;">SLIP GAJI · {slip['period']}</h2>
        <div style="font-size:12px; color:#71717a; margin-top:4px;">{company} · HR Department</div>
      </td></tr>
      <tr><td style="padding: 20px 0;">
        Halo <strong>{slip['name']}</strong>,<br/><br/>
        Berikut adalah ringkasan slip gaji Anda untuk periode <strong>{slip['period']}</strong>.
        Slip lengkap terlampir sebagai PDF.
      </td></tr>
      <tr><td>
        <table width="100%" cellpadding="8" cellspacing="0" style="border-collapse: collapse; font-size: 13px;">
          <tr style="background:#f4f4f5;"><th align="left" style="border-bottom:1px solid #e4e4e7;">PENDAPATAN</th><th align="right" style="border-bottom:1px solid #e4e4e7;">Rp</th></tr>
          <tr><td>Gaji Pokok</td><td align="right" style="font-family:monospace;">{_format_idr(e['basic_salary'])}</td></tr>
          <tr><td>Tunjangan</td><td align="right" style="font-family:monospace;">{_format_idr(e['fixed_allowance'])}</td></tr>
          <tr><td>Lembur</td><td align="right" style="font-family:monospace;">{_format_idr(e['overtime'])}</td></tr>
          <tr><td>Bonus</td><td align="right" style="font-family:monospace;">{_format_idr(e['bonus'])}</td></tr>
          <tr style="font-weight:bold; border-top:1px solid #a1a1aa;"><td>Total Bruto</td><td align="right" style="font-family:monospace;">{_format_idr(e['gross'])}</td></tr>
          <tr style="background:#f4f4f5;"><th align="left" style="border-bottom:1px solid #e4e4e7;">POTONGAN</th><th align="right" style="border-bottom:1px solid #e4e4e7;">Rp</th></tr>
          {deduction_rows_html}
          {sisa_row}
          <tr style="font-weight:bold; border-top:1px solid #a1a1aa;"><td>Total Potongan</td><td align="right" style="font-family:monospace;">{_format_idr(d['total'])}</td></tr>
        </table>
      </td></tr>
      <tr><td style="padding:16px; background:#18181b; color:white; margin-top:12px;">
        <table width="100%"><tr>
          <td style="font-size:11px; color:#a1a1aa; text-transform:uppercase; letter-spacing:1px;">Take Home Pay</td>
          <td align="right" style="font-family:monospace; font-size:22px; font-weight:bold;">{_format_idr(slip['net_salary'])}</td>
        </tr></table>
      </td></tr>
      <tr><td style="padding: 16px 0; font-size:11px; color:#71717a;">
        Email otomatis dari sistem payroll. Jangan dibalas. Untuk pertanyaan hubungi HR.
      </td></tr>
    </table>
    """


def _send_email_via_resend(to_email: str, subject: str, html: str, pdf_bytes: bytes, pdf_filename: str) -> Dict[str, Any]:
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    sender = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")
    if not api_key:
        return {"status": "mocked", "to": to_email, "subject": subject, "message": "RESEND_API_KEY belum diatur — email tidak dikirim (mode mock)."}
    try:
        import resend
        import base64 as b64
        resend.api_key = api_key
        params = {
            "from": sender,
            "to": [to_email],
            "subject": subject,
            "html": html,
            "attachments": [{
                "filename": pdf_filename,
                "content": b64.b64encode(pdf_bytes).decode("utf-8"),
            }],
        }
        result = resend.Emails.send(params)
        return {"status": "sent", "to": to_email, "email_id": result.get("id")}
    except Exception as ex:
        return {"status": "failed", "to": to_email, "error": str(ex)[:200]}


def _send_simple_email(to_email: str, subject: str, html: str) -> Dict[str, Any]:
    api_key = os.environ.get("RESEND_API_KEY", "").strip()
    sender = os.environ.get("SENDER_EMAIL", "onboarding@resend.dev")
    if not api_key or not to_email:
        return {"status": "mocked", "to": to_email}
    try:
        import resend
        resend.api_key = api_key
        result = resend.Emails.send({"from": sender, "to": [to_email], "subject": subject, "html": html})
        return {"status": "sent", "to": to_email, "email_id": result.get("id")}
    except Exception as ex:
        return {"status": "failed", "to": to_email, "error": str(ex)[:200]}


# [MOVED to routers/payroll.py]
#   POST /payroll/payslip/{slip_id}/email
#   POST /payroll/runs/{period}/email-all


# ---------------- Bank Transfer Export ----------------
def _format_bank_export(slips: List[Dict[str, Any]], emp_by_id: Dict[str, Dict[str, Any]], fmt: str, period: str) -> tuple:
    """Returns (content_bytes, filename, mimetype) for given format."""
    rows = []
    for s in slips:
        emp = emp_by_id.get(s["employee_id"], {})
        # Gunakan bank_account_holder bila diisi (kasus rekening atas nama istri/orangtua), fallback ke nama karyawan
        recipient_name = emp.get("bank_account_holder") or s["name"]
        rows.append({
            "nik": s["nik"],
            "name": recipient_name,
            "bank": emp.get("bank_name") or "",
            "account": emp.get("bank_account") or "",
            "amount": round(float(s["net_salary"])),
            "note": f"Gaji {period}",
        })

    out = io.StringIO()
    if fmt == "bca":
        # Format KlikBCA Business sederhana: SourceAcc|Date|DestAcc|Amount|Reference
        out.write("ACCOUNT_NUMBER|DATE|DESTINATION_ACCOUNT|AMOUNT|REFERENCE\n")
        for r in rows:
            out.write(f"SENDER|{period}-25|{r['account']}|{r['amount']}|{r['note']}\n")
        filename = f"bca-payroll-{period}.txt"
        mime = "text/plain"
    elif fmt == "mandiri":
        # Mandiri Cash Management CSV
        out.write("Nama Penerima,Bank Penerima,Nomor Rekening,Jumlah,Berita\n")
        for r in rows:
            out.write(f'"{r["name"]}","{r["bank"]}","{r["account"]}",{r["amount"]},"{r["note"]}"\n')
        filename = f"mandiri-payroll-{period}.csv"
        mime = "text/csv"
    elif fmt == "bni":
        out.write("NomorRekening,NamaPenerima,Bank,Jumlah,Keterangan\n")
        for r in rows:
            out.write(f'"{r["account"]}","{r["name"]}","{r["bank"]}",{r["amount"]},"{r["note"]}"\n')
        filename = f"bni-payroll-{period}.csv"
        mime = "text/csv"
    elif fmt == "bri":
        out.write("NoRekening;NamaPenerima;BankPenerima;Nominal;Berita\n")
        for r in rows:
            out.write(f'{r["account"]};{r["name"]};{r["bank"]};{r["amount"]};{r["note"]}\n')
        filename = f"bri-payroll-{period}.csv"
        mime = "text/csv"
    else:  # generic
        out.write("NIK,Nama,Bank,No Rekening,Jumlah,Keterangan\n")
        for r in rows:
            out.write(f'"{r["nik"]}","{r["name"]}","{r["bank"]}","{r["account"]}",{r["amount"]},"{r["note"]}"\n')
        filename = f"payroll-{period}.csv"
        mime = "text/csv"
    return out.getvalue().encode("utf-8"), filename, mime


# [MOVED to routers/payroll.py]  GET /payroll/runs/{period}/bank-export


# ---------------- Database Backup & Restore (Admin) ----------------
BACKUP_COLLECTIONS = [
    "users",
    "employees",
    "payslips",
    "payroll_runs",
    "thr_slips",
    "thr_runs",
    "attendance_imports",
    "app_config",
    "email_logs",
    "portal_reset_tokens",
    "leave_requests",
]


def _json_default(o):
    if isinstance(o, datetime):
        return o.isoformat()
    if hasattr(o, "isoformat"):
        return o.isoformat()
    return str(o)


@api_router.get("/admin/export-database")
async def export_database(user: dict = Depends(require_super_admin)):
    """Download a full JSON backup of all collections (admin only)."""
    import json as _json

    snapshot: Dict[str, Any] = {
        "_meta": {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "exported_by": user.get("email"),
            "db_name": os.environ.get("DB_NAME"),
            "version": 1,
        }
    }

    for col_name in BACKUP_COLLECTIONS:
        docs = await db[col_name].find({}, {"_id": 0}).to_list(length=100000)
        snapshot[col_name] = docs

    content = _json.dumps(snapshot, default=_json_default, ensure_ascii=False, indent=2).encode("utf-8")
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    filename = f"payroll-backup-{stamp}.json"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.post("/admin/import-database")
async def import_database(
    file: UploadFile = File(...),
    mode: str = "merge",  # 'replace' or 'merge' (default: safer 'merge')
    user: dict = Depends(require_super_admin),
):
    """Restore database from a JSON backup created via export-database.
    mode='replace': drop existing data in each collection then insert from backup.
    mode='merge': upsert by `id` field (keeps current admin and non-backed-up records).
    """
    import json as _json

    if mode not in {"replace", "merge"}:
        raise HTTPException(status_code=400, detail="mode harus 'replace' atau 'merge'")

    raw = await file.read()
    try:
        snapshot = _json.loads(raw.decode("utf-8-sig"))
    except Exception as ex:
        raise HTTPException(status_code=400, detail=f"File JSON tidak valid: {ex}")

    if not isinstance(snapshot, dict) or "_meta" not in snapshot:
        raise HTTPException(status_code=400, detail="Format backup tidak dikenali")

    summary: Dict[str, Any] = {"mode": mode, "restored": {}, "errors": []}

    for col_name in BACKUP_COLLECTIONS:
        docs = snapshot.get(col_name) or []
        if not isinstance(docs, list):
            summary["errors"].append(f"{col_name}: bukan list, dilewati")
            continue
        try:
            if mode == "replace":
                await db[col_name].delete_many({})
                if docs:
                    await db[col_name].insert_many(docs)
            else:  # merge
                for d in docs:
                    if "id" in d:
                        await db[col_name].update_one({"id": d["id"]}, {"$set": d}, upsert=True)
                    else:
                        await db[col_name].insert_one(d)
            summary["restored"][col_name] = len(docs)
        except Exception as ex:
            summary["errors"].append(f"{col_name}: {str(ex)[:200]}")

    # Reload config in memory
    await _load_config_from_db()

    return summary


# ---------------- WhatsApp via Fonnte ----------------
def _normalize_phone_id(phone: str) -> Optional[str]:
    """Normalize Indonesian phone to '62xxx' format. Returns None if invalid."""
    if not phone:
        return None
    p = "".join(c for c in str(phone) if c.isdigit())
    if not p:
        return None
    if p.startswith("0"):
        p = "62" + p[1:]
    elif p.startswith("62"):
        pass
    elif p.startswith("8"):
        p = "62" + p
    if len(p) < 10 or len(p) > 15:
        return None
    return p


def _whatsapp_slip_message(slip: Dict[str, Any], employee: Dict[str, Any]) -> str:
    company = os.environ.get("COMPANY_NAME", "PLAZAKREASI DIGITAL PRINTING")
    portal = (os.environ.get("PUBLIC_APP_URL", "").rstrip("/")) + "/portal/login"
    take_home = f"{int(round(slip['net_salary'])):,}".replace(",", ".")
    gross = f"{int(round(slip['earnings']['gross'])):,}".replace(",", ".")
    pph = f"{int(round(slip['deductions']['pph21'])):,}".replace(",", ".")
    return (
        f"Halo {employee['name']},\n\n"
        f"Slip gaji periode *{slip['period']}* telah tersedia.\n\n"
        f"💰 Take Home: Rp {take_home}\n"
        f"📊 Bruto: Rp {gross}\n"
        f"🧾 PPh 21: Rp {pph}\n\n"
        f"Lihat slip lengkap di portal:\n{portal}\n"
        f"(Email: {employee.get('email') or '-'}, NIK: {employee['nik']})\n\n"
        f"— {company}"
    )


async def _send_whatsapp(phone: str, message: str) -> Dict[str, Any]:
    """Kirim WhatsApp via Fonnte. Return: {status: 'sent'|'mocked'|'failed', ...}.

    Format nomor: normalize ke '62xxx' via `_normalize_phone_id`. Fonnte akan menerima
    nomor dgn awalan 62 langsung — tidak perlu kirim `countryCode` lagi (menghindari
    prefix ganda seperti '6262...').
    """
    import httpx

    target = _normalize_phone_id(phone)
    if not target:
        return {"status": "failed", "phone": phone, "reason": "Nomor WhatsApp tidak valid"}

    token = os.environ.get("FONNTE_TOKEN", "").strip()
    if not token:
        logger.info(f"[MOCK WA] to {target}: {message[:60]}...")
        return {"status": "mocked", "phone": target, "reason": "FONNTE_TOKEN belum diatur di .env"}

    base_url = os.environ.get("FONNTE_BASE_URL", "https://api.fonnte.com").rstrip("/")
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            resp = await client.post(
                f"{base_url}/send",
                data={"target": target, "message": message},
                headers={"Authorization": token},
            )
        try:
            payload = resp.json()
        except Exception:
            payload = {"raw": resp.text}

        # Fonnte success: HTTP 200 + {"status": true, "id": [...]}
        # Fonnte error:   HTTP 4xx/5xx atau {"status": false, "reason": "..."}
        is_ok = resp.status_code < 400 and payload.get("status") is not False
        if not is_ok:
            raw_reason = payload.get("reason") or payload.get("message") or f"HTTP {resp.status_code}"
            reason = str(raw_reason).strip()
            # Terjemahkan error Fonnte umum ke bahasa yg mudah dipahami
            reason_lc = reason.lower()
            if "disconnected" in reason_lc or "logged out" in reason_lc:
                reason = (
                    "Perangkat WhatsApp di dashboard Fonnte sedang OFFLINE / logout. "
                    "Silakan login ke https://md.fonnte.com dan scan ulang QR untuk menyambungkan WhatsApp Anda."
                )
            elif "quota" in reason_lc or "limit" in reason_lc:
                reason = f"Kuota Fonnte habis atau melewati limit. Detail: {reason}"
            elif "invalid token" in reason_lc or "unauthorized" in reason_lc:
                reason = "Token FONNTE_TOKEN salah atau kadaluarsa. Perbarui token di .env dan restart backend."
            elif "target invalid" in reason_lc or "not valid" in reason_lc:
                reason = f"Nomor WhatsApp tidak valid / belum terdaftar di WhatsApp. Detail: {reason}"
            logger.warning(f"Fonnte send failed to {target}: {raw_reason} | resp={payload}")
            return {
                "status": "failed",
                "phone": target,
                "reason": reason[:400],
                "http_status": resp.status_code,
            }

        # Success
        fonnte_id = payload.get("id")
        if isinstance(fonnte_id, list) and fonnte_id:
            fonnte_id = fonnte_id[0]
        logger.info(f"Fonnte OK → {target} (id={fonnte_id})")
        return {"status": "sent", "phone": target, "fonnte_id": fonnte_id}
    except httpx.TimeoutException:
        return {"status": "failed", "phone": target, "reason": "Timeout menghubungi Fonnte (>20s)"}
    except httpx.RequestError as ex:
        return {"status": "failed", "phone": target, "reason": f"Koneksi gagal: {str(ex)[:150]}"}
    except Exception as ex:
        logger.error(f"Fonnte send exception: {ex}")
        return {"status": "failed", "phone": phone, "reason": str(ex)[:200]}


# [MOVED to routers/payroll.py]
#   POST /payroll/payslip/{slip_id}/whatsapp
#   POST /payroll/runs/{period}/whatsapp-all


@api_router.get("/admin/whatsapp/status")
async def whatsapp_status(user: dict = Depends(require_super_admin)):
    has_token = bool(os.environ.get("FONNTE_TOKEN", "").strip())
    return {
        "configured": has_token,
        "provider": "Fonnte",
        "mode": "live" if has_token else "mock",
    }


# ---------------- Leave & Permission Module ----------------
LEAVE_TYPES = {"terlambat", "pulang_awal", "tidak_masuk", "sakit", "lembur"}
LEAVE_TYPE_LABELS = {
    "terlambat": "Izin Datang Terlambat",
    "pulang_awal": "Izin Pulang Awal",
    "tidak_masuk": "Izin Tidak Masuk",
    "sakit": "Izin Sakit",
    "lembur": "Izin Lembur",
}
MAX_ATTACHMENT_SIZE = 2 * 1024 * 1024  # 2 MB
ALLOWED_ATTACHMENT_MIME = {"application/pdf", "image/jpeg", "image/jpg", "image/png"}


def _leave_view(doc: dict, include_attachment_meta: bool = True) -> dict:
    out = {
        "id": doc["id"],
        "employee_id": doc["employee_id"],
        "employee_name": doc.get("employee_name"),
        "employee_nik": doc.get("employee_nik"),
        "department": doc.get("department"),
        "type": doc["type"],
        "type_label": LEAVE_TYPE_LABELS.get(doc["type"], doc["type"]),
        "date_start": doc["date_start"],
        "date_end": doc["date_end"],
        "time_minutes": doc.get("time_minutes"),
        "time_start": doc.get("time_start"),
        "time_end": doc.get("time_end"),
        "reason": doc.get("reason", ""),
        "status": doc.get("status", "pending"),
        "hr_note": doc.get("hr_note"),
        "submitted_at": doc.get("submitted_at"),
        "reviewed_at": doc.get("reviewed_at"),
        "reviewed_by": doc.get("reviewed_by"),
    }
    if include_attachment_meta and doc.get("attachment"):
        att = doc["attachment"]
        out["attachment"] = {
            "filename": att.get("filename"),
            "mime": att.get("mime"),
            "size": att.get("size"),
        }
    else:
        out["attachment"] = None
    return out


# [MOVED to routers/portal.py]
#   POST   /portal/leave
#   GET    /portal/leave
#   DELETE /portal/leave/{leave_id}
#   GET    /portal/leave/{leave_id}/attachment


# ----- HR Admin endpoints -----
@api_router.get("/leave")
async def admin_leave_list(
    status: Optional[str] = None,
    type: Optional[str] = None,
    user: dict = Depends(require_leave_access),
):
    q: Dict[str, Any] = {}
    if status:
        q["status"] = status
    if type:
        q["type"] = type
    items = await db.leave_requests.find(q, {"_id": 0, "attachment.data_base64": 0}).sort("submitted_at", -1).to_list(length=1000)
    return [_leave_view(x) for x in items]


@api_router.get("/leave/stats")
async def admin_leave_stats(user: dict = Depends(require_leave_access)):
    pending = await db.leave_requests.count_documents({"status": "pending"})
    approved = await db.leave_requests.count_documents({"status": "approved"})
    rejected = await db.leave_requests.count_documents({"status": "rejected"})
    return {"pending": pending, "approved": approved, "rejected": rejected, "total": pending + approved + rejected}


@api_router.get("/leave/{leave_id}")
async def admin_leave_detail(leave_id: str, user: dict = Depends(require_leave_access)):
    doc = await db.leave_requests.find_one({"id": leave_id}, {"_id": 0, "attachment.data_base64": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    return _leave_view(doc)


@api_router.get("/leave/{leave_id}/attachment")
async def admin_leave_attachment(leave_id: str, user: dict = Depends(require_leave_access)):
    doc = await db.leave_requests.find_one({"id": leave_id})
    if not doc or not doc.get("attachment"):
        raise HTTPException(status_code=404, detail="Lampiran tidak ditemukan")
    att = doc["attachment"]
    data = base64.b64decode(att["data_base64"])
    return Response(
        content=data,
        media_type=att["mime"],
        headers={"Content-Disposition": f'inline; filename="{att["filename"]}"'},
    )


@api_router.delete("/leave/{leave_id}")
async def admin_leave_delete(leave_id: str, user: dict = Depends(require_leave_access)):
    doc = await db.leave_requests.find_one({"id": leave_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    await db.leave_requests.delete_one({"id": leave_id})
    return {"ok": True}


class LeaveReviewIn(BaseModel):
    hr_note: Optional[str] = ""


@api_router.put("/leave/{leave_id}/approve")
async def admin_leave_approve(leave_id: str, payload: LeaveReviewIn, user: dict = Depends(require_leave_access)):
    doc = await db.leave_requests.find_one({"id": leave_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    await db.leave_requests.update_one(
        {"id": leave_id},
        {"$set": {
            "status": "approved",
            "hr_note": (payload.hr_note or "").strip(),
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "reviewed_by": user.get("email"),
        }},
    )
    updated = await db.leave_requests.find_one({"id": leave_id}, {"_id": 0, "attachment.data_base64": 0})
    _notify_employee_leave_status(updated, "approved")
    return _leave_view(updated)


@api_router.put("/leave/{leave_id}/reject")
async def admin_leave_reject(leave_id: str, payload: LeaveReviewIn, user: dict = Depends(require_leave_access)):
    doc = await db.leave_requests.find_one({"id": leave_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Pengajuan tidak ditemukan")
    if not (payload.hr_note or "").strip():
        raise HTTPException(status_code=400, detail="Alasan penolakan wajib diisi")
    await db.leave_requests.update_one(
        {"id": leave_id},
        {"$set": {
            "status": "rejected",
            "hr_note": payload.hr_note.strip(),
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
            "reviewed_by": user.get("email"),
        }},
    )
    updated = await db.leave_requests.find_one({"id": leave_id}, {"_id": 0, "attachment.data_base64": 0})
    _notify_employee_leave_status(updated, "rejected")
    return _leave_view(updated)


def _notify_employee_leave_status(leave_doc: dict, status: str):
    try:
        asyncio.create_task(_send_leave_status_email(leave_doc, status))
    except Exception as ex:
        logger.warning(f"Failed to schedule leave notif: {ex}")


async def _send_leave_status_email(leave_doc: dict, status: str):
    try:
        emp = await db.employees.find_one({"id": leave_doc["employee_id"]}, {"_id": 0, "email": 1, "name": 1})
        if not emp or not emp.get("email"):
            return
        label_status = "DISETUJUI" if status == "approved" else "DITOLAK"
        color = "#059669" if status == "approved" else "#dc2626"
        type_label = LEAVE_TYPE_LABELS.get(leave_doc.get("type"), leave_doc.get("type"))
        period = leave_doc["date_start"]
        if leave_doc.get("date_end") and leave_doc["date_end"] != leave_doc["date_start"]:
            period = f"{leave_doc['date_start']} s/d {leave_doc['date_end']}"
        note_html = ""
        if leave_doc.get("hr_note"):
            note_html = f'<p style="margin-top:12px;padding:10px;background:#f5f5f5;border-left:3px solid {color};font-size:13px"><b>Catatan HR:</b> {leave_doc["hr_note"]}</p>'
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:560px">
          <h2 style="color:{color};margin:0 0 8px">Pengajuan Izin {label_status}</h2>
          <p style="color:#444">Halo {emp.get('name', '')},</p>
          <p style="color:#444">Pengajuan izin Anda telah diproses dengan detail berikut:</p>
          <table style="border-collapse:collapse;width:100%;font-size:14px">
            <tr><td style="padding:6px 0;color:#666">Jenis</td><td><b>{type_label}</b></td></tr>
            <tr><td style="padding:6px 0;color:#666">Tanggal</td><td>{period}</td></tr>
            <tr><td style="padding:6px 0;color:#666">Status</td><td style="color:{color}"><b>{label_status}</b></td></tr>
          </table>
          {note_html}
          <p style="margin-top:16px;color:#666;font-size:13px">Terima kasih.</p>
        </div>
        """
        _send_simple_email(emp["email"], f"[Payroll] Pengajuan Izin {label_status}", html)
    except Exception as ex:
        logger.warning(f"Failed to send leave status email: {ex}")


# ----- Leave Monthly Report Export (Excel + PDF) -----
def _parse_month(period: str):
    """period format: YYYY-MM. Returns (start_date, end_date) as YYYY-MM-DD strings."""
    try:
        year, month = period.split("-")
        year, month = int(year), int(month)
        from calendar import monthrange
        last_day = monthrange(year, month)[1]
        return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}", year, month
    except Exception:
        raise HTTPException(status_code=400, detail="Format periode tidak valid (gunakan YYYY-MM)")


async def _fetch_monthly_leave(period: str):
    start, end, year, month = _parse_month(period)
    # Get leaves where the request date range overlaps the month
    items = await db.leave_requests.find(
        {"date_start": {"$lte": end}, "date_end": {"$gte": start}},
        {"_id": 0, "attachment.data_base64": 0},
    ).sort([("date_start", 1), ("employee_name", 1)]).to_list(length=5000)
    return items, start, end, year, month


@api_router.get("/leave/report/{period}/excel")
async def leave_report_excel(period: str, user: dict = Depends(require_leave_access)):
    items, start, end, year, month = await _fetch_monthly_leave(period)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan {period}"

    # Title
    ws["A1"] = f"LAPORAN IZIN KARYAWAN — Periode {period}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="002FA7")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Perusahaan: {os.environ.get('COMPANY_NAME', 'PLAZAKREASI DIGITAL PRINTING')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws.merge_cells("A2:I2")

    # Header row
    headers = ["No", "NIK", "Nama", "Departemen", "Jenis Izin", "Tanggal Mulai", "Tanggal Selesai", "Durasi (menit)", "Alasan"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="27272A")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="thin", color="888888"))

    # Status column added at the end
    ws.cell(row=header_row, column=10, value="Status").font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=header_row, column=10).fill = PatternFill("solid", fgColor="27272A")
    ws.cell(row=header_row, column=10).alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=11, value="Catatan HR").font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=header_row, column=11).fill = PatternFill("solid", fgColor="27272A")
    ws.cell(row=header_row, column=11).alignment = Alignment(horizontal="center")

    status_colors = {"pending": "FEF3C7", "approved": "D1FAE5", "rejected": "FEE2E2"}

    for idx, x in enumerate(items, start=1):
        row = header_row + idx
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=x.get("employee_nik") or "")
        ws.cell(row=row, column=3, value=x.get("employee_name") or "")
        ws.cell(row=row, column=4, value=x.get("department") or "")
        ws.cell(row=row, column=5, value=LEAVE_TYPE_LABELS.get(x.get("type"), x.get("type")))
        ws.cell(row=row, column=6, value=x.get("date_start"))
        ws.cell(row=row, column=7, value=x.get("date_end"))
        # Durasi cell: show detail based on type
        dur_text = ""
        if x.get("time_minutes"):
            if x.get("type") == "pulang_awal" and x.get("time_end"):
                dur_text = f"Pulang {x['time_end']} ({x['time_minutes']} menit lebih awal)"
            elif x.get("time_start") and x.get("time_end"):
                dur_text = f"{x['time_minutes']} menit ({x['time_start']}-{x['time_end']})"
            else:
                dur_text = f"{x['time_minutes']} menit"
        ws.cell(row=row, column=8, value=dur_text)
        ws.cell(row=row, column=9, value=x.get("reason") or "")
        status_cell = ws.cell(row=row, column=10, value=STATUS_LABEL_MAP.get(x.get("status"), x.get("status")))
        status_cell.fill = PatternFill("solid", fgColor=status_colors.get(x.get("status"), "FFFFFF"))
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.font = Font(bold=True, size=9)
        ws.cell(row=row, column=11, value=x.get("hr_note") or "")

    # Summary at the bottom
    summary_row = header_row + len(items) + 2
    pending = sum(1 for x in items if x.get("status") == "pending")
    approved = sum(1 for x in items if x.get("status") == "approved")
    rejected = sum(1 for x in items if x.get("status") == "rejected")
    ws.cell(row=summary_row, column=1, value="RINGKASAN").font = Font(bold=True, size=11)
    ws.cell(row=summary_row + 1, column=1, value="Total Pengajuan")
    ws.cell(row=summary_row + 1, column=2, value=len(items)).font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=1, value="Menunggu")
    ws.cell(row=summary_row + 2, column=2, value=pending)
    ws.cell(row=summary_row + 3, column=1, value="Disetujui")
    ws.cell(row=summary_row + 3, column=2, value=approved)
    ws.cell(row=summary_row + 4, column=1, value="Ditolak")
    ws.cell(row=summary_row + 4, column=2, value=rejected)

    # Column widths
    widths = [5, 14, 24, 18, 20, 14, 14, 12, 36, 12, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="laporan-izin-{period}.xlsx"'},
    )


@api_router.get("/leave/report/{period}/pdf")
async def leave_report_pdf(period: str, user: dict = Depends(require_leave_access)):
    items, start, end, year, month = await _fetch_monthly_leave(period)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=14, textColor=colors.HexColor("#002FA7"), spaceAfter=4)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#666666"))
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=8, leading=10)

    story = []
    story.append(Paragraph("LAPORAN IZIN KARYAWAN", title_style))
    story.append(Paragraph(f"Periode: <b>{period}</b> &nbsp;|&nbsp; Perusahaan: {os.environ.get('COMPANY_NAME', 'PLAZAKREASI DIGITAL PRINTING')}", sub_style))
    story.append(Paragraph(f"Dicetak: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", sub_style))
    story.append(Spacer(1, 8))

    # Table
    table_data = [["No", "NIK", "Nama", "Departemen", "Jenis", "Tgl Mulai", "Tgl Selesai", "Durasi", "Status", "Alasan / Catatan HR"]]
    for idx, x in enumerate(items, start=1):
        note_parts = []
        if x.get("reason"):
            note_parts.append(x["reason"])
        if x.get("hr_note"):
            note_parts.append(f"<i>HR: {x['hr_note']}</i>")
        note_html = "<br/>".join(note_parts) or "-"
        dur_text = ""
        if x.get("time_minutes"):
            if x.get("type") == "pulang_awal" and x.get("time_end"):
                dur_text = f"Pulang {x['time_end']}<br/>{x['time_minutes']}m lebih awal"
            elif x.get("time_start") and x.get("time_end"):
                dur_text = f"{x['time_start']}-{x['time_end']}<br/>{x['time_minutes']}m"
            else:
                dur_text = f"{x['time_minutes']}m"
        table_data.append([
            str(idx),
            x.get("employee_nik") or "",
            Paragraph(x.get("employee_name") or "", cell_style),
            x.get("department") or "",
            LEAVE_TYPE_LABELS.get(x.get("type"), x.get("type")),
            x.get("date_start") or "",
            x.get("date_end") or "",
            Paragraph(dur_text, cell_style),
            STATUS_LABEL_MAP.get(x.get("status"), x.get("status")),
            Paragraph(note_html, cell_style),
        ])

    col_widths = [10 * mm, 20 * mm, 38 * mm, 28 * mm, 32 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm, 52 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#27272A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (5, 0), (8, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFA")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])
    # Color status column
    for ridx, x in enumerate(items, start=1):
        st = x.get("status")
        if st == "pending":
            style.add("BACKGROUND", (8, ridx), (8, ridx), colors.HexColor("#FEF3C7"))
        elif st == "approved":
            style.add("BACKGROUND", (8, ridx), (8, ridx), colors.HexColor("#D1FAE5"))
        elif st == "rejected":
            style.add("BACKGROUND", (8, ridx), (8, ridx), colors.HexColor("#FEE2E2"))
    table.setStyle(style)
    story.append(table)

    # Summary
    pending = sum(1 for x in items if x.get("status") == "pending")
    approved = sum(1 for x in items if x.get("status") == "approved")
    rejected = sum(1 for x in items if x.get("status") == "rejected")
    story.append(Spacer(1, 12))
    summary_data = [
        ["RINGKASAN", "", "", ""],
        ["Total Pengajuan", str(len(items)), "Menunggu", str(pending)],
        ["Disetujui", str(approved), "Ditolak", str(rejected)],
    ]
    summary_table = Table(summary_data, colWidths=[40 * mm, 25 * mm, 40 * mm, 25 * mm])
    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("SPAN", (0, 0), (3, 0)),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)

    # Signature
    story.append(Spacer(1, 30))
    sig_data = [
        ["Diketahui,", "", "Dibuat oleh,"],
        ["", "", ""],
        ["", "", ""],
        ["(_______________________)", "", "(_______________________)"],
        ["Direktur", "", "HR Department"],
    ]
    sig_table = Table(sig_data, colWidths=[80 * mm, 80 * mm, 80 * mm])
    sig_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(sig_table)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan-izin-{period}.pdf"'},
    )


STATUS_LABEL_MAP = {"pending": "Menunggu", "approved": "Disetujui", "rejected": "Ditolak"}


# ---------------- User Management (Super Admin only) ----------------
class UserCreateIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str  # "super_admin" or "admin_privileged"
    permissions: Optional[List[str]] = None
    branch: Optional[str] = None  # "plaza" | "kastem" | None


class UserUpdateIn(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    password: Optional[str] = None
    permissions: Optional[List[str]] = None
    branch: Optional[str] = None


def _sanitize_permissions(role: str, perms: Optional[List[str]]) -> List[str]:
    if role == ROLE_SUPER_ADMIN:
        return list(MENU_KEYS)  # super admin implicitly has all
    valid = [p for p in (perms or []) if p in MENU_KEYS]
    # de-dup while preserving order
    seen = set()
    out = []
    for p in valid:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


VALID_BRANCHES = {"plaza", "kastem"}


def _sanitize_branch(val: Optional[str]) -> Optional[str]:
    if not val:
        return None
    v = str(val).strip().lower()
    return v if v in VALID_BRANCHES else None


def _user_view(u: dict) -> dict:
    role = u.get("role")
    if role == "admin":
        role = ROLE_SUPER_ADMIN
    if role == "hr_leave":
        role = ROLE_ADMIN_PRIVILEGED
    perms = u.get("permissions") or []
    if role == ROLE_SUPER_ADMIN:
        perms = list(MENU_KEYS)
    else:
        perms = [p for p in perms if p in MENU_KEYS]
    return {
        "id": u["id"],
        "email": u["email"],
        "name": u.get("name", ""),
        "role": role,
        "permissions": perms,
        "branch": _sanitize_branch(u.get("branch")),
        "created_at": u.get("created_at"),
    }


@api_router.get("/users")
async def list_users(user: dict = Depends(require_super_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(length=500)
    return [_user_view(u) for u in users]


@api_router.post("/users")
async def create_user(payload: UserCreateIn, user: dict = Depends(require_super_admin)):
    if payload.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Role tidak valid. Pilihan: {', '.join(sorted(VALID_ROLES))}")
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
    email = payload.email.lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email sudah terdaftar")
    doc = {
        "id": str(uuid.uuid4()),
        "email": email,
        "name": payload.name.strip(),
        "password_hash": hash_password(payload.password),
        "role": payload.role,
        "permissions": _sanitize_permissions(payload.role, payload.permissions),
        "branch": _sanitize_branch(payload.branch),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    return _user_view(doc)


@api_router.put("/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdateIn, user: dict = Depends(require_super_admin)):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    update: Dict[str, Any] = {}
    if payload.name is not None:
        update["name"] = payload.name.strip()
    new_role = payload.role
    if new_role is not None:
        if new_role not in VALID_ROLES:
            raise HTTPException(status_code=400, detail="Role tidak valid")
        # Prevent demoting the last super_admin
        if target.get("role") in {ROLE_SUPER_ADMIN, "admin"} and new_role != ROLE_SUPER_ADMIN:
            super_admins = await db.users.count_documents({"role": {"$in": [ROLE_SUPER_ADMIN, "admin"]}})
            if super_admins <= 1:
                raise HTTPException(status_code=400, detail="Tidak dapat mengubah role: minimal harus ada 1 Super Admin")
        update["role"] = new_role
    if payload.password:
        if len(payload.password) < 6:
            raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
        update["password_hash"] = hash_password(payload.password)
    # Permissions: always sanitize according to final role
    if payload.permissions is not None or new_role is not None:
        role_effective = new_role or target.get("role") or ROLE_ADMIN_PRIVILEGED
        if role_effective == "admin":
            role_effective = ROLE_SUPER_ADMIN
        if role_effective == "hr_leave":
            role_effective = ROLE_ADMIN_PRIVILEGED
        perms_input = payload.permissions if payload.permissions is not None else (target.get("permissions") or [])
        update["permissions"] = _sanitize_permissions(role_effective, perms_input)
    if payload.branch is not None:
        # Allow explicit empty string / null to clear
        update["branch"] = _sanitize_branch(payload.branch)
    if update:
        await db.users.update_one({"id": user_id}, {"$set": update})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return _user_view(updated)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_super_admin)):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if target["id"] == user["id"]:
        raise HTTPException(status_code=400, detail="Tidak dapat menghapus akun sendiri")
    if target.get("role") in {ROLE_SUPER_ADMIN, "admin"}:
        super_admins = await db.users.count_documents({"role": {"$in": [ROLE_SUPER_ADMIN, "admin"]}})
        if super_admins <= 1:
            raise HTTPException(status_code=400, detail="Tidak dapat menghapus: minimal harus ada 1 Super Admin")
    await db.users.delete_one({"id": user_id})
    return {"ok": True}


# ---------------- Health ----------------
@api_router.get("/")
async def root():
    return {"message": "HRIS API", "ok": True}


# ---------------- Startup ----------------
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.employees.create_index("nik", unique=True)
    await db.payslips.create_index([("period", 1), ("employee_id", 1)])
    await db.payroll_runs.create_index("period", unique=True)
    await db.thr_runs.create_index("period", unique=True)
    await db.thr_slips.create_index([("period", 1), ("employee_id", 1)])
    await db.app_config.create_index("id", unique=True)
    await db.email_logs.create_index("sent_at")
    await db.leave_requests.create_index([("employee_id", 1), ("submitted_at", -1)])
    await db.leave_requests.create_index("status")

    # Load config overrides
    await _load_config_from_db()

    admin_email = os.environ.get("ADMIN_EMAIL", "admin@payroll.id").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin HR",
            "role": ROLE_SUPER_ADMIN,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Seeded super admin user: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}},
        )
        logger.info("Updated admin password from .env")

    # Migrate legacy "admin" role -> "super_admin"
    migrated = await db.users.update_many({"role": "admin"}, {"$set": {"role": ROLE_SUPER_ADMIN}})
    if migrated.modified_count:
        logger.info(f"Migrated {migrated.modified_count} legacy 'admin' role(s) to '{ROLE_SUPER_ADMIN}'")

    # Migrate legacy "hr_leave" role -> "admin_privileged" with izin_cuti permission
    hr_leave_users = await db.users.find({"role": "hr_leave"}, {"_id": 0}).to_list(length=200)
    for u in hr_leave_users:
        perms = list(set((u.get("permissions") or []) + ["izin_cuti"]))
        await db.users.update_one(
            {"id": u["id"]},
            {"$set": {"role": ROLE_ADMIN_PRIVILEGED, "permissions": perms}},
        )
    if hr_leave_users:
        logger.info(f"Migrated {len(hr_leave_users)} 'hr_leave' role(s) to '{ROLE_ADMIN_PRIVILEGED}'")

    # Ensure super_admin users have full permissions array populated (idempotent)
    await db.users.update_many(
        {"role": ROLE_SUPER_ADMIN},
        {"$set": {"permissions": list(MENU_KEYS)}},
    )
    # Ensure admin_privileged users have a permissions field
    await db.users.update_many(
        {"role": ROLE_ADMIN_PRIVILEGED, "permissions": {"$exists": False}},
        {"$set": {"permissions": []}},
    )


@app.on_event("shutdown")
async def shutdown():
    client.close()


# ---------------- Inventory Module ----------------
class MaterialIn(BaseModel):
    name: str
    category: str  # flexy | sticker | tinta | lainnya
    unit: str  # meter | roll | liter | pcs
    current_stock: float = 0
    purchase_price: float = 0  # harga beli per unit
    selling_price: float = 0  # harga jual per m² (untuk POS Sales)
    min_stock: float = 0
    supplier_default: Optional[str] = None
    notes: Optional[str] = None
    active: bool = True


class StockInIn(BaseModel):
    material_id: str
    quantity: float
    unit_price: float
    supplier: Optional[str] = None
    invoice_no: Optional[str] = None
    date: str  # ISO YYYY-MM-DD
    notes: Optional[str] = None


class WasteIn(BaseModel):
    material_id: str
    quantity: float
    reason: str  # rusak | rijek | kadaluarsa | lainnya
    date: str
    reported_by: Optional[str] = None
    notes: Optional[str] = None


MATERIAL_CATEGORIES = ["flexy", "sticker", "tinta", "lainnya"]
MATERIAL_UNITS = ["meter", "roll", "liter", "pcs"]
WASTE_REASONS = ["rusak", "rijek", "kadaluarsa", "lainnya"]


# ----- Materials CRUD -----
@api_router.get("/inventory/materials")
async def inv_list_materials(user: dict = Depends(require_super_admin)):
    cursor = db.materials.find({}, {"_id": 0}).sort("created_at", -1)
    return await cursor.to_list(length=2000)


@api_router.post("/inventory/materials")
async def inv_create_material(payload: MaterialIn, user: dict = Depends(require_super_admin)):
    if not (payload.category or "").strip():
        raise HTTPException(status_code=400, detail="Kategori wajib diisi")
    if payload.unit not in MATERIAL_UNITS:
        raise HTTPException(status_code=400, detail=f"Satuan tidak valid")
    doc = payload.model_dump()
    doc["category"] = (doc.get("category") or "").strip()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["updated_at"] = doc["created_at"]
    await db.materials.insert_one(doc)
    await _upsert_category("material", doc["category"])
    doc.pop("_id", None)
    return doc


@api_router.put("/inventory/materials/{material_id}")
async def inv_update_material(material_id: str, payload: MaterialIn, user: dict = Depends(require_super_admin)):
    if not (payload.category or "").strip():
        raise HTTPException(status_code=400, detail="Kategori wajib diisi")
    if payload.unit not in MATERIAL_UNITS:
        raise HTTPException(status_code=400, detail=f"Satuan tidak valid")
    existing = await db.materials.find_one({"id": material_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bahan tidak ditemukan")
    update = payload.model_dump()
    update["category"] = (update.get("category") or "").strip()
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.materials.update_one({"id": material_id}, {"$set": update})
    await _upsert_category("material", update["category"])
    return await db.materials.find_one({"id": material_id}, {"_id": 0})


@api_router.delete("/inventory/materials/{material_id}")
async def inv_delete_material(material_id: str, user: dict = Depends(require_super_admin)):
    existing = await db.materials.find_one({"id": material_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bahan tidak ditemukan")
    has_in = await db.stock_in.find_one({"material_id": material_id})
    has_waste = await db.waste.find_one({"material_id": material_id})
    if has_in or has_waste:
        await db.materials.update_one({"id": material_id}, {"$set": {"active": False, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"ok": True, "soft_deleted": True}
    await db.materials.delete_one({"id": material_id})
    return {"ok": True, "soft_deleted": False}


# ----- Stock In (Barang Masuk) -----
async def _enrich_with_material(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    mats = {m["id"]: m async for m in db.materials.find({}, {"_id": 0})}
    for it in items:
        mat = mats.get(it.get("material_id")) or {}
        it["material_name"] = mat.get("name") or "-"
        it["material_unit"] = mat.get("unit") or ""
        it["material_category"] = mat.get("category") or ""
    return items


@api_router.get("/inventory/stock-in")
async def inv_list_stock_in(user: dict = Depends(require_super_admin), material_id: Optional[str] = None):
    q = {}
    if material_id:
        q["material_id"] = material_id
    cursor = db.stock_in.find(q, {"_id": 0}).sort("date", -1)
    items = await cursor.to_list(length=2000)
    return await _enrich_with_material(items)


@api_router.post("/inventory/stock-in")
async def inv_create_stock_in(payload: StockInIn, user: dict = Depends(require_super_admin)):
    mat = await db.materials.find_one({"id": payload.material_id})
    if not mat:
        raise HTTPException(status_code=404, detail="Bahan tidak ditemukan")
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Kuantitas harus > 0")
    total = round(float(payload.quantity) * float(payload.unit_price), 2)
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["total_price"] = total
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user.get("email")
    await db.stock_in.insert_one(doc)
    new_stock = round(float(mat.get("current_stock", 0)) + float(payload.quantity), 4)
    await db.materials.update_one(
        {"id": payload.material_id},
        {"$set": {
            "current_stock": new_stock,
            "purchase_price": float(payload.unit_price),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    doc.pop("_id", None)
    doc["new_stock"] = new_stock
    return doc


@api_router.delete("/inventory/stock-in/{stock_in_id}")
async def inv_delete_stock_in(stock_in_id: str, user: dict = Depends(require_super_admin)):
    doc = await db.stock_in.find_one({"id": stock_in_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Data barang masuk tidak ditemukan")
    mat = await db.materials.find_one({"id": doc["material_id"]})
    if mat:
        new_stock = round(float(mat.get("current_stock", 0)) - float(doc.get("quantity", 0)), 4)
        await db.materials.update_one(
            {"id": doc["material_id"]},
            {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}},
        )
    await db.stock_in.delete_one({"id": stock_in_id})
    return {"ok": True}


# ----- Waste (Sisa/Rijek) -----
@api_router.get("/inventory/waste")
async def inv_list_waste(user: dict = Depends(require_super_admin), material_id: Optional[str] = None):
    q = {}
    if material_id:
        q["material_id"] = material_id
    cursor = db.waste.find(q, {"_id": 0}).sort("date", -1)
    items = await cursor.to_list(length=2000)
    return await _enrich_with_material(items)


@api_router.post("/inventory/waste")
async def inv_create_waste(payload: WasteIn, user: dict = Depends(require_super_admin)):
    mat = await db.materials.find_one({"id": payload.material_id})
    if not mat:
        raise HTTPException(status_code=404, detail="Bahan tidak ditemukan")
    if payload.quantity <= 0:
        raise HTTPException(status_code=400, detail="Kuantitas harus > 0")
    if payload.reason not in WASTE_REASONS:
        raise HTTPException(status_code=400, detail=f"Alasan tidak valid")
    unit_price = float(mat.get("purchase_price", 0))
    estimated_loss = round(float(payload.quantity) * unit_price, 2)
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["unit_price_snapshot"] = unit_price
    doc["estimated_loss"] = estimated_loss
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user.get("email")
    await db.waste.insert_one(doc)
    new_stock = round(float(mat.get("current_stock", 0)) - float(payload.quantity), 4)
    await db.materials.update_one(
        {"id": payload.material_id},
        {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    doc.pop("_id", None)
    doc["new_stock"] = new_stock
    return doc


@api_router.delete("/inventory/waste/{waste_id}")
async def inv_delete_waste(waste_id: str, user: dict = Depends(require_super_admin)):
    doc = await db.waste.find_one({"id": waste_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Data waste tidak ditemukan")
    mat = await db.materials.find_one({"id": doc["material_id"]})
    if mat:
        new_stock = round(float(mat.get("current_stock", 0)) + float(doc.get("quantity", 0)), 4)
        await db.materials.update_one(
            {"id": doc["material_id"]},
            {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}},
        )
    await db.waste.delete_one({"id": waste_id})
    return {"ok": True}


# ----- Inventory Stats -----
@api_router.get("/inventory/stats")
async def inv_stats(user: dict = Depends(require_super_admin)):
    materials = await db.materials.find({}, {"_id": 0}).to_list(length=5000)
    total_stock_value = 0.0
    low_stock = []
    for m in materials:
        stock = float(m.get("current_stock", 0))
        price = float(m.get("purchase_price", 0))
        total_stock_value += stock * price
        min_stock = float(m.get("min_stock", 0))
        if min_stock > 0 and stock <= min_stock and m.get("active", True):
            low_stock.append({
                "id": m["id"], "name": m["name"], "category": m.get("category"),
                "current_stock": stock, "min_stock": min_stock, "unit": m.get("unit"),
            })
    today = datetime.now(timezone.utc).date()
    month_start = today.replace(day=1).isoformat()
    waste_month = await db.waste.find({"date": {"$gte": month_start}}, {"_id": 0}).to_list(length=5000)
    total_waste_this_month = sum(float(w.get("estimated_loss", 0)) for w in waste_month)
    # Top waste bulan ini per material
    top_agg: Dict[str, Dict[str, Any]] = {}
    mat_by_id = {m["id"]: m for m in materials}
    for w in waste_month:
        mid = w.get("material_id")
        if not mid:
            continue
        mat = mat_by_id.get(mid, {})
        row = top_agg.setdefault(mid, {
            "material_id": mid, "material_name": mat.get("name") or "-",
            "material_unit": mat.get("unit") or "", "material_category": mat.get("category") or "",
            "qty": 0.0, "loss": 0.0, "records": 0,
        })
        row["qty"] += float(w.get("quantity", 0))
        row["loss"] += float(w.get("estimated_loss", 0))
        row["records"] += 1
    top_waste = sorted(top_agg.values(), key=lambda r: r["loss"], reverse=True)[:5]
    for r in top_waste:
        r["qty"] = round(r["qty"], 4)
        r["loss"] = round(r["loss"], 2)
    return {
        "total_materials": sum(1 for m in materials if m.get("active", True)),
        "total_stock_value": round(total_stock_value, 2),
        "low_stock_count": len(low_stock),
        "low_stock": low_stock[:10],
        "total_waste_this_month": round(total_waste_this_month, 2),
        "waste_records_this_month": len(waste_month),
        "top_waste": top_waste,
    }


# ---------------- Stock Adjustment (Opname) ----------------
class StockAdjustIn(BaseModel):
    material_id: str
    new_stock: float  # nilai stok setelah opname
    reason: str = "opname"  # opname | koreksi | lainnya
    date: str
    notes: Optional[str] = None


@api_router.get("/inventory/stock-adjust")
async def inv_list_stock_adjust(user: dict = Depends(require_super_admin), material_id: Optional[str] = None):
    q = {}
    if material_id:
        q["material_id"] = material_id
    items = await db.stock_adjust.find(q, {"_id": 0}).sort("date", -1).to_list(length=2000)
    return await _enrich_with_material(items)


@api_router.post("/inventory/stock-adjust")
async def inv_create_stock_adjust(payload: StockAdjustIn, user: dict = Depends(require_super_admin)):
    mat = await db.materials.find_one({"id": payload.material_id})
    if not mat:
        raise HTTPException(status_code=404, detail="Bahan tidak ditemukan")
    stock_before = float(mat.get("current_stock", 0))
    new_stock = float(payload.new_stock)
    delta = round(new_stock - stock_before, 4)
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["stock_before"] = round(stock_before, 4)
    doc["stock_after"] = round(new_stock, 4)
    doc["delta"] = delta
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user.get("email")
    await db.stock_adjust.insert_one(doc)
    await db.materials.update_one(
        {"id": payload.material_id},
        {"$set": {"current_stock": round(new_stock, 4), "updated_at": doc["created_at"]}},
    )
    doc.pop("_id", None)
    return doc


@api_router.delete("/inventory/stock-adjust/{adj_id}")
async def inv_delete_stock_adjust(adj_id: str, user: dict = Depends(require_super_admin)):
    doc = await db.stock_adjust.find_one({"id": adj_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Data opname tidak ditemukan")
    # Kembalikan stok ke nilai sebelum adjustment
    await db.materials.update_one(
        {"id": doc["material_id"]},
        {"$set": {"current_stock": float(doc.get("stock_before", 0)), "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    await db.stock_adjust.delete_one({"id": adj_id})
    return {"ok": True}


# ---------------- Job Order (Produksi) ----------------
class JobItemIn(BaseModel):
    material_id: str
    quantity: float  # qty bahan dikonsumsi total


class JobOrderIn(BaseModel):
    order_no: Optional[str] = None  # auto-gen bila kosong
    customer: str
    product_name: str
    quantity: int = 1
    unit_price: float = 0
    start_date: str  # ISO
    due_date: Optional[str] = None
    items: List[JobItemIn] = []
    notes: Optional[str] = None


async def _next_order_no() -> str:
    today = datetime.now(timezone.utc).date()
    prefix = f"JO-{today.strftime('%Y%m')}-"
    count = await db.job_orders.count_documents({"order_no": {"$regex": f"^{prefix}"}})
    return f"{prefix}{count + 1:04d}"


async def _reverse_job_stock(job: Dict[str, Any]) -> None:
    """Kembalikan stok yg pernah dikurangi oleh job (dipakai saat cancel/delete)."""
    for it in job.get("items") or []:
        mid = it.get("material_id")
        qty = float(it.get("quantity", 0))
        if not mid or qty <= 0:
            continue
        mat = await db.materials.find_one({"id": mid})
        if mat:
            new_stock = round(float(mat.get("current_stock", 0)) + qty, 4)
            await db.materials.update_one(
                {"id": mid},
                {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )


@api_router.get("/inventory/orders")
async def inv_list_orders(user: dict = Depends(require_super_admin), status: Optional[str] = None):
    q = {}
    if status:
        q["status"] = status
    items = await db.job_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(length=2000)
    # Enrich items dgn nama material
    all_mats = {m["id"]: m async for m in db.materials.find({}, {"_id": 0})}
    for job in items:
        for it in job.get("items") or []:
            mat = all_mats.get(it.get("material_id")) or {}
            it["material_name"] = mat.get("name") or "-"
            it["material_unit"] = mat.get("unit") or ""
    return items


@api_router.post("/inventory/orders")
async def inv_create_order(payload: JobOrderIn, user: dict = Depends(require_super_admin)):
    if not payload.customer or not payload.product_name:
        raise HTTPException(status_code=400, detail="Customer & nama produk wajib diisi")
    # Validate materials + hitung total kerugian bahan snapshot
    items_out = []
    total_material_cost = 0.0
    for it in payload.items:
        mat = await db.materials.find_one({"id": it.material_id})
        if not mat:
            raise HTTPException(status_code=400, detail=f"Bahan {it.material_id} tidak ditemukan")
        if it.quantity <= 0:
            raise HTTPException(status_code=400, detail=f"Kuantitas bahan {mat.get('name')} harus > 0")
        stock = float(mat.get("current_stock", 0))
        if it.quantity > stock:
            raise HTTPException(status_code=400, detail=f"Stok {mat.get('name')} tidak cukup (ada {stock}, butuh {it.quantity})")
        unit_price = float(mat.get("purchase_price", 0))
        items_out.append({
            "material_id": it.material_id,
            "quantity": float(it.quantity),
            "unit_price_snapshot": unit_price,
            "cost": round(float(it.quantity) * unit_price, 2),
        })
        total_material_cost += float(it.quantity) * unit_price

    order_no = payload.order_no or await _next_order_no()
    doc = payload.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["order_no"] = order_no
    doc["items"] = items_out
    doc["status"] = "aktif"
    doc["total_material_cost"] = round(total_material_cost, 2)
    doc["total_price"] = round(float(payload.quantity) * float(payload.unit_price), 2)
    doc["gross_margin"] = round(doc["total_price"] - doc["total_material_cost"], 2)
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["created_by"] = user.get("email")
    await db.job_orders.insert_one(doc)

    # Kurangi stok tiap bahan
    for it in items_out:
        mat = await db.materials.find_one({"id": it["material_id"]})
        if mat:
            new_stock = round(float(mat.get("current_stock", 0)) - float(it["quantity"]), 4)
            await db.materials.update_one(
                {"id": it["material_id"]},
                {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}},
            )
    doc.pop("_id", None)
    return doc


@api_router.put("/inventory/orders/{order_id}/complete")
async def inv_complete_order(order_id: str, user: dict = Depends(require_super_admin)):
    job = await db.job_orders.find_one({"id": order_id})
    if not job:
        raise HTTPException(status_code=404, detail="Order tidak ditemukan")
    if job.get("status") == "batal":
        raise HTTPException(status_code=400, detail="Order sudah dibatalkan")
    await db.job_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "selesai", "completed_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"ok": True, "status": "selesai"}


@api_router.put("/inventory/orders/{order_id}/cancel")
async def inv_cancel_order(order_id: str, user: dict = Depends(require_super_admin)):
    job = await db.job_orders.find_one({"id": order_id})
    if not job:
        raise HTTPException(status_code=404, detail="Order tidak ditemukan")
    if job.get("status") == "batal":
        return {"ok": True, "status": "batal"}
    # Rollback stok jika sebelumnya aktif (belum pernah dibatalkan)
    if job.get("status") in ("aktif", "selesai"):
        await _reverse_job_stock(job)
    await db.job_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "batal", "cancelled_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"ok": True, "status": "batal"}


@api_router.delete("/inventory/orders/{order_id}")
async def inv_delete_order(order_id: str, user: dict = Depends(require_super_admin)):
    job = await db.job_orders.find_one({"id": order_id})
    if not job:
        raise HTTPException(status_code=404, detail="Order tidak ditemukan")
    # Rollback bila order aktif/selesai (yg pernah kurangi stok)
    if job.get("status") in ("aktif", "selesai"):
        await _reverse_job_stock(job)
    await db.job_orders.delete_one({"id": order_id})
    return {"ok": True}


# ---------------- Waste Report (Excel + PDF) ----------------
async def _fetch_monthly_waste(period: str):
    start, end, year, month = _parse_month(period)
    cursor = db.waste.find({"date": {"$gte": start, "$lte": end}}, {"_id": 0}).sort("date", 1)
    items = await cursor.to_list(length=5000)
    items = await _enrich_with_material(items)
    total_loss = sum(float(x.get("estimated_loss", 0)) for x in items)
    total_qty = sum(float(x.get("quantity", 0)) for x in items)
    return items, total_loss, total_qty, year, month


@api_router.get("/inventory/waste/report/{period}/excel")
async def inv_waste_report_excel(period: str, user: dict = Depends(require_super_admin)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    items, total_loss, total_qty, year, month = await _fetch_monthly_waste(period)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Waste {period}"
    ws.append([f"LAPORAN WASTE / RIJEK BULANAN — {year}-{month:02d}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Tanggal", "Bahan", "Kategori", "Alasan", "Qty", "Satuan", "Harga/Unit", "Kerugian (Rp)", "Pelapor", "Catatan"]
    ws.append(headers)
    header_fill = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for it in items:
        ws.append([
            it.get("date", ""),
            it.get("material_name", ""),
            it.get("material_category", ""),
            it.get("reason", ""),
            float(it.get("quantity", 0)),
            it.get("material_unit", ""),
            float(it.get("unit_price_snapshot", 0)),
            float(it.get("estimated_loss", 0)),
            it.get("reported_by", ""),
            it.get("notes", ""),
        ])
    ws.append([])
    total_row = ["TOTAL", "", "", "", total_qty, "", "", total_loss, "", ""]
    ws.append(total_row)
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)
    for col_idx, width in enumerate([12, 30, 12, 14, 10, 10, 14, 16, 20, 30], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="laporan-waste-{period}.xlsx"'},
    )


@api_router.get("/inventory/waste/report/{period}/pdf")
async def inv_waste_report_pdf(period: str, user: dict = Depends(require_super_admin)):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    items, total_loss, total_qty, year, month = await _fetch_monthly_waste(period)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elems = [
        Paragraph(f"<b>LAPORAN WASTE / RIJEK BULANAN</b>", styles["Title"]),
        Paragraph(f"Periode: {year}-{month:02d}", styles["Normal"]),
        Spacer(1, 8),
    ]
    data = [["Tanggal", "Bahan", "Alasan", "Qty", "Satuan", "Kerugian (Rp)", "Pelapor"]]
    for it in items:
        data.append([
            it.get("date", ""),
            it.get("material_name", ""),
            it.get("reason", ""),
            f"{float(it.get('quantity', 0)):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
            it.get("material_unit", ""),
            f"Rp {float(it.get('estimated_loss', 0)):,.0f}".replace(",", "."),
            it.get("reported_by", "") or "-",
        ])
    data.append(["", "", "TOTAL", f"{total_qty:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), "", f"Rp {total_loss:,.0f}".replace(",", "."), ""])
    tbl = Table(data, colWidths=[22 * mm, 60 * mm, 25 * mm, 22 * mm, 20 * mm, 40 * mm, 40 * mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (3, 1), (5, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f5f5f5")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems.append(tbl)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laporan-waste-{period}.pdf"'},
    )


# ---------------- Customer Master ----------------
class CustomerIn(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    npwp: Optional[str] = None
    contact_person: Optional[str] = None
    category: Optional[str] = None
    notes: Optional[str] = None
    active: bool = True


# ================================================================
# ==================== MASTER KATEGORI (unified) ==================
# ================================================================
CATEGORY_TYPES = ("material", "product", "supplier", "customer")

class CategoryIn(BaseModel):
    type: str  # material | product | supplier | customer
    name: str
    description: Optional[str] = None
    color: Optional[str] = None  # hex mis. #002FA7
    active: bool = True


async def _upsert_category(cat_type: str, name: Optional[str]):
    """Idempotent upsert kategori. Dipanggil dari CRUD master lain."""
    if not name or not name.strip() or cat_type not in CATEGORY_TYPES:
        return
    nm = name.strip()
    # Case-insensitive dedupe
    exists = await db.categories.find_one({"type": cat_type, "name": {"$regex": f"^{re.escape(nm)}$", "$options": "i"}})
    if exists:
        return
    await db.categories.insert_one({
        "id": str(uuid.uuid4()),
        "type": cat_type,
        "name": nm,
        "description": None,
        "color": None,
        "active": True,
        "auto_created": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })


async def _backfill_categories():
    """One-shot backfill: scan existing masters lalu upsert kategori-kategori yang belum ada."""
    for src_coll, cat_type in [
        ("materials", "material"),
        ("products", "product"),
        ("suppliers", "supplier"),
        ("customers", "customer"),
    ]:
        try:
            names = await db[src_coll].distinct("category")
            for n in names or []:
                if n and isinstance(n, str) and n.strip():
                    await _upsert_category(cat_type, n.strip())
        except Exception as ex:
            logger.warning(f"Backfill category from {src_coll} failed: {ex}")


@api_router.get("/categories")
async def categories_list(user: dict = Depends(require_super_admin), type: Optional[str] = None, only_active: bool = False):
    q: Dict[str, Any] = {}
    if type:
        if type not in CATEGORY_TYPES:
            raise HTTPException(status_code=400, detail=f"Type harus salah satu: {CATEGORY_TYPES}")
        q["type"] = type
    if only_active:
        q["active"] = {"$ne": False}
    items = await db.categories.find(q, {"_id": 0}).sort([("type", 1), ("name", 1)]).to_list(length=5000)
    return items


@api_router.post("/categories/backfill")
async def categories_backfill(user: dict = Depends(require_super_admin)):
    before = await db.categories.count_documents({})
    await _backfill_categories()
    after = await db.categories.count_documents({})
    return {"added": after - before, "total": after}


@api_router.post("/categories")
async def categories_create(payload: CategoryIn, user: dict = Depends(require_super_admin)):
    if payload.type not in CATEGORY_TYPES:
        raise HTTPException(status_code=400, detail=f"Type harus salah satu: {CATEGORY_TYPES}")
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama kategori wajib")
    exists = await db.categories.find_one({"type": payload.type, "name": {"$regex": f"^{re.escape(payload.name.strip())}$", "$options": "i"}})
    if exists:
        raise HTTPException(status_code=400, detail=f"Kategori '{payload.name}' sudah ada untuk tipe {payload.type}")
    doc = {
        "id": str(uuid.uuid4()),
        "type": payload.type,
        "name": payload.name.strip(),
        "description": (payload.description or "").strip() or None,
        "color": (payload.color or "").strip() or None,
        "active": payload.active,
        "auto_created": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.categories.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/categories/{cat_id}")
async def categories_update(cat_id: str, payload: CategoryIn, user: dict = Depends(require_super_admin)):
    existing = await db.categories.find_one({"id": cat_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    new_name = payload.name.strip()
    if payload.type not in CATEGORY_TYPES:
        raise HTTPException(status_code=400, detail=f"Type harus salah satu: {CATEGORY_TYPES}")
    if not new_name:
        raise HTTPException(status_code=400, detail="Nama kategori wajib")
    if new_name.lower() != (existing.get("name") or "").lower() or payload.type != existing.get("type"):
        dup = await db.categories.find_one({
            "type": payload.type,
            "name": {"$regex": f"^{re.escape(new_name)}$", "$options": "i"},
            "id": {"$ne": cat_id},
        })
        if dup:
            raise HTTPException(status_code=400, detail=f"Kategori '{new_name}' sudah ada untuk tipe {payload.type}")
    # Kalau rename, cascade update di collections terkait
    old_name = existing.get("name") or ""
    old_type = existing.get("type")
    upd = {
        "type": payload.type,
        "name": new_name,
        "description": (payload.description or "").strip() or None,
        "color": (payload.color or "").strip() or None,
        "active": payload.active,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.categories.update_one({"id": cat_id}, {"$set": upd})
    # Cascade rename di master data terkait
    if old_name and old_name != new_name and old_type == payload.type:
        coll_map = {"material": "materials", "product": "products", "supplier": "suppliers", "customer": "customers"}
        coll = coll_map.get(payload.type)
        if coll:
            await db[coll].update_many({"category": old_name}, {"$set": {"category": new_name}})
    return await db.categories.find_one({"id": cat_id}, {"_id": 0})


@api_router.delete("/categories/{cat_id}")
async def categories_delete(cat_id: str, user: dict = Depends(require_super_admin)):
    existing = await db.categories.find_one({"id": cat_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Kategori tidak ditemukan")
    # Cek apakah masih dipakai
    coll_map = {"material": "materials", "product": "products", "supplier": "suppliers", "customer": "customers"}
    coll = coll_map.get(existing.get("type"))
    if coll:
        used = await db[coll].count_documents({"category": existing.get("name")})
        if used > 0:
            raise HTTPException(status_code=400, detail=f"Kategori masih dipakai di {used} data {existing.get('type')}. Non-aktifkan saja atau ubah data yang menggunakan.")
    await db.categories.delete_one({"id": cat_id})
    return {"ok": True}


@api_router.get("/categories/stats")
async def categories_stats(user: dict = Depends(require_super_admin)):
    """Return count per type + usage per category."""
    coll_map = {"material": "materials", "product": "products", "supplier": "suppliers", "customer": "customers"}
    stats = {}
    for t in CATEGORY_TYPES:
        total = await db.categories.count_documents({"type": t})
        active = await db.categories.count_documents({"type": t, "active": {"$ne": False}})
        stats[t] = {"total": total, "active": active}
    return stats


@api_router.get("/inventory/customers")
async def cust_list(user: dict = Depends(require_super_admin)):
    items = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(length=5000)
    # Enrich dengan agregat order per customer
    orders = await db.job_orders.find({}, {"_id": 0}).to_list(length=10000)
    by_name: Dict[str, Dict[str, Any]] = {}
    for o in orders:
        key = (o.get("customer") or "").strip().lower()
        if not key:
            continue
        row = by_name.setdefault(key, {"count": 0, "revenue": 0.0, "material_cost": 0.0})
        if o.get("status") != "batal":
            row["count"] += 1
            row["revenue"] += float(o.get("total_price", 0))
            row["material_cost"] += float(o.get("total_material_cost", 0))
    for c in items:
        agg = by_name.get(c.get("name", "").strip().lower()) or {}
        c["order_count"] = agg.get("count", 0)
        c["total_revenue"] = round(agg.get("revenue", 0.0), 2)
        c["total_material_cost"] = round(agg.get("material_cost", 0.0), 2)
    return items


@api_router.post("/inventory/customers")
async def cust_create(payload: CustomerIn, user: dict = Depends(require_super_admin)):
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama customer wajib diisi")
    # Cek duplicate name (case-insensitive) — escape regex special chars
    safe_name = re.escape(payload.name.strip())
    existing = await db.customers.find_one({"name": {"$regex": f"^{safe_name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Customer dengan nama tersebut sudah ada")
    doc = payload.model_dump()
    doc["name"] = payload.name.strip()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.insert_one(doc)
    await _upsert_category("customer", doc.get("category"))
    doc.pop("_id", None)
    return doc


@api_router.put("/inventory/customers/{customer_id}")
async def cust_update(customer_id: str, payload: CustomerIn, user: dict = Depends(require_super_admin)):
    existing = await db.customers.find_one({"id": customer_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer tidak ditemukan")
    upd = payload.model_dump()
    upd["name"] = payload.name.strip()
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.update_one({"id": customer_id}, {"$set": upd})
    await _upsert_category("customer", upd.get("category"))
    return await db.customers.find_one({"id": customer_id}, {"_id": 0})


@api_router.delete("/inventory/customers/{customer_id}")
async def cust_delete(customer_id: str, user: dict = Depends(require_super_admin)):
    existing = await db.customers.find_one({"id": customer_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer tidak ditemukan")
    await db.customers.delete_one({"id": customer_id})
    return {"ok": True}


# ---------------- Broadcast WhatsApp ke Pelanggan ----------------
class CustomerBroadcastIn(BaseModel):
    message: str
    customer_ids: Optional[List[str]] = None  # None/empty = semua pelanggan aktif yg punya phone
    preview_only: bool = False


@api_router.post("/inventory/customers/broadcast-whatsapp")
async def customers_broadcast_whatsapp(payload: CustomerBroadcastIn, user: dict = Depends(require_super_admin)):
    msg_template = (payload.message or "").strip()
    if not msg_template:
        raise HTTPException(status_code=400, detail="Pesan tidak boleh kosong")
    if len(msg_template) > 3000:
        raise HTTPException(status_code=400, detail="Pesan terlalu panjang (max 3000 karakter)")

    # Ambil target customers
    query: Dict[str, Any] = {"active": {"$ne": False}}
    if payload.customer_ids:
        query["id"] = {"$in": payload.customer_ids}
    all_customers = await db.customers.find(query, {"_id": 0}).to_list(length=5000)
    # Filter yang punya phone
    targets = [c for c in all_customers if (c.get("phone") or "").strip()]
    skipped_no_phone = len(all_customers) - len(targets)

    if payload.preview_only:
        return {
            "preview_only": True,
            "total_selected": len(all_customers),
            "total_with_phone": len(targets),
            "skipped_no_phone": skipped_no_phone,
            "sample_targets": [{"name": c["name"], "phone": c.get("phone")} for c in targets[:5]],
        }

    if not targets:
        raise HTTPException(status_code=400, detail="Tidak ada pelanggan dengan nomor WhatsApp yang bisa dikirim")

    results = []
    sent = failed = mocked = 0
    for c in targets:
        # Replace variabel {name} & {phone}
        personalized = msg_template.replace("{name}", c.get("name", "")).replace("{phone}", c.get("phone", ""))
        res = await _send_whatsapp(c.get("phone", ""), personalized)
        status = res.get("status", "failed")
        if status == "sent":
            sent += 1
        elif status == "mocked":
            mocked += 1
        else:
            failed += 1
        results.append({
            "customer_id": c.get("id"),
            "name": c.get("name"),
            "phone": res.get("phone") or c.get("phone"),
            "status": status,
            "reason": res.get("reason"),
        })
        # Log ke db (opsional, best-effort)
        try:
            await db.whatsapp_logs.insert_one({
                "id": str(uuid.uuid4()),
                "type": "customer_broadcast",
                "customer_id": c.get("id"),
                "customer_name": c.get("name"),
                "phone": res.get("phone") or c.get("phone"),
                "message_preview": personalized[:200],
                "status": status,
                "reason": res.get("reason"),
                "sent_at": datetime.now(timezone.utc).isoformat(),
                "sent_by": user.get("email"),
            })
        except Exception:
            pass
        # Gentle pacing untuk Fonnte free plan
        await asyncio.sleep(0.3)

    return {
        "preview_only": False,
        "total": len(targets),
        "sent": sent,
        "failed": failed,
        "mocked": mocked,
        "skipped_no_phone": skipped_no_phone,
        "results": results,
    }


# ---------------- Laporan Laba/Rugi (Profit & Loss) ----------------
async def _payroll_cost_for_month(period: str) -> tuple:
    """Return (total_net, employee_count) untuk payroll_runs dgn period=YYYY-MM. Fallback 0."""
    run = await db.payroll_runs.find_one({"period": period}, {"_id": 0})
    if not run:
        return 0.0, 0
    return float(run.get("total_net", 0)), int(run.get("employee_count", 0))


@api_router.get("/reports/product-margin/{period}")
async def product_margin_report(period: str, user: dict = Depends(require_super_admin)):
    """Ranking margin per produk untuk periode YYYY-MM.
    Data source: db.sales items dengan status ∈ ["paid","dp"].
    Cost = sum(component.consumption × material.purchase_price).
    """
    start, end, _year, _month = _parse_month(period)
    sales = await db.sales.find(
        {"date": {"$gte": start, "$lte": end}, "status": {"$in": ["paid", "dp"]}}, {"_id": 0},
    ).to_list(length=50000)

    # Cache materials untuk lookup purchase_price
    material_cache: Dict[str, Dict[str, Any]] = {}
    async def _mat(mid: str):
        if mid in material_cache:
            return material_cache[mid]
        m = await db.materials.find_one({"id": mid}, {"_id": 0, "name": 1, "purchase_price": 1, "unit": 1})
        material_cache[mid] = m or {}
        return material_cache[mid]

    # Group by product identity
    # Key: (product_id or "") + "::" + product_name (untuk group produk yg sama)
    groups: Dict[str, Dict[str, Any]] = {}
    for s in sales:
        for it in (s.get("items") or []):
            key = f"{it.get('product_id') or ''}::{(it.get('product_name') or '-').strip()}"
            row = groups.setdefault(key, {
                "product_id": it.get("product_id"),
                "product_name": (it.get("product_name") or "-").strip() or "-",
                "is_bom": bool(it.get("product_id")),
                "sale_count": 0,
                "qty_total": 0,
                "revenue": 0.0,
                "cost": 0.0,
                "materials_used": {},  # name -> total consumption
            })
            row["sale_count"] += 1
            row["qty_total"] += int(it.get("quantity", 0) or 0)
            row["revenue"] += float(it.get("subtotal", 0) or 0)
            # Cost from components (BOM) or legacy area_total × purchase_price
            comps = it.get("components") or []
            if comps:
                for c in comps:
                    m = await _mat(c.get("material_id"))
                    cost = float(c.get("consumption", 0) or 0) * float(m.get("purchase_price", 0) or 0)
                    row["cost"] += cost
                    mname = m.get("name") or c.get("material_name") or "-"
                    mu = row["materials_used"].setdefault(mname, {"consumption": 0.0, "unit": m.get("unit") or c.get("material_unit") or ""})
                    mu["consumption"] += float(c.get("consumption", 0) or 0)
            else:
                # Legacy: material_id + area_total
                mid = it.get("material_id")
                if mid:
                    m = await _mat(mid)
                    area_total = float(it.get("area_total", 0) or 0)
                    row["cost"] += area_total * float(m.get("purchase_price", 0) or 0)
                    mname = m.get("name") or it.get("material_name") or "-"
                    mu = row["materials_used"].setdefault(mname, {"consumption": 0.0, "unit": m.get("unit") or it.get("material_unit") or ""})
                    mu["consumption"] += area_total

    rows = []
    for _k, r in groups.items():
        margin = r["revenue"] - r["cost"]
        pct = (margin / r["revenue"] * 100) if r["revenue"] > 0 else 0.0
        rows.append({
            "product_id": r["product_id"],
            "product_name": r["product_name"],
            "is_bom": r["is_bom"],
            "sale_count": r["sale_count"],
            "qty_total": r["qty_total"],
            "revenue": round(r["revenue"], 2),
            "cost": round(r["cost"], 2),
            "margin": round(margin, 2),
            "margin_pct": round(pct, 2),
            "materials_used": [
                {"name": k, "consumption": round(v["consumption"], 4), "unit": v["unit"]}
                for k, v in sorted(r["materials_used"].items(), key=lambda x: -x[1]["consumption"])
            ],
        })
    # Sort by revenue desc default (frontend bisa re-sort)
    rows.sort(key=lambda x: x["revenue"], reverse=True)
    total_revenue = sum(r["revenue"] for r in rows)
    total_cost = sum(r["cost"] for r in rows)
    total_margin = total_revenue - total_cost
    return {
        "period": period,
        "period_start": start,
        "period_end": end,
        "total_products": len(rows),
        "total_revenue": round(total_revenue, 2),
        "total_cost": round(total_cost, 2),
        "total_margin": round(total_margin, 2),
        "total_margin_pct": round((total_margin / total_revenue * 100) if total_revenue > 0 else 0, 2),
        "products": rows,
    }


@api_router.get("/reports/profit-loss-latest-period")
async def profit_loss_latest_period(user: dict = Depends(require_super_admin)):
    """Return bulan (YYYY-MM) TERBARU yang punya data P&L (sales paid/dp, waste, atau payroll).

    Dipakai oleh halaman Laporan Laba/Rugi untuk set default period ke bulan yg realistis
    (bukan bulan kalender kosong). Fallback ke bulan sekarang jika DB kosong total.
    """
    today = datetime.now(timezone.utc).date()
    fallback = f"{today.year:04d}-{today.month:02d}"

    # Sumber data terbaru: max sales.date (paid/dp), max waste.date, max payroll_runs.period
    latest_dates: List[str] = []
    sale = await db.sales.find_one(
        {"status": {"$in": ["paid", "dp"]}, "date": {"$exists": True, "$ne": None}},
        {"_id": 0, "date": 1},
        sort=[("date", -1)],
    )
    if sale and sale.get("date"):
        latest_dates.append(str(sale["date"])[:7])
    waste_doc = await db.waste.find_one({"date": {"$exists": True, "$ne": None}}, {"_id": 0, "date": 1}, sort=[("date", -1)])
    if waste_doc and waste_doc.get("date"):
        latest_dates.append(str(waste_doc["date"])[:7])
    payroll_run = await db.payroll_runs.find_one({"period": {"$exists": True, "$ne": None}}, {"_id": 0, "period": 1}, sort=[("period", -1)])
    if payroll_run and payroll_run.get("period"):
        latest_dates.append(str(payroll_run["period"])[:7])

    period = max(latest_dates) if latest_dates else fallback
    return {"period": period, "fallback_used": not bool(latest_dates)}


@api_router.get("/reports/profit-loss/{period}")
async def profit_loss_report(period: str, user: dict = Depends(require_super_admin)):
    """P&L bulanan — Struktur baru (permintaan user 2026-08-14):
      PENDAPATAN: dari db.sales (status ∈ paid/dp), Σ items[].subtotal.
      BEBAN ADMINISTRASI & UMUM: dari db.cash_transactions (type=out) per account_code.
        a. Gaji (505)
        b. ATK, Fotocopy, Dll (104)
        c. Telephone, Listrik & Air (502)
        d. Keperluan Kantor (104)
        e. Jasa handling barang (507 + 106)
        f. Penyusutan GA (513)
        g. Perbaikan & Perawatan Kendaraan (105)
        h. Operasional Kendaraan (501)
        i. Administrasi Bank (511)
        j. Pajak (514)
        k. Perbaikan Mesin (402)
        l. Pembelian Bahan Baku Mesin (103-01)
      Laba Bersih = Pendapatan − Total Beban A&U.
      Catatan: Akun 401 (Pembelian Bahan Baku umum) SENGAJA tidak muncul di laporan.
    """
    start, end, year, month = _parse_month(period)

    # --- PENDAPATAN dari sales ---
    sales = await db.sales.find(
        {"date": {"$gte": start, "$lte": end}, "status": {"$in": ["paid", "dp"]}},
        {"_id": 0, "total": 1, "items": 1, "customer_name": 1},
    ).to_list(length=50000)

    # Material lookup untuk COGS per-customer (dipakai top_customers saja)
    material_cache: Dict[str, Dict[str, Any]] = {}
    async def _mat(mid: Optional[str]):
        if not mid:
            return {}
        if mid in material_cache:
            return material_cache[mid]
        m = await db.materials.find_one({"id": mid}, {"_id": 0, "purchase_price": 1}) or {}
        material_cache[mid] = m
        return m

    revenue = 0.0
    cust_agg: Dict[str, Dict[str, Any]] = {}
    for s in sales:
        sale_revenue = 0.0
        sale_cogs = 0.0
        for it in (s.get("items") or []):
            sale_revenue += float(it.get("subtotal", 0) or 0)
            comps = it.get("components") or []
            if comps:
                for c in comps:
                    m = await _mat(c.get("material_id"))
                    sale_cogs += float(c.get("consumption", 0) or 0) * float(m.get("purchase_price", 0) or 0)
            elif it.get("material_id"):
                m = await _mat(it.get("material_id"))
                sale_cogs += float(it.get("area_total", 0) or 0) * float(m.get("purchase_price", 0) or 0)
        revenue += sale_revenue
        key = (s.get("customer_name") or "-").strip() or "-"
        row = cust_agg.setdefault(key, {"customer": key, "orders": 0, "revenue": 0.0, "material_cost": 0.0})
        row["orders"] += 1
        row["revenue"] += sale_revenue
        row["material_cost"] += sale_cogs

    # --- BEBAN A&U dari cash_transactions (type=out) per account_code ---
    expense_tx = await db.cash_transactions.find(
        {"date": {"$gte": start, "$lte": end}, "type": "out"},
        {"_id": 0, "account_code": 1, "amount": 1},
    ).to_list(length=200000)
    code_totals: Dict[str, float] = {}
    for t in expense_tx:
        code = t.get("account_code") or ""
        code_totals[code] = code_totals.get(code, 0.0) + float(t.get("amount") or 0)

    def _sum(*codes: str) -> float:
        return round(sum(code_totals.get(c, 0.0) for c in codes), 2)

    # Struktur sesuai permintaan user (12 item, urutan tetap)
    expenses = [
        {"key": "gaji", "label": "By. Gaji", "codes": ["505"], "amount": _sum("505")},
        {"key": "atk_fc", "label": "By. ATK, Fotocopy, Dll", "codes": ["104"], "amount": _sum("104")},
        {"key": "telp_listrik_air", "label": "By. Telephone, Listrik & Air", "codes": ["502"], "amount": _sum("502")},
        {"key": "keperluan_kantor", "label": "By. Keperluan Kantor", "codes": ["104"], "amount": _sum("104")},
        {"key": "handling", "label": "By. Jasa Handling Barang", "codes": ["507", "106"], "amount": _sum("507", "106")},
        {"key": "penyusutan_ga", "label": "By. Penyusutan GA", "codes": ["513"], "amount": _sum("513")},
        {"key": "perbaikan_kendaraan", "label": "By. Perbaikan & Perawatan Kendaraan", "codes": ["105"], "amount": _sum("105")},
        {"key": "operasional_kendaraan", "label": "By. Operasional Kendaraan", "codes": ["501"], "amount": _sum("501")},
        {"key": "adm_bank", "label": "By. Administrasi Bank", "codes": ["511"], "amount": _sum("511")},
        {"key": "pajak", "label": "By. Pajak", "codes": ["514"], "amount": _sum("514")},
        {"key": "perbaikan_mesin", "label": "By. Perbaikan Mesin", "codes": ["402"], "amount": _sum("402")},
        {"key": "bahan_baku_mesin", "label": "By. Pembelian Bahan Baku Mesin", "codes": ["103-01"], "amount": _sum("103-01")},
    ]
    total_expenses = round(sum(e["amount"] for e in expenses), 2)
    net_profit = round(revenue - total_expenses, 2)

    # Top customer
    top_customers = sorted(cust_agg.values(), key=lambda r: r["revenue"], reverse=True)[:10]
    for r in top_customers:
        r["revenue"] = round(r["revenue"], 2)
        r["material_cost"] = round(r["material_cost"], 2)
        r["margin"] = round(r["revenue"] - r["material_cost"], 2)

    return {
        "period": period,
        "revenue": round(revenue, 2),
        "order_count": len(sales),
        "expenses": expenses,
        "total_expenses": total_expenses,
        "net_profit": net_profit,
        "net_margin_pct": round((net_profit / revenue * 100) if revenue > 0 else 0, 2),
        "top_customers": top_customers,
        # Legacy fields (dipakai grafik/PDF lama) — biar backward compatible
        "cogs": 0.0,
        "gross_profit": round(revenue, 2),
        "gross_margin_pct": 100.0 if revenue > 0 else 0.0,
        "waste_loss": 0.0,
        "waste_records": 0,
        "payroll_cost": expenses[0]["amount"],
        "employee_count": 0,
    }


@api_router.get("/reports/profit-loss-trend")
async def profit_loss_trend(months: int = 12, user: dict = Depends(require_super_admin)):
    """Return P&L summary utk N bulan terakhir termasuk bulan sekarang. Plus data YoY (bulan sama tahun lalu)."""
    if months < 1 or months > 36:
        raise HTTPException(status_code=400, detail="months harus 1-36")
    today = datetime.now(timezone.utc).date()
    # Buat list periode dari (months-1) bulan lalu → sekarang
    periods = []
    y, m = today.year, today.month
    for _ in range(months):
        periods.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    periods.reverse()
    # YoY: bulan-bulan yg sama tahun sebelumnya
    yoy_periods = []
    for p in periods:
        yy, mm = int(p[:4]) - 1, int(p[5:7])
        yoy_periods.append(f"{yy:04d}-{mm:02d}")

    async def _summary(period: str):
        start, end, _, _ = _parse_month(period)
        # Sales paid/dp — revenue & COGS via BOM (konsisten dgn profit_loss_report)
        sales = await db.sales.find(
            {"date": {"$gte": start, "$lte": end}, "status": {"$in": ["paid", "dp"]}},
            {"_id": 0, "total": 1, "items": 1},
        ).to_list(length=50000)
        revenue = 0.0
        cogs = 0.0
        _mcache: Dict[str, Dict[str, Any]] = {}
        async def _mp(mid: Optional[str]):
            if not mid: return 0.0
            if mid in _mcache:
                return float(_mcache[mid].get("purchase_price", 0) or 0)
            m = await db.materials.find_one({"id": mid}, {"_id": 0, "purchase_price": 1}) or {}
            _mcache[mid] = m
            return float(m.get("purchase_price", 0) or 0)
        for s in sales:
            for it in (s.get("items") or []):
                revenue += float(it.get("subtotal", 0) or 0)
                comps = it.get("components") or []
                if comps:
                    for c in comps:
                        cogs += float(c.get("consumption", 0) or 0) * await _mp(c.get("material_id"))
                elif it.get("material_id"):
                    cogs += float(it.get("area_total", 0) or 0) * await _mp(it.get("material_id"))
        waste_docs = await db.waste.find({"date": {"$gte": start, "$lte": end}}, {"_id": 0, "estimated_loss": 1}).to_list(length=5000)
        waste_loss = sum(float(w.get("estimated_loss", 0)) for w in waste_docs)
        payroll_cost, _ = await _payroll_cost_for_month(period)
        gross = revenue - cogs
        net = gross - (waste_loss + payroll_cost)
        return {
            "period": period,
            "revenue": round(revenue, 2),
            "cogs": round(cogs, 2),
            "waste_loss": round(waste_loss, 2),
            "payroll_cost": round(payroll_cost, 2),
            "gross_profit": round(gross, 2),
            "net_profit": round(net, 2),
            "order_count": len(sales),
        }

    # Fetch parallel
    curr_data, yoy_data = await asyncio.gather(
        asyncio.gather(*[_summary(p) for p in periods]),
        asyncio.gather(*[_summary(p) for p in yoy_periods]),
    )
    # Gabungkan yoy ke curr_data
    yoy_map = {y["period"]: y for y in yoy_data}
    for i, row in enumerate(curr_data):
        yoy_row = yoy_map.get(yoy_periods[i], {})
        row["yoy_period"] = yoy_periods[i]
        row["yoy_revenue"] = yoy_row.get("revenue", 0)
        row["yoy_net_profit"] = yoy_row.get("net_profit", 0)
        # Growth %
        base_rev = row["yoy_revenue"]
        base_np = row["yoy_net_profit"]
        row["revenue_growth_pct"] = round(((row["revenue"] - base_rev) / abs(base_rev) * 100) if base_rev != 0 else 0, 2) if base_rev != 0 else None
        row["net_profit_growth_pct"] = round(((row["net_profit"] - base_np) / abs(base_np) * 100) if base_np != 0 else 0, 2) if base_np != 0 else None
    # Rangkuman total periode
    total_revenue = sum(r["revenue"] for r in curr_data)
    total_net = sum(r["net_profit"] for r in curr_data)
    total_yoy_revenue = sum(r["yoy_revenue"] for r in curr_data)
    total_yoy_net = sum(r["yoy_net_profit"] for r in curr_data)
    return {
        "months": months,
        "periods": periods,
        "data": curr_data,
        "totals": {
            "revenue": round(total_revenue, 2),
            "net_profit": round(total_net, 2),
            "yoy_revenue": round(total_yoy_revenue, 2),
            "yoy_net_profit": round(total_yoy_net, 2),
            "revenue_growth_pct": round(((total_revenue - total_yoy_revenue) / abs(total_yoy_revenue) * 100) if total_yoy_revenue != 0 else 0, 2) if total_yoy_revenue != 0 else None,
            "net_profit_growth_pct": round(((total_net - total_yoy_net) / abs(total_yoy_net) * 100) if total_yoy_net != 0 else 0, 2) if total_yoy_net != 0 else None,
        },
    }


@api_router.get("/reports/profit-loss/{period}/pdf")
async def profit_loss_pdf(period: str, user: dict = Depends(require_super_admin)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    r = await profit_loss_report(period, user)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Title"], fontSize=16, spaceAfter=6)
    elems = [
        Paragraph("<b>LAPORAN LABA / RUGI BULANAN</b>", h1),
        Paragraph(f"Periode: <b>{period}</b>", styles["Normal"]),
        Spacer(1, 10),
    ]
    def _fmt(n):
        return f"Rp {n:,.0f}".replace(",", ".")
    rows = [
        ["Uraian", "Jumlah"],
        [f"Pendapatan Penjualan ({r['order_count']} order)", _fmt(r['revenue'])],
        ["(-) Biaya Bahan Baku (COGS)", _fmt(r['cogs'])],
        [f"LABA KOTOR ({r['gross_margin_pct']}%)", _fmt(r['gross_profit'])],
        [f"(-) Kerugian Waste/Rijek ({r['waste_records']} record)", _fmt(r['waste_loss'])],
        [f"(-) Biaya Gaji Karyawan ({r['employee_count']} karyawan)", _fmt(r['payroll_cost'])],
        ["Total Beban Operasional", _fmt(r['total_expenses'])],
        [f"LABA / RUGI BERSIH ({r['net_margin_pct']}%)", _fmt(r['net_profit'])],
    ]
    tbl = Table(rows, colWidths=[100 * mm, 60 * mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#e8f0ff")),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#002FA7") if r["net_profit"] >= 0 else colors.HexColor("#E81123")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(tbl)
    if r["top_customers"]:
        elems.append(Spacer(1, 14))
        elems.append(Paragraph("<b>Top Customer</b>", styles["Heading3"]))
        crows = [["#", "Customer", "Order", "Revenue", "Margin"]]
        for i, c in enumerate(r["top_customers"], 1):
            crows.append([str(i), c["customer"], str(c["orders"]), _fmt(c["revenue"]), _fmt(c["margin"])])
        ctbl = Table(crows, colWidths=[10 * mm, 70 * mm, 20 * mm, 35 * mm, 35 * mm])
        ctbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (2, 1), (4, -1), "RIGHT"),
        ]))
        elems.append(ctbl)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="laba-rugi-{period}.pdf"'},
    )



# ---------------- Purchasing Module ----------------
class SupplierIn(BaseModel):
    name: str
    phone: Optional[str] = None
    address: Optional[str] = None
    email: Optional[str] = None
    contact_person: Optional[str] = None
    category: Optional[str] = None
    notes: Optional[str] = None
    active: bool = True


class POItemIn(BaseModel):
    material_id: str
    quantity: float
    unit_price: float


class PurchaseOrderIn(BaseModel):
    supplier_id: Optional[str] = None
    supplier_name: Optional[str] = None  # jika supplier belum di-master
    po_no: Optional[str] = None
    date: str  # ISO YYYY-MM-DD
    items: List[POItemIn] = []
    tax_pct: float = 0
    notes: Optional[str] = None
    invoice_no: Optional[str] = None


PO_STATUS = ["draft", "diterima", "batal"]
PAY_STATUS = ["belum_lunas", "sebagian", "lunas"]


async def _next_po_no() -> str:
    today = datetime.now(timezone.utc).date()
    prefix = f"PO-{today.strftime('%Y%m')}-"
    count = await db.purchase_orders.count_documents({"po_no": {"$regex": f"^{re.escape(prefix)}"}})
    return f"{prefix}{count + 1:04d}"


# ----- Supplier CRUD -----
@api_router.get("/purchasing/suppliers")
async def sup_list(user: dict = Depends(require_super_admin)):
    items = await db.suppliers.find({}, {"_id": 0}).sort("name", 1).to_list(length=5000)
    # Enrich dgn agregat PO
    pos = await db.purchase_orders.find({}, {"_id": 0, "supplier_id": 1, "total": 1, "status": 1, "payment_status": 1, "amount_paid": 1}).to_list(length=20000)
    agg: Dict[str, Dict[str, Any]] = {}
    for p in pos:
        sid = p.get("supplier_id")
        if not sid or p.get("status") == "batal":
            continue
        row = agg.setdefault(sid, {"po_count": 0, "total_purchase": 0.0, "outstanding": 0.0})
        row["po_count"] += 1
        total = float(p.get("total", 0))
        paid = float(p.get("amount_paid", 0))
        row["total_purchase"] += total
        if p.get("payment_status") != "lunas":
            row["outstanding"] += max(total - paid, 0)
    for s in items:
        a = agg.get(s.get("id")) or {}
        s["po_count"] = a.get("po_count", 0)
        s["total_purchase"] = round(a.get("total_purchase", 0), 2)
        s["outstanding"] = round(a.get("outstanding", 0), 2)
    return items


@api_router.post("/purchasing/suppliers")
async def sup_create(payload: SupplierIn, user: dict = Depends(require_super_admin)):
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama supplier wajib diisi")
    safe = re.escape(payload.name.strip())
    exists = await db.suppliers.find_one({"name": {"$regex": f"^{safe}$", "$options": "i"}})
    if exists:
        raise HTTPException(status_code=400, detail="Supplier dengan nama tersebut sudah ada")
    doc = payload.model_dump()
    doc["name"] = payload.name.strip()
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.suppliers.insert_one(doc)
    await _upsert_category("supplier", doc.get("category"))
    doc.pop("_id", None)
    return doc


@api_router.put("/purchasing/suppliers/{sid}")
async def sup_update(sid: str, payload: SupplierIn, user: dict = Depends(require_super_admin)):
    existing = await db.suppliers.find_one({"id": sid})
    if not existing:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")
    new_name = payload.name.strip()
    if new_name.lower() != (existing.get("name") or "").lower():
        safe = re.escape(new_name)
        dup = await db.suppliers.find_one({"name": {"$regex": f"^{safe}$", "$options": "i"}, "id": {"$ne": sid}})
        if dup:
            raise HTTPException(status_code=400, detail="Supplier dengan nama tersebut sudah ada")
    upd = payload.model_dump()
    upd["name"] = new_name
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.suppliers.update_one({"id": sid}, {"$set": upd})
    await _upsert_category("supplier", upd.get("category"))
    return await db.suppliers.find_one({"id": sid}, {"_id": 0})


@api_router.delete("/purchasing/suppliers/{sid}")
async def sup_delete(sid: str, user: dict = Depends(require_super_admin)):
    existing = await db.suppliers.find_one({"id": sid})
    if not existing:
        raise HTTPException(status_code=404, detail="Supplier tidak ditemukan")
    has_po = await db.purchase_orders.find_one({"supplier_id": sid})
    if has_po:
        await db.suppliers.update_one({"id": sid}, {"$set": {"active": False, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"ok": True, "soft_deleted": True}
    await db.suppliers.delete_one({"id": sid})
    return {"ok": True, "soft_deleted": False}


# ----- Purchase Order CRUD -----
async def _enrich_po(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    mats = {m["id"]: m async for m in db.materials.find({}, {"_id": 0})}
    for po in items:
        for it in po.get("items") or []:
            mat = mats.get(it.get("material_id")) or {}
            it["material_name"] = mat.get("name") or "-"
            it["material_unit"] = mat.get("unit") or ""
    return items


@api_router.get("/purchasing/purchase-orders")
async def po_list(user: dict = Depends(require_super_admin), status: Optional[str] = None, payment_status: Optional[str] = None):
    q = {}
    if status:
        q["status"] = status
    if payment_status:
        q["payment_status"] = payment_status
    items = await db.purchase_orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(length=5000)
    return await _enrich_po(items)


@api_router.post("/purchasing/purchase-orders")
async def po_create(payload: PurchaseOrderIn, user: dict = Depends(require_super_admin)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="PO harus memiliki minimal 1 item bahan")
    supplier_name = payload.supplier_name
    if payload.supplier_id:
        sup = await db.suppliers.find_one({"id": payload.supplier_id})
        if not sup:
            raise HTTPException(status_code=400, detail="Supplier tidak ditemukan")
        supplier_name = sup.get("name")
    if not supplier_name:
        raise HTTPException(status_code=400, detail="Supplier wajib diisi")
    items_out = []
    subtotal = 0.0
    for it in payload.items:
        mat = await db.materials.find_one({"id": it.material_id})
        if not mat:
            raise HTTPException(status_code=400, detail=f"Bahan {it.material_id} tidak ditemukan")
        if it.quantity <= 0 or it.unit_price < 0:
            raise HTTPException(status_code=400, detail=f"Qty & harga bahan {mat.get('name')} tidak valid")
        line_total = round(float(it.quantity) * float(it.unit_price), 2)
        items_out.append({
            "material_id": it.material_id,
            "quantity": float(it.quantity),
            "unit_price": float(it.unit_price),
            "total": line_total,
        })
        subtotal += line_total
    tax = round(subtotal * float(payload.tax_pct or 0) / 100, 2)
    total = round(subtotal + tax, 2)
    po_no = payload.po_no or await _next_po_no()
    doc = {
        "id": str(uuid.uuid4()),
        "po_no": po_no,
        "supplier_id": payload.supplier_id,
        "supplier_name": supplier_name,
        "date": payload.date,
        "items": items_out,
        "subtotal": round(subtotal, 2),
        "tax_pct": float(payload.tax_pct or 0),
        "tax_amount": tax,
        "total": total,
        "status": "draft",
        "payment_status": "belum_lunas",
        "amount_paid": 0.0,
        "invoice_no": payload.invoice_no,
        "notes": payload.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("email"),
    }
    await db.purchase_orders.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/purchasing/purchase-orders/{po_id}/receive")
async def po_receive(po_id: str, user: dict = Depends(require_super_admin)):
    """Tandai PO diterima → auto-buat stock_in entry per item + update material stok & harga beli terbaru."""
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="PO tidak ditemukan")
    if po.get("status") == "diterima":
        return {"ok": True, "already": True}
    if po.get("status") == "batal":
        raise HTTPException(status_code=400, detail="PO sudah dibatalkan")
    now = datetime.now(timezone.utc).isoformat()
    for it in po.get("items") or []:
        mat = await db.materials.find_one({"id": it["material_id"]})
        if not mat:
            continue
        qty = float(it["quantity"])
        unit_price = float(it["unit_price"])
        # Insert stock_in entry
        si_doc = {
            "id": str(uuid.uuid4()),
            "material_id": it["material_id"],
            "quantity": qty,
            "unit_price": unit_price,
            "total_price": round(qty * unit_price, 2),
            "supplier": po.get("supplier_name"),
            "invoice_no": po.get("invoice_no") or po.get("po_no"),
            "date": po.get("date"),
            "notes": f"Auto dari PO {po.get('po_no')}",
            "po_id": po_id,
            "po_no": po.get("po_no"),
            "created_at": now,
            "created_by": user.get("email"),
        }
        await db.stock_in.insert_one(si_doc)
        # Update material stok & harga beli
        new_stock = round(float(mat.get("current_stock", 0)) + qty, 4)
        await db.materials.update_one(
            {"id": it["material_id"]},
            {"$set": {"current_stock": new_stock, "purchase_price": unit_price, "updated_at": now}},
        )
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"status": "diterima", "received_at": now, "received_by": user.get("email")}},
    )
    return {"ok": True, "received_at": now}


@api_router.put("/purchasing/purchase-orders/{po_id}/cancel")
async def po_cancel(po_id: str, user: dict = Depends(require_super_admin)):
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="PO tidak ditemukan")
    if po.get("status") == "diterima":
        raise HTTPException(status_code=400, detail="PO sudah diterima, batalkan penerimaan dulu bila perlu")
    await db.purchase_orders.update_one({"id": po_id}, {"$set": {"status": "batal", "cancelled_at": datetime.now(timezone.utc).isoformat()}})
    return {"ok": True}


@api_router.put("/purchasing/purchase-orders/{po_id}/pay")
async def po_pay(po_id: str, payload: Dict[str, Any] = Body(...), user: dict = Depends(require_super_admin)):
    amount = float(payload.get("amount", 0))
    if amount < 0:
        raise HTTPException(status_code=400, detail="Jumlah pembayaran tidak valid")
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="PO tidak ditemukan")
    new_paid = round(float(po.get("amount_paid", 0)) + amount, 2)
    total = float(po.get("total", 0))
    if new_paid >= total:
        new_paid = total
        payment_status = "lunas"
    elif new_paid > 0:
        payment_status = "sebagian"
    else:
        payment_status = "belum_lunas"
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"amount_paid": new_paid, "payment_status": payment_status, "last_payment_at": datetime.now(timezone.utc).isoformat()}},
    )
    # Auto-insert ke Kas Operasional (Pengeluaran Bayar Utang Usaha)
    if amount > 0:
        try:
            await _insert_cash_transaction(
                account_code="201",
                description=f"Bayar PO {po.get('po_no', po_id)} — {po.get('supplier_name', '')}".strip(" —"),
                amount=amount,
                reference=po.get("po_no") or po_id,
                date_iso=datetime.now(timezone.utc).date().isoformat(),
                auto=True,
                created_by=user.get("email"),
            )
        except Exception as ex:
            logger.warning(f"Cashbook auto-insert (PO payment) failed: {ex}")
    doc = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    return doc


@api_router.delete("/purchasing/purchase-orders/{po_id}")
async def po_delete(po_id: str, user: dict = Depends(require_super_admin)):
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="PO tidak ditemukan")
    if po.get("status") == "diterima":
        # Rollback stock_in entries + kurangi stok
        si_docs = await db.stock_in.find({"po_id": po_id}, {"_id": 0}).to_list(length=1000)
        for si in si_docs:
            mat = await db.materials.find_one({"id": si["material_id"]})
            if mat:
                new_stock = round(float(mat.get("current_stock", 0)) - float(si.get("quantity", 0)), 4)
                await db.materials.update_one({"id": si["material_id"]}, {"$set": {"current_stock": new_stock, "updated_at": datetime.now(timezone.utc).isoformat()}})
        await db.stock_in.delete_many({"po_id": po_id})
    # Rollback cash transactions AUTO yang dibuat dari pembayaran PO ini
    po_no = po.get("po_no")
    if po_no:
        await db.cash_transactions.delete_many({"reference": po_no, "auto": True, "account_code": "201"})
    await db.purchase_orders.delete_one({"id": po_id})
    return {"ok": True}


# ----- Price History -----
@api_router.get("/purchasing/price-history")
async def price_history(material_id: Optional[str] = None, user: dict = Depends(require_super_admin)):
    """Riwayat harga beli — dari stock_in (semua sumber: manual + auto-PO)."""
    q = {}
    if material_id:
        q["material_id"] = material_id
    items = await db.stock_in.find(q, {"_id": 0}).sort("date", 1).to_list(length=10000)
    items = await _enrich_with_material(items)
    # Group by material — return per material terpisah bila material_id tidak diisi
    grouped: Dict[str, Dict[str, Any]] = {}
    for it in items:
        mid = it.get("material_id")
        if not mid:
            continue
        g = grouped.setdefault(mid, {
            "material_id": mid,
            "material_name": it.get("material_name"),
            "material_unit": it.get("material_unit"),
            "history": [],
        })
        g["history"].append({
            "date": it.get("date"),
            "unit_price": float(it.get("unit_price", 0)),
            "quantity": float(it.get("quantity", 0)),
            "supplier": it.get("supplier"),
            "po_no": it.get("po_no"),
            "invoice_no": it.get("invoice_no"),
        })
    # Hitung min/max/avg + first/last
    result = []
    for g in grouped.values():
        hist = g["history"]
        prices = [h["unit_price"] for h in hist]
        g["min_price"] = round(min(prices), 2) if prices else 0
        g["max_price"] = round(max(prices), 2) if prices else 0
        g["avg_price"] = round(sum(prices) / len(prices), 2) if prices else 0
        g["current_price"] = hist[-1]["unit_price"] if hist else 0
        g["first_price"] = hist[0]["unit_price"] if hist else 0
        g["change_pct"] = round(((g["current_price"] - g["first_price"]) / g["first_price"] * 100), 2) if g["first_price"] > 0 else 0
        result.append(g)
    result.sort(key=lambda x: x.get("material_name") or "")
    return {"count": len(result), "items": result}


# ----- Purchasing Stats -----
@api_router.get("/purchasing/stats")
async def purchasing_stats(user: dict = Depends(require_super_admin)):
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(length=20000)
    total_po = len(pos)
    total_purchase = sum(float(p.get("total", 0)) for p in pos if p.get("status") != "batal")
    outstanding = 0.0
    unpaid_pos = 0
    for p in pos:
        if p.get("status") == "batal":
            continue
        if p.get("payment_status") != "lunas":
            outstanding += max(float(p.get("total", 0)) - float(p.get("amount_paid", 0)), 0)
            if p.get("payment_status") == "belum_lunas":
                unpaid_pos += 1
    total_suppliers = await db.suppliers.count_documents({"active": True})
    return {
        "total_po": total_po,
        "total_purchase": round(total_purchase, 2),
        "outstanding": round(outstanding, 2),
        "unpaid_pos": unpaid_pos,
        "total_suppliers": total_suppliers,
    }



# ---------------- Sales / POS Module ----------------
# Master Produk dengan BOM (Bill of Materials)
PRODUCT_FORMULAS = ("fixed", "per_qty", "area", "length")

class ProductComponent(BaseModel):
    material_id: str
    formula: str  # "fixed" | "per_qty" | "area" | "length"
    quantity: float = 1.0  # faktor konsumsi (untuk tier A / non-size / S-XL)
    # NEW: konsumsi untuk tier B (XXL keatas). None = pakai quantity yg sama
    quantity_size_b: Optional[float] = None

class ProductIn(BaseModel):
    code: Optional[str] = None
    name: str
    category: Optional[str] = None
    pricing_mode: str = "fixed"  # "fixed" (per unit) | "per_area" (per m²)
    unit_price: float = 0  # harga jual default (dipakai bila has_sizes=False)
    purchase_price: float = 0  # harga beli / modal per unit (opsional, bisa auto dari BOM)
    current_stock: float = 0  # stok produk jadi (untuk finished goods yang di-stok)
    components: List[ProductComponent] = []
    active: bool = True
    # NEW: sizing untuk produk kaos/jersey
    has_sizes: bool = False
    sizes: List[str] = []  # subset dari ["S","M","L","XL","XXL","XXXL"]
    price_size_a: float = 0  # harga untuk S–XL (dipakai kalau has_sizes=True)
    price_size_b: float = 0  # harga untuk XXL keatas
    # NEW: panjang per pcs (meter) — dipakai di Laporan Penjualan untuk hitung total meter
    length_meter: float = 0  # 0 = tidak ada info panjang (mis. produk non-linear seperti kaos)


# Klasifikasi tier size
SIZE_TIER_A = {"S", "M", "L", "XL"}
# Everything else (XXL, XXXL, 2XL, 3XL, dsb) = tier B


def _size_tier(size: Optional[str]) -> str:
    """Return 'A' untuk S-XL, 'B' untuk XXL+ atau default 'A' bila tidak ada."""
    if not size:
        return "A"
    return "A" if size.strip().upper() in SIZE_TIER_A else "B"


def _product_requires_dimensions(components: List[Dict[str, Any]]) -> bool:
    return any(c.get("formula") in ("area", "length") for c in components)


async def _enrich_product(p: Dict[str, Any]) -> Dict[str, Any]:
    """Isi snapshot material_name & material_unit ke tiap component. Compute bom_cost & stock_value."""
    if not p:
        return p
    bom_cost = 0.0
    for c in p.get("components", []):
        mat = await db.materials.find_one({"id": c.get("material_id")}, {"_id": 0, "name": 1, "unit": 1, "current_stock": 1, "purchase_price": 1})
        if mat:
            c["material_name"] = mat.get("name")
            c["material_unit"] = mat.get("unit")
            c["material_stock"] = mat.get("current_stock")
            c["material_purchase_price"] = mat.get("purchase_price", 0)
        else:
            c["material_name"] = "(bahan dihapus)"
            c["material_unit"] = ""
            c["material_stock"] = 0
            c["material_purchase_price"] = 0
        # BOM cost per unit produk (untuk formula per_qty & fixed cukup faktor × harga beli)
        # Untuk area/length butuh dimensi standard — kita hitung per unit dasar (asumsi 1m²/1m)
        factor = float(c.get("quantity", 1) or 1)
        buy = float(c.get("material_purchase_price", 0) or 0)
        formula = c.get("formula", "")
        if formula in ("fixed", "per_qty"):
            bom_cost += factor * buy
        elif formula in ("area", "length"):
            # Cost per 1m² atau 1m (sebagai reference); pengguna bisa hitung berdasar ukuran actual
            bom_cost += factor * buy
    p["requires_dimensions"] = _product_requires_dimensions(p.get("components", []))
    p["bom_cost"] = round(bom_cost, 2)  # modal bahan per unit (reference)
    p["stock_value"] = round(float(p.get("current_stock", 0) or 0) * float(p.get("purchase_price", 0) or 0), 2)
    return p


def _compute_component_consumption(
    formula: str, factor: float, length_m: float, width_m: float, qty: int
) -> float:
    """Hitung konsumsi bahan untuk 1 component sale."""
    f = (formula or "").lower()
    factor = float(factor or 0)
    q = int(qty or 0)
    if f == "fixed":
        return round(factor, 4)
    if f == "per_qty":
        return round(factor * q, 4)
    if f == "area":
        return round(factor * float(length_m or 0) * float(width_m or 0) * q, 4)
    if f == "length":
        return round(factor * float(length_m or 0) * q, 4)
    return 0.0


@api_router.get("/products")
async def products_list(user: dict = Depends(require_super_admin), only_active: bool = False):
    q: Dict[str, Any] = {}
    if only_active:
        q["active"] = {"$ne": False}
    items = await db.products.find(q, {"_id": 0}).sort("name", 1).to_list(length=2000)
    for p in items:
        await _enrich_product(p)
    return items


@api_router.get("/products/{product_id}")
async def products_get(product_id: str, user: dict = Depends(require_super_admin)):
    p = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    await _enrich_product(p)
    return p


@api_router.post("/products")
async def products_create(payload: ProductIn, user: dict = Depends(require_super_admin)):
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama produk wajib")
    if payload.pricing_mode not in ("fixed", "per_area"):
        raise HTTPException(status_code=400, detail="pricing_mode harus 'fixed' atau 'per_area'")
    if payload.unit_price < 0:
        raise HTTPException(status_code=400, detail="Harga tidak boleh negatif")
    for c in payload.components:
        if c.formula not in PRODUCT_FORMULAS:
            raise HTTPException(status_code=400, detail=f"Formula '{c.formula}' tidak valid. Pilih: {PRODUCT_FORMULAS}")
        if c.quantity <= 0:
            raise HTTPException(status_code=400, detail="Quantity komponen harus > 0")
        m = await db.materials.find_one({"id": c.material_id})
        if not m:
            raise HTTPException(status_code=400, detail=f"Bahan {c.material_id} tidak ditemukan")
    # Validasi sizing
    if payload.has_sizes:
        if not payload.sizes:
            raise HTTPException(status_code=400, detail="Pilih minimal 1 ukuran untuk produk ini")
        if payload.price_size_a <= 0:
            raise HTTPException(status_code=400, detail="Harga S-XL harus > 0")
        has_tier_b = any(_size_tier(s) == "B" for s in payload.sizes)
        if has_tier_b and payload.price_size_b <= 0:
            raise HTTPException(status_code=400, detail="Harga XXL keatas harus > 0")
    # Cek duplicate nama
    safe = re.escape(payload.name.strip())
    exists = await db.products.find_one({"name": {"$regex": f"^{safe}$", "$options": "i"}})
    if exists:
        raise HTTPException(status_code=400, detail="Produk dengan nama tersebut sudah ada")
    doc = {
        "id": str(uuid.uuid4()),
        "code": (payload.code or "").strip() or None,
        "name": payload.name.strip(),
        "category": (payload.category or "").strip() or None,
        "pricing_mode": payload.pricing_mode,
        "unit_price": round(float(payload.unit_price), 2),
        "purchase_price": round(float(payload.purchase_price or 0), 2),
        "current_stock": round(float(payload.current_stock or 0), 4),
        "components": [c.model_dump() for c in payload.components],
        "active": payload.active,
        # Sizing (kaos/jersey)
        "has_sizes": bool(payload.has_sizes),
        "sizes": list(payload.sizes) if payload.has_sizes else [],
        "price_size_a": round(float(payload.price_size_a or 0), 2) if payload.has_sizes else 0,
        "price_size_b": round(float(payload.price_size_b or 0), 2) if payload.has_sizes else 0,
        "length_meter": round(float(payload.length_meter or 0), 4),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.products.insert_one(doc)
    await _upsert_category("product", doc.get("category"))
    doc.pop("_id", None)
    await _enrich_product(doc)
    return doc


@api_router.put("/products/{product_id}")
async def products_update(product_id: str, payload: ProductIn, user: dict = Depends(require_super_admin)):
    existing = await db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    if payload.pricing_mode not in ("fixed", "per_area"):
        raise HTTPException(status_code=400, detail="pricing_mode harus 'fixed' atau 'per_area'")
    for c in payload.components:
        if c.formula not in PRODUCT_FORMULAS:
            raise HTTPException(status_code=400, detail=f"Formula '{c.formula}' tidak valid")
        if c.quantity <= 0:
            raise HTTPException(status_code=400, detail="Quantity komponen harus > 0")
        m = await db.materials.find_one({"id": c.material_id})
        if not m:
            raise HTTPException(status_code=400, detail=f"Bahan {c.material_id} tidak ditemukan")
    # Validasi sizing
    if payload.has_sizes:
        if not payload.sizes:
            raise HTTPException(status_code=400, detail="Pilih minimal 1 ukuran untuk produk ini")
        if payload.price_size_a <= 0:
            raise HTTPException(status_code=400, detail="Harga S-XL harus > 0")
        has_tier_b = any(_size_tier(s) == "B" for s in payload.sizes)
        if has_tier_b and payload.price_size_b <= 0:
            raise HTTPException(status_code=400, detail="Harga XXL keatas harus > 0")
    upd = {
        "code": (payload.code or "").strip() or None,
        "name": payload.name.strip(),
        "category": (payload.category or "").strip() or None,
        "pricing_mode": payload.pricing_mode,
        "unit_price": round(float(payload.unit_price), 2),
        "purchase_price": round(float(payload.purchase_price or 0), 2),
        "current_stock": round(float(payload.current_stock or 0), 4),
        "components": [c.model_dump() for c in payload.components],
        "active": payload.active,
        # Sizing
        "has_sizes": bool(payload.has_sizes),
        "sizes": list(payload.sizes) if payload.has_sizes else [],
        "price_size_a": round(float(payload.price_size_a or 0), 2) if payload.has_sizes else 0,
        "price_size_b": round(float(payload.price_size_b or 0), 2) if payload.has_sizes else 0,
        "length_meter": round(float(payload.length_meter or 0), 4),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.products.update_one({"id": product_id}, {"$set": upd})
    await _upsert_category("product", upd.get("category"))
    doc = await db.products.find_one({"id": product_id}, {"_id": 0})
    await _enrich_product(doc)
    return doc


@api_router.delete("/products/{product_id}")
async def products_delete(product_id: str, user: dict = Depends(require_super_admin)):
    existing = await db.products.find_one({"id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Produk tidak ditemukan")
    used = await db.sales.count_documents({"items.product_id": product_id})
    if used > 0:
        raise HTTPException(status_code=400, detail=f"Produk masih dipakai di {used} transaksi. Non-aktifkan saja.")
    await db.products.delete_one({"id": product_id})
    return {"ok": True}


PAYMENT_ACCOUNT_MAP = {
    "cash": "101",             # Penjualan Tunai → langsung masuk Kas Utama (2026-08-08)
    "transfer_bca": "301-BCA", # Transfer BCA
    "transfer_mandiri": "301-MDR",
    "shopee_plaza": "301-SPP",
    "shopee_kastem": "301-SPK",
}


def _resolve_payment_account(payment_method: str, payment_bank: Optional[str] = None) -> Tuple[str, str]:
    """Return (account_code, human_label) untuk auto cash tx."""
    pm = (payment_method or "cash").lower()
    if pm == "transfer":
        b = (payment_bank or "").strip().lower()
        if b == "mandiri":
            return ("301-MDR", "Transfer Mandiri")
        return ("301-BCA", "Transfer BCA")
    if pm == "shopee_plaza":
        return ("301-SPP", "Shopee Plaza")
    if pm == "shopee_kastem":
        return ("301-SPK", "Shopee Kastem")
    if pm in ("cash", "tunai"):
        # 2026-08-08: cash sales langsung masuk akun 101 Kas Utama supaya
        # Saldo Kas Real-time (header) sinkron dengan Jurnal Akuntansi.
        return ("101", "Penjualan Tunai")
    # Fallback (legacy "tunai" atau lainnya)
    return ("101", "Penjualan Tunai")



def _company_info() -> Dict[str, str]:
    return {
        "name": os.environ.get("COMPANY_NAME", "PLAZAKREASI DIGITAL PRINTING"),
        "address": os.environ.get("COMPANY_ADDRESS", "Jl. Ruko Sentralan B72 Driyorejo Gresik"),
        "phone": os.environ.get("COMPANY_PHONE", "081235598288"),
    }


# ================================================================
# ==================== KAS OPERASIONAL (Cash Book) ================
# ================================================================
# Chart of Accounts default (bisa di-extend via UI)
DEFAULT_CASH_ACCOUNTS = [
    # Pemasukan (income) — per metode pembayaran
    {"code": "301", "name": "Penjualan Tunai", "type": "in", "system": True},
    {"code": "301-BCA", "name": "Penjualan via Transfer BCA", "type": "in", "system": True},
    {"code": "301-MDR", "name": "Penjualan via Transfer Mandiri", "type": "in", "system": True},
    {"code": "301-SPP", "name": "Penjualan via Shopee Plaza", "type": "in", "system": True},
    {"code": "301-SPK", "name": "Penjualan via Shopee Kastem", "type": "in", "system": True},
    {"code": "302", "name": "Terima Piutang", "type": "in", "system": False},
    {"code": "303", "name": "Modal / Setoran Kas", "type": "in", "system": False},
    {"code": "304", "name": "Pendapatan Lain-lain", "type": "in", "system": False},
    # Kode Akun Akuntansi Standar (Assets & Expense) — permintaan user
    {"code": "101", "name": "Kas", "type": "in", "system": False},
    # 102-PTP: Piutang Perusahaan — KHUSUS Jurnal Akuntansi.
    # Dipakai untuk menyeimbangkan saldo minus di Jurnal (transfer dari akun minus).
    # DILARANG muncul di Buku Kas — di-filter di frontend (filteredJournal).
    {"code": "102-PTP", "name": "Piutang Perusahaan", "type": "in", "system": True},
    {"code": "103", "name": "Persediaan Barang", "type": "out", "system": False},
    {"code": "103-01", "name": "Bahan Baku Mesin", "type": "out", "system": False},
    {"code": "104", "name": "By. ATK & Keperluan Kantor", "type": "out", "system": False},
    {"code": "105", "name": "BBM dan Maintenance Kendaraan", "type": "out", "system": False},
    {"code": "106", "name": "Pengiriman Dokumen", "type": "out", "system": False},
    {"code": "108", "name": "Makan dan Entertainment", "type": "out", "system": False},
    # Pengeluaran (expense) — kode lama tetap kompatibel
    {"code": "201", "name": "Bayar Utang Usaha", "type": "out", "system": True},
    {"code": "401", "name": "Pembelian Bahan Baku", "type": "out", "system": False},
    {"code": "402", "name": "By. Perbaikan Mesin", "type": "out", "system": False},
    {"code": "403", "name": "Alat Tulis Kantor", "type": "out", "system": False},
    {"code": "501", "name": "BBM, Parkir & Maintenance Kendaraan", "type": "out", "system": False},
    {"code": "502", "name": "Beban Listrik, Air, Telepon", "type": "out", "system": False},
    {"code": "503", "name": "Sewa Kendaraan", "type": "out", "system": False},
    {"code": "504", "name": "Sewa Bangunan / Mess", "type": "out", "system": False},
    {"code": "505", "name": "Gaji Karyawan", "type": "out", "system": False},
    {"code": "506", "name": "Makan & Entertainment", "type": "out", "system": False},
    {"code": "507", "name": "Pengiriman Dokumen / Barang", "type": "out", "system": False},
    {"code": "508", "name": "Promosi & Iklan", "type": "out", "system": False},
    {"code": "509", "name": "Percetakan", "type": "out", "system": False},
    {"code": "510", "name": "Jasa Freelancer", "type": "out", "system": False},
    {"code": "511", "name": "Biaya Administrasi Bank", "type": "out", "system": False},
    {"code": "512", "name": "Kasbon Karyawan", "type": "out", "system": False},
    {"code": "513", "name": "By. Penyusutan GA", "type": "out", "system": False},
    {"code": "514", "name": "By. Pajak", "type": "out", "system": False},
    {"code": "599", "name": "Lain-lain", "type": "out", "system": False},
]


# Kode akun yang butuh RENAME dari nama lama → nama baru (migrasi 2026-08-14)
_ACCOUNT_RENAMES = {
    "402": "By. Perbaikan Mesin",
    "104": "By. ATK & Keperluan Kantor",
}


async def _ensure_cash_accounts():
    """Seed default chart of accounts idempotently (per kode akun).
    Selain seed baru, juga migrasi rename untuk kode di _ACCOUNT_RENAMES.
    """
    for a in DEFAULT_CASH_ACCOUNTS:
        exists = await db.cash_accounts.find_one({"code": a["code"]})
        if exists:
            # Terapkan rename bila kode termasuk dalam _ACCOUNT_RENAMES dan namanya beda.
            new_name = _ACCOUNT_RENAMES.get(a["code"])
            if new_name and exists.get("name") != new_name:
                await db.cash_accounts.update_one({"code": a["code"]}, {"$set": {"name": new_name}})
            continue
        await db.cash_accounts.insert_one({
            "id": str(uuid.uuid4()),
            "code": a["code"],
            "name": a["name"],
            "type": a["type"],
            "system": a["system"],
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })


async def _ensure_shopee_admin_fee_account():
    """Ensure account 502-SHP 'Biaya Admin Shopee' exists (type=out)."""
    exists = await db.cash_accounts.find_one({"code": "502-SHP"})
    if exists:
        return
    await db.cash_accounts.insert_one({
        "id": str(uuid.uuid4()),
        "code": "502-SHP",
        "name": "Biaya Admin Shopee",
        "type": "out",
        "system": True,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })




async def _insert_cash_transaction(
    account_code: str,
    description: str,
    amount: float,
    reference: Optional[str] = None,
    date_iso: Optional[str] = None,
    auto: bool = False,
    created_by: Optional[str] = None,
) -> Dict[str, Any]:
    """Insert satu transaksi kas. Return doc."""
    await _ensure_cash_accounts()
    acc = await db.cash_accounts.find_one({"code": account_code}, {"_id": 0})
    if not acc:
        raise HTTPException(status_code=404, detail=f"Akun {account_code} tidak ditemukan")
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Jumlah harus > 0")
    doc = {
        "id": str(uuid.uuid4()),
        "date": date_iso or datetime.now(timezone.utc).date().isoformat(),
        "account_code": acc["code"],
        "account_name": acc["name"],
        "type": acc["type"],  # "in" atau "out"
        "description": description.strip(),
        "amount": round(float(amount), 2),
        "reference": reference,
        "auto": auto,
        "created_by": created_by,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.cash_transactions.insert_one(doc)
    doc.pop("_id", None)
    return doc





# ---------------- Cashbook Router (extracted 2026-08-01, part 2) ----------------
from routers.cashbook import make_router as _make_cashbook_router
api_router.include_router(_make_cashbook_router(
    db=db,
    require_super_admin=require_super_admin,
    logger=logger,
    _insert_cash_transaction=_insert_cash_transaction,
    _ensure_cash_accounts=_ensure_cash_accounts,
    _resolve_payment_account=_resolve_payment_account,
    PAYMENT_ACCOUNT_MAP=PAYMENT_ACCOUNT_MAP,
    _company_info=_company_info,
))

# ---------------- Sales Router (extracted 2026-08-01, part 3) ----------------
from routers.sales import make_router as _make_sales_router
api_router.include_router(_make_sales_router(
    db=db,
    require_super_admin=require_super_admin,
    logger=logger,
    _insert_cash_transaction=_insert_cash_transaction,
    _resolve_payment_account=_resolve_payment_account,
    PAYMENT_ACCOUNT_MAP=PAYMENT_ACCOUNT_MAP,
    _size_tier=_size_tier,
    _compute_component_consumption=_compute_component_consumption,
    _sanitize_branch=_sanitize_branch,
    _company_info=_company_info,
))


# ---------------- Backup Router (added 2026-08-01) ----------------
from routers.backup import make_router as _make_backup_router
api_router.include_router(_make_backup_router(db=db, require_super_admin=require_super_admin, logger=logger))


# ---------------- Payroll Router (extracted 2026-08-04, part 5) ----------------
from routers.payroll import make_router as _make_payroll_router
api_router.include_router(_make_payroll_router(
    db=db,
    require_super_admin=require_super_admin,
    logger=logger,
    calculate_payslip=calculate_payslip,
    _calculate_thr=_calculate_thr,
    _build_payslip_pdf=_build_payslip_pdf,
    _payslip_html=_payslip_html,
    _send_email_via_resend=_send_email_via_resend,
    _whatsapp_slip_message=_whatsapp_slip_message,
    _send_whatsapp=_send_whatsapp,
    _format_bank_export=_format_bank_export,
    _build_annual_summary=_build_annual_summary,
    _build_bukti_potong_pdf=_build_bukti_potong_pdf,
))


# ---------------- Employees Router (extracted 2026-08-04, part 6) ----------------
from routers.employees import make_router as _make_employees_router
api_router.include_router(_make_employees_router(
    db=db,
    require_super_admin=require_super_admin,
    logger=logger,
    EmployeeIn=EmployeeIn,
    EMPLOYEE_CSV_HEADERS=EMPLOYEE_CSV_HEADERS,
))


# ---------------- Portal Router (extracted 2026-08-04, part 7) ----------------
from routers.portal import make_router as _make_portal_router
api_router.include_router(_make_portal_router(
    db=db,
    logger=logger,
    get_current_employee=get_current_employee,
    create_portal_token=create_portal_token,
    _build_payslip_pdf=_build_payslip_pdf,
    _build_annual_summary=_build_annual_summary,
    _build_bukti_potong_pdf=_build_bukti_potong_pdf,
    _send_simple_email=_send_simple_email,
    _leave_view=_leave_view,
    LEAVE_TYPES=LEAVE_TYPES,
    LEAVE_TYPE_LABELS=LEAVE_TYPE_LABELS,
    MAX_ATTACHMENT_SIZE=MAX_ATTACHMENT_SIZE,
    ALLOWED_ATTACHMENT_MIME=ALLOWED_ATTACHMENT_MIME,
))






# Include router
app.include_router(api_router)


# ---------------- RBAC Middleware ----------------
# Enforces per-menu access for role=admin_privileged based on request path prefix.
# Super admin, portal endpoints, and auth endpoints bypass this check.
@app.middleware("http")
async def rbac_middleware(request: Request, call_next):
    path = request.url.path
    # Only guard /api paths that aren't in bypass list
    if not path.startswith("/api"):
        return await call_next(request)
    if any(path.startswith(pref) for pref in RBAC_BYPASS_PREFIXES):
        return await call_next(request)

    # Extract token (cookie or bearer). No token -> let endpoint dep return 401.
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        return await call_next(request)

    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            return await call_next(request)
        user_id = payload.get("sub")
    except jwt.PyJWTError:
        return await call_next(request)

    if not user_id:
        return await call_next(request)
    u = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    if not u:
        return await call_next(request)

    role = u.get("role")
    if role == "admin":
        role = ROLE_SUPER_ADMIN
    if role == "hr_leave":
        role = ROLE_ADMIN_PRIVILEGED
    if role == ROLE_SUPER_ADMIN:
        return await call_next(request)
    if role != ROLE_ADMIN_PRIVILEGED:
        return await call_next(request)

    # admin_privileged → check menu perm
    menu = _menu_for_path(path)
    if menu is None:
        # Path not mapped to any menu → allow (fallthrough to endpoint auth)
        return await call_next(request)
    perms = u.get("permissions") or []
    if menu not in perms:
        from starlette.responses import JSONResponse
        return JSONResponse(
            status_code=403,
            content={"detail": f"Akses ditolak: Anda tidak memiliki izin untuk menu '{menu}'"},
        )
    return await call_next(request)


# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
