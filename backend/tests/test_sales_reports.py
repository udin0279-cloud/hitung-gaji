"""Tests for Sales Report/Invoice endpoints (Nota A4 PDF, Laporan Bulanan PDF, Export Excel).

Endpoints covered:
  - GET /api/sales/{sale_id}/invoice-pdf   -> application/pdf (Nota A4)
  - GET /api/sales/report/pdf?month=YYYY-MM  -> application/pdf (landscape laporan)
  - GET /api/sales/report/pdf?date_from=&date_to=  -> application/pdf
  - GET /api/sales/report/excel?month=YYYY-MM  -> .xlsx
"""
import io
import os
import re
import zipfile
import pytest
import requests

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/") if os.environ.get("REACT_APP_BACKEND_URL") else None
if not BASE_URL:
    from pathlib import Path
    for line in Path("/app/frontend/.env").read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
            break

PDF_MIME = "application/pdf"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------- Fixtures ----------------

@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json={"email": "admin@payroll.id", "password": "admin123"})
    if r.status_code != 200:
        pytest.skip(f"Login failed: {r.status_code} {r.text}")
    return s


@pytest.fixture(scope="module")
def existing_sale(client):
    """Ambil salah satu sale existing untuk test invoice-pdf; fallback membuat sale baru bila kosong."""
    r = client.get(f"{BASE_URL}/api/sales?limit=5")
    assert r.status_code == 200, r.text
    lst = r.json()
    if isinstance(lst, list) and lst:
        return lst[0]

    # Fallback: buat material + sale sederhana (agar test tetap berjalan pada DB kosong)
    m = client.post(f"{BASE_URL}/api/inventory/materials", json={
        "name": "TEST_REPORT_Flex", "category": "flexy", "unit": "meter",
        "current_stock": 100.0, "purchase_price": 10000, "selling_price": 20000,
        "min_stock": 0, "active": True,
    })
    assert m.status_code == 200, m.text
    mat = m.json()
    r2 = client.post(f"{BASE_URL}/api/sales", json={
        "customer_name": "TEST_REPORT_Umum", "customer_phone": "081200",
        "items": [{
            "material_id": mat["id"], "product_name": "Banner Uji Report",
            "length_m": 2, "width_m": 1, "quantity": 1, "unit_price": 20000,
        }],
        "discount": 0, "cash_paid": 40000, "payment_method": "tunai",
    })
    assert r2.status_code == 200, r2.text
    return r2.json()


# ---------------- Invoice A4 PDF ----------------

class TestInvoicePDF:
    def test_invoice_pdf_success(self, client, existing_sale):
        sid = existing_sale["id"]
        r = client.get(f"{BASE_URL}/api/sales/{sid}/invoice-pdf")
        assert r.status_code == 200, r.text[:200]
        # Content-Type
        assert PDF_MIME in r.headers.get("content-type", ""), r.headers.get("content-type")
        # Magic bytes
        assert r.content[:4] == b"%PDF", f"Not a PDF, first bytes: {r.content[:8]}"
        # Non-trivial payload
        assert len(r.content) > 2000
        # Filename hint in Content-Disposition
        cd = r.headers.get("content-disposition", "")
        assert "Nota_" in cd or "filename" in cd.lower()

    def test_invoice_pdf_404_on_missing_sale(self, client):
        r = client.get(f"{BASE_URL}/api/sales/does-not-exist-xyz/invoice-pdf")
        assert r.status_code == 404

    def test_invoice_pdf_requires_auth(self):
        # unauth session -> harus 401/403
        anon = requests.Session()
        r = anon.get(f"{BASE_URL}/api/sales/anything/invoice-pdf")
        assert r.status_code in (401, 403), r.status_code


# ---------------- Report PDF ----------------

class TestReportPDF:
    def test_report_pdf_by_month(self, client, existing_sale):
        # gunakan bulan dari sale existing agar ada isi
        d = existing_sale.get("date", "")  # YYYY-MM-DD
        month = d[:7] if len(d) >= 7 else "2026-07"
        r = client.get(f"{BASE_URL}/api/sales/report/pdf", params={"month": month})
        assert r.status_code == 200, r.text[:200]
        assert PDF_MIME in r.headers.get("content-type", "")
        assert r.content[:4] == b"%PDF"
        assert len(r.content) > 1500

    def test_report_pdf_empty_month_still_200(self, client):
        # bulan kosong harus tetap 200 dengan pesan 'Belum ada transaksi'
        r = client.get(f"{BASE_URL}/api/sales/report/pdf", params={"month": "2020-01"})
        assert r.status_code == 200, r.text[:200]
        assert PDF_MIME in r.headers.get("content-type", "")
        assert r.content[:4] == b"%PDF"

    def test_report_pdf_by_date_range(self, client, existing_sale):
        d = existing_sale.get("date", "2026-07-11")
        r = client.get(f"{BASE_URL}/api/sales/report/pdf", params={
            "date_from": d, "date_to": "2099-12-31",
        })
        assert r.status_code == 200, r.text[:200]
        assert PDF_MIME in r.headers.get("content-type", "")
        assert r.content[:4] == b"%PDF"

    def test_report_pdf_invalid_month(self, client):
        r = client.get(f"{BASE_URL}/api/sales/report/pdf", params={"month": "not-a-month"})
        assert r.status_code == 400


# ---------------- Report Excel ----------------

class TestReportExcel:
    EXPECTED_COLUMNS = [
        "Tanggal", "No. Nota", "Pelanggan", "No. Telepon", "Kasir",
        "Jumlah Item", "Subtotal", "Diskon", "Total", "Bayar Tunai",
        "Kembali", "Metode", "Catatan",
    ]

    def test_excel_success_and_structure(self, client, existing_sale):
        d = existing_sale.get("date", "")
        month = d[:7] if len(d) >= 7 else "2026-07"
        r = client.get(f"{BASE_URL}/api/sales/report/excel", params={"month": month})
        assert r.status_code == 200, r.text[:200]
        assert XLSX_MIME in r.headers.get("content-type", "")
        # .xlsx = ZIP
        assert r.content[:2] == b"PK", f"Not a ZIP/xlsx: {r.content[:8]}"
        # Content-Disposition harus attachment
        assert "attachment" in r.headers.get("content-disposition", "").lower()

        # Parse xlsx dengan openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available for xlsx parsing")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Row 1-3 = header perusahaan; Row 4 = kolom (startrow=3 zero-index)
        a1 = str(ws["A1"].value or "")
        a2 = str(ws["A2"].value or "")
        a3 = str(ws["A3"].value or "")
        assert a1.strip() != "", "A1 kosong (harus nama perusahaan)"
        assert "Laporan Penjualan" in a3, f"A3 tidak berisi label periode: {a3}"

        # Baris kolom pada row 4
        header_values = [str(ws.cell(row=4, column=c).value or "") for c in range(1, 14)]
        # Semua kolom expected ada
        for col in self.EXPECTED_COLUMNS:
            assert col in header_values, f"Kolom '{col}' hilang dari header {header_values}"
        # Tepat 13 kolom
        assert len([v for v in header_values if v]) == 13, header_values

    def test_excel_empty_month_still_200(self, client):
        r = client.get(f"{BASE_URL}/api/sales/report/excel", params={"month": "2020-01"})
        assert r.status_code == 200, r.text[:200]
        assert XLSX_MIME in r.headers.get("content-type", "")
        assert r.content[:2] == b"PK"
        # Harus tetap valid zipfile
        assert zipfile.is_zipfile(io.BytesIO(r.content))

    def test_excel_by_date_range(self, client, existing_sale):
        d = existing_sale.get("date", "2026-07-11")
        r = client.get(f"{BASE_URL}/api/sales/report/excel", params={
            "date_from": d, "date_to": "2099-12-31",
        })
        assert r.status_code == 200
        assert XLSX_MIME in r.headers.get("content-type", "")
        assert r.content[:2] == b"PK"

    def test_excel_total_footer_present(self, client, existing_sale):
        """Row TOTAL harus ada di footer bila ada transaksi."""
        d = existing_sale.get("date", "")
        month = d[:7] if len(d) >= 7 else "2026-07"
        r = client.get(f"{BASE_URL}/api/sales/report/excel", params={"month": month})
        assert r.status_code == 200
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # cari cell dengan value "TOTAL" pada kolom A di baris > 4
        found = False
        for row in range(5, ws.max_row + 1):
            v = str(ws.cell(row=row, column=1).value or "")
            if v.strip().upper() == "TOTAL":
                found = True
                # cek total kolom 9 (Total) berupa angka
                total_val = ws.cell(row=row, column=9).value
                assert isinstance(total_val, (int, float)), f"Total bukan numeric: {total_val!r}"
                break
        assert found, "Row TOTAL tidak ditemukan di footer"

    def test_excel_invalid_month(self, client):
        r = client.get(f"{BASE_URL}/api/sales/report/excel", params={"month": "abc-xy"})
        assert r.status_code == 400


# ---------------- Regression: existing sales endpoints tetap OK ----------------

class TestSalesRegression:
    def test_sales_list_still_ok(self, client):
        r = client.get(f"{BASE_URL}/api/sales?limit=5")
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_sales_stats_today_still_ok(self, client):
        r = client.get(f"{BASE_URL}/api/sales/stats/today")
        assert r.status_code == 200
        d = r.json()
        for k in ["count_today", "total_today", "count_month", "total_month"]:
            assert k in d
