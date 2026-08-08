"""Sales router — extracted from server.py (POC refactor 2026-08-01, part 3).

Endpoints (18+):
  GET/POST/PUT/DELETE /sales, /sales/{id}
  POST                /sales/bulk-tag-branch
  GET (HTML)          /sales/{id}/receipt
  GET (PDF)           /sales/{id}/invoice-pdf
  GET (PDF/Excel)     /sales/report/pdf, /sales/report/excel
  GET                 /sales/report/shopee-rincian
  GET                 /sales/report/analytics
  PATCH               /sales/{id}/saldo-masuk
  POST                /sales/{id}/pay-remaining
  GET                 /sales/{id}/payments
  GET                 /sales/stats/today
  POST                /sales/shopee/bulk-set-admin-fee

Usage in server.py:
    from routers.sales import make_router as _make_sales_router
    api_router.include_router(_make_sales_router(
        db=db, require_super_admin=require_super_admin, logger=logger,
        _insert_cash_transaction=_insert_cash_transaction,
        _resolve_payment_account=_resolve_payment_account,
        PAYMENT_ACCOUNT_MAP=PAYMENT_ACCOUNT_MAP,
        _size_tier=_size_tier,
        _compute_component_consumption=_compute_component_consumption,
        _sanitize_branch=_sanitize_branch,
        _company_info=_company_info,
    ))
"""
import io
import os
import re
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import HTMLResponse
from pydantic import BaseModel


# ---------- Module-level Pydantic Models ----------
class SaleItemIn(BaseModel):
    # MODE 1 (backward compat): pilih material langsung
    material_id: Optional[str] = None
    # MODE 2 (baru): pilih product dengan BOM
    product_id: Optional[str] = None
    product_name: str  # nama produk/jasa (mis. "Banner 3x2m", "Slayer")
    length_m: float = 0
    width_m: float = 0
    quantity: int = 1
    unit_price: float  # harga per m² (mode material) ATAU harga per unit (mode product fixed) ATAU per m² (product per_area)
    size: Optional[str] = None  # NEW: untuk produk yang has_sizes (S/M/L/XL/XXL/XXXL)


class SaleIn(BaseModel):
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    items: List[SaleItemIn] = []
    discount: float = 0
    cash_paid: float = 0
    payment_method: str = "cash"  # cash | transfer | shopee_plaza | shopee_kastem
    payment_bank: Optional[str] = None  # BCA | Mandiri (khusus transfer)
    payment_notes: Optional[str] = None  # keterangan tambahan (khusus transfer)
    shopee_admin_fee: float = 0  # biaya admin Shopee — dicatat sebagai pengeluaran terpisah, mengurangi netto omzet
    notes: Optional[str] = None
    branch: Optional[str] = None  # "plaza" | "kastem" — WAJIB untuk super_admin agar sale ter-tag cabang; kasir cabang otomatis pakai branch usernya.


# Mapping payment method → account_code untuk auto cash tx

class SaldoMasukIn(BaseModel):
    saldo_masuk: Optional[float] = None  # None = clear

class PayRemainingIn(BaseModel):
    amount: float
    payment_method: str = "cash"  # cash | transfer | shopee_plaza | shopee_kastem
    payment_bank: Optional[str] = None  # bca | mandiri (jika transfer)
    date: Optional[str] = None  # YYYY-MM-DD (opsional, default hari ini)
    notes: Optional[str] = None



def make_router(
    *,
    db,
    require_super_admin,
    logger,
    _insert_cash_transaction,
    _resolve_payment_account,
    PAYMENT_ACCOUNT_MAP,
    _size_tier,
    _compute_component_consumption,
    _sanitize_branch,
    _company_info,
):
    """Build the sales sub-router using injected dependencies."""
    router = APIRouter()

    async def _next_sale_no() -> str:
        today = datetime.now(timezone.utc).date()
        prefix = f"NOTA-{today.strftime('%Y%m%d')}-"
        count = await db.sales.count_documents({"sale_no": {"$regex": f"^{re.escape(prefix)}"}})
        return f"{prefix}{count + 1:04d}"


    @router.get("/sales")
    async def sales_list(
        user: dict = Depends(require_super_admin),
        limit: int = 200,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        page: int = 1,
        page_size: int = 10,
        q: Optional[str] = None,
        paginate: bool = False,
    ):
        """Return either plain list (backward compat) atau {items,total,page,page_size,pages} bila paginate=true."""
        query: Dict[str, Any] = {}
        if date_from or date_to:
            rng = {}
            if date_from:
                rng["$gte"] = date_from
            if date_to:
                rng["$lte"] = date_to
            query["date"] = rng
        if q and q.strip():
            safe = re.escape(q.strip())
            query["$or"] = [
                {"sale_no": {"$regex": safe, "$options": "i"}},
                {"customer_name": {"$regex": safe, "$options": "i"}},
                {"customer_phone": {"$regex": safe, "$options": "i"}},
            ]
        if paginate:
            page = max(1, int(page))
            page_size = max(1, min(int(page_size), 100))
            total = await db.sales.count_documents(query)
            skip = (page - 1) * page_size
            items = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(length=page_size)
            pages = (total + page_size - 1) // page_size if page_size else 0
            return {"items": items, "total": total, "page": page, "page_size": page_size, "pages": pages}
        # Backward-compat: plain list
        items = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=max(1, min(limit, 2000)))
        return items


    @router.get("/sales/{sale_id}")
    async def sales_get(sale_id: str, user: dict = Depends(require_super_admin)):
        s = await db.sales.find_one({"id": sale_id}, {"_id": 0})
        if not s:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        return s


    async def _build_and_persist_sale(
        payload: SaleIn,
        user: dict,
        *,
        sale_no: Optional[str] = None,
        sale_id: Optional[str] = None,
        created_at_iso: Optional[str] = None,
        date_iso: Optional[str] = None,
        is_update: bool = False,
    ) -> Dict[str, Any]:
        """Compute + validate + persist a sale (create atau update).
        Untuk update, pastikan rollback stok+cash sudah dilakukan SEBELUM memanggil helper ini."""
        if not payload.items:
            raise HTTPException(status_code=400, detail="Item transaksi tidak boleh kosong")
        items_out: List[Dict[str, Any]] = []
        subtotal = 0.0
        stock_deductions: Dict[str, float] = {}
        material_cache: Dict[str, Dict[str, Any]] = {}

        async def _get_mat(mid: str) -> Dict[str, Any]:
            if mid in material_cache:
                return material_cache[mid]
            m = await db.materials.find_one({"id": mid})
            if not m:
                raise HTTPException(status_code=400, detail="Bahan tidak ditemukan")
            material_cache[mid] = m
            return m

        for it in payload.items:
            if it.quantity <= 0:
                raise HTTPException(status_code=400, detail=f"Qty {it.product_name} harus > 0")
            if it.product_id:
                prod = await db.products.find_one({"id": it.product_id}, {"_id": 0})
                if not prod:
                    raise HTTPException(status_code=400, detail="Produk tidak ditemukan")
                components = prod.get("components") or []
                needs_dim = any(c.get("formula") in ("area", "length") for c in components)
                if needs_dim and (it.length_m <= 0 or it.width_m <= 0):
                    bad = any(c.get("formula") == "area" for c in components) and (it.length_m <= 0 or it.width_m <= 0)
                    bad_len = any(c.get("formula") == "length" for c in components) and it.length_m <= 0
                    if bad or bad_len:
                        raise HTTPException(status_code=400, detail=f"Ukuran P×L wajib diisi untuk {it.product_name}")
                has_sizes = bool(prod.get("has_sizes"))
                size_tier = "A"
                size_used = None
                if has_sizes:
                    if not it.size:
                        raise HTTPException(status_code=400, detail=f"Ukuran wajib dipilih untuk {it.product_name}")
                    available = prod.get("sizes") or []
                    if it.size not in available:
                        raise HTTPException(status_code=400, detail=f"Ukuran '{it.size}' tidak tersedia untuk {it.product_name}. Pilihan: {', '.join(available)}")
                    size_used = it.size
                    size_tier = _size_tier(it.size)
                pricing = prod.get("pricing_mode") or "fixed"
                if has_sizes:
                    unit_price_use = float(prod.get("price_size_b", 0) if size_tier == "B" else prod.get("price_size_a", 0))
                    if unit_price_use <= 0:
                        unit_price_use = float(prod.get("unit_price", 0))
                else:
                    unit_price_use = float(it.unit_price if it.unit_price > 0 else prod.get("unit_price", 0))
                if pricing == "per_area":
                    area_pc = float(it.length_m or 0) * float(it.width_m or 0)
                    area_total = round(area_pc * int(it.quantity), 4)
                    line_subtotal = round(area_total * unit_price_use, 2)
                else:
                    area_pc = float(it.length_m or 0) * float(it.width_m or 0)
                    area_total = round(area_pc * int(it.quantity), 4)
                    line_subtotal = round(unit_price_use * int(it.quantity), 2)
                item_components = []
                for c in components:
                    factor_use = float(c.get("quantity", 1) or 0)
                    if has_sizes and size_tier == "B":
                        qsb = c.get("quantity_size_b")
                        if qsb is not None:
                            factor_use = float(qsb or 0)
                    cons = _compute_component_consumption(
                        c["formula"], factor_use, it.length_m, it.width_m, it.quantity,
                    )
                    mat = await _get_mat(c["material_id"])
                    stock_deductions[c["material_id"]] = stock_deductions.get(c["material_id"], 0) + cons
                    item_components.append({
                        "material_id": c["material_id"],
                        "material_name": mat.get("name"),
                        "material_unit": mat.get("unit"),
                        "formula": c["formula"],
                        "factor": factor_use,
                        "consumption": cons,
                    })
                items_out.append({
                    "product_id": it.product_id,
                    "product_code": prod.get("code"),
                    "product_name": it.product_name or prod.get("name"),
                    "product_pricing_mode": pricing,
                    "length_m": float(it.length_m or 0),
                    "width_m": float(it.width_m or 0),
                    "quantity": int(it.quantity),
                    "area_per_pc": round(area_pc, 4),
                    "area_total": area_total,
                    "unit_price": unit_price_use,
                    "subtotal": line_subtotal,
                    "components": item_components,
                    "size": size_used,
                    "size_tier": size_tier if has_sizes else None,
                    "material_id": None,
                    "material_name": ", ".join(c["material_name"] for c in item_components) or "-",
                    "material_unit": item_components[0]["material_unit"] if item_components else "",
                })
                subtotal += line_subtotal
            else:
                if not it.material_id:
                    raise HTTPException(status_code=400, detail=f"{it.product_name}: pilih Produk atau Bahan")
                if it.length_m <= 0 or it.width_m <= 0:
                    raise HTTPException(status_code=400, detail=f"Ukuran P×L {it.product_name} harus > 0")
                mat = await _get_mat(it.material_id)
                area_per_pc = float(it.length_m) * float(it.width_m)
                area_total = round(area_per_pc * int(it.quantity), 4)
                line_subtotal = round(area_total * float(it.unit_price), 2)
                stock_deductions[it.material_id] = stock_deductions.get(it.material_id, 0) + area_total
                items_out.append({
                    "material_id": it.material_id,
                    "material_name": mat.get("name"),
                    "material_unit": mat.get("unit"),
                    "product_id": None,
                    "product_name": it.product_name,
                    "length_m": float(it.length_m),
                    "width_m": float(it.width_m),
                    "quantity": int(it.quantity),
                    "area_per_pc": round(area_per_pc, 4),
                    "area_total": area_total,
                    "unit_price": float(it.unit_price),
                    "subtotal": line_subtotal,
                    "components": [{
                        "material_id": it.material_id,
                        "material_name": mat.get("name"),
                        "material_unit": mat.get("unit"),
                        "formula": "area",
                        "factor": 1.0,
                        "consumption": area_total,
                    }],
                })
                subtotal += line_subtotal

        # Validasi stok
        for mid, total_needed in stock_deductions.items():
            mat = material_cache.get(mid) or await _get_mat(mid)
            current = float(mat.get("current_stock", 0))
            if total_needed > current + 1e-6:
                raise HTTPException(
                    status_code=400,
                    detail=f"Stok {mat.get('name')} tidak cukup (butuh {round(total_needed, 4)} {mat.get('unit')}, tersedia {round(current, 4)})",
                )

        discount = float(payload.discount or 0)
        total = round(subtotal - discount, 2)
        cash_paid = float(payload.cash_paid or 0)
        if cash_paid < 0:
            raise HTTPException(status_code=400, detail="Nominal diterima tidak boleh negatif")
        # DP support: jika cash_paid < total, sisanya jadi piutang (Sisa Tagihan)
        sisa_tagihan = round(max(0.0, total - cash_paid), 2)
        change = round(max(0.0, cash_paid - total), 2)
        payment_status = "dp" if sisa_tagihan > 0.01 else "paid"
        now = datetime.now(timezone.utc)
        final_sale_no = sale_no or await _next_sale_no()
        final_id = sale_id or str(uuid.uuid4())
        final_created = created_at_iso or now.isoformat()
        final_date = date_iso or now.date().isoformat()
        # Seed initial payment record (is_initial=True). Selalu ada meski cash_paid=0
        # (edge case: full DP tanpa DP di muka). Ini adalah entry pertama di history pembayaran.
        initial_payment = {
            "id": str(uuid.uuid4()),
            "amount": round(min(cash_paid, total), 2),  # exclude kembalian
            "payment_method": payload.payment_method or "cash",
            "payment_bank": (payload.payment_bank or "").strip() or None,
            "date": final_date,
            "notes": (payload.payment_notes or "").strip() or None,
            "is_initial": True,
            "created_at": final_created,
            "created_by": user.get("email"),
        }
        doc = {
            "id": final_id,
            "sale_no": final_sale_no,
            "date": final_date,
            "customer_name": (payload.customer_name or "Umum").strip() or "Umum",
            "customer_phone": (payload.customer_phone or "").strip(),
            "cashier": user.get("email"),
            "cashier_name": user.get("name") or user.get("email"),
            # Branch resolution: payload.branch (form) DIUTAMAKAN bila super_admin memilih;
            # fallback ke user.branch (cabang kasir) bila payload kosong.
            "branch": _sanitize_branch(payload.branch) or _sanitize_branch(user.get("branch")),
            "items": items_out,
            "subtotal": round(subtotal, 2),
            "discount": round(discount, 2),
            "total": total,
            "cash_paid": round(cash_paid, 2),
            "change": change,
            "sisa_tagihan": sisa_tagihan,
            "payment_method": payload.payment_method or "cash",
            "payment_bank": (payload.payment_bank or "").strip() or None,
            "payment_notes": (payload.payment_notes or "").strip() or None,
            "shopee_admin_fee": round(float(payload.shopee_admin_fee or 0), 2) if payload.payment_method in ("shopee_plaza", "shopee_kastem") else 0.0,
            "notes": payload.notes,
            "status": payment_status,  # "paid" (LUNAS) atau "dp"
            "payments": [initial_payment],  # unified payment history (DP + pelunasan)
            "created_at": final_created,
        }
        if is_update:
            doc["updated_at"] = now.isoformat()
            doc["updated_by"] = user.get("email")
            await db.sales.update_one({"id": final_id}, {"$set": doc})
        else:
            await db.sales.insert_one(doc)
        # Auto cash tx — akun tergantung metode pembayaran
        # Auto cash tx — akun tergantung metode pembayaran. Untuk DP, hanya cash_paid yg masuk.
        cash_recorded = round(min(cash_paid, total), 2)  # exclude kembalian dari kas
        try:
            acc_code, acc_label = _resolve_payment_account(payload.payment_method, payload.payment_bank)
            desc = f"Penjualan {final_sale_no} — {doc['customer_name']} · {acc_label}"
            if payment_status == "dp":
                desc += f" · DP (sisa Rp {sisa_tagihan:,.0f})"
            if doc.get("payment_notes"):
                desc += f" ({doc['payment_notes']})"
            # Untuk Shopee, cash tx pemasukan dicatat NETTO (gross - admin fee)
            shopee_admin_fee = float(doc.get("shopee_admin_fee") or 0)
            is_shopee = payload.payment_method in ("shopee_plaza", "shopee_kastem")
            if is_shopee and shopee_admin_fee > 0:
                netto_recorded = max(0.0, round(cash_recorded - shopee_admin_fee, 2))
                if netto_recorded > 0:
                    await _insert_cash_transaction(
                        account_code=acc_code,
                        description=desc + f" · − Admin Rp {shopee_admin_fee:,.0f}",
                        amount=netto_recorded,
                        reference=final_sale_no,
                        date_iso=doc["date"],
                        auto=True,
                        created_by=user.get("email"),
                    )
            else:
                if cash_recorded > 0:
                    await _insert_cash_transaction(
                        account_code=acc_code,
                        description=desc,
                        amount=cash_recorded,
                        reference=final_sale_no,
                        date_iso=doc["date"],
                        auto=True,
                        created_by=user.get("email"),
                    )
        except Exception as ex:
            logger.warning(f"Cashbook auto-insert (sale) failed: {ex}")
        # Apply stock deduction (net dari state saat ini)
        for mid, qty_used in stock_deductions.items():
            mat = material_cache.get(mid)
            if mat:
                new_stock = round(float(mat.get("current_stock", 0)) - float(qty_used), 4)
                await db.materials.update_one(
                    {"id": mid},
                    {"$set": {"current_stock": new_stock, "updated_at": now.isoformat()}},
                )
        doc.pop("_id", None)
        return doc


    @router.post("/sales")
    async def sales_create(payload: SaleIn, user: dict = Depends(require_super_admin)):
        return await _build_and_persist_sale(payload, user)


    @router.post("/sales/bulk-tag-branch")
    async def sales_bulk_tag_branch(
        branch: str,
        only_untagged: bool = True,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        user: dict = Depends(require_super_admin),
    ):
        """Bulk-tag branch (plaza/kastem) untuk sale yg belum bercabang (branch=null).
        Filter opsional: date_from, date_to.
        Query param `only_untagged=false` = timpa juga sale yg sudah bercabang.
        """
        b = _sanitize_branch(branch)
        if not b:
            raise HTTPException(status_code=400, detail="Branch harus 'plaza' atau 'kastem'")
        q: Dict[str, Any] = {}
        if only_untagged:
            q["$or"] = [{"branch": None}, {"branch": {"$exists": False}}, {"branch": ""}]
        if date_from or date_to:
            q["date"] = {}
            if date_from:
                q["date"]["$gte"] = date_from
            if date_to:
                q["date"]["$lte"] = date_to
        now_iso = datetime.now(timezone.utc).isoformat()
        res = await db.sales.update_many(q, {"$set": {"branch": b, "updated_at": now_iso}})
        return {"ok": True, "matched": res.matched_count, "modified": res.modified_count, "branch": b}


    async def _rollback_sale_effects(sale: Dict[str, Any]) -> None:
        """Rollback stock deduction dan hapus auto cash transaction untuk sale ini."""
        rollback: Dict[str, float] = {}
        for it in sale.get("items") or []:
            comps = it.get("components")
            if comps:
                for c in comps:
                    mid = c.get("material_id")
                    if mid:
                        rollback[mid] = rollback.get(mid, 0) + float(c.get("consumption", 0))
            else:
                mid = it.get("material_id")
                if mid:
                    rollback[mid] = rollback.get(mid, 0) + float(it.get("area_total", 0))
        now_iso = datetime.now(timezone.utc).isoformat()
        for mid, qty in rollback.items():
            mat = await db.materials.find_one({"id": mid})
            if mat:
                new_stock = round(float(mat.get("current_stock", 0)) + float(qty), 4)
                await db.materials.update_one(
                    {"id": mid},
                    {"$set": {"current_stock": new_stock, "updated_at": now_iso}},
                )
        try:
            # Hapus semua auto cash tx untuk sale ini (semua account_code payment method)
            payment_codes = list(PAYMENT_ACCOUNT_MAP.values())
            await db.cash_transactions.delete_many({
                "reference": sale.get("sale_no"),
                "auto": True,
                "account_code": {"$in": payment_codes},
            })
        except Exception as ex:
            logger.warning(f"Cashbook rollback (sale) failed: {ex}")


    @router.put("/sales/{sale_id}")
    async def sales_update(sale_id: str, payload: SaleIn, user: dict = Depends(require_super_admin)):
        existing = await db.sales.find_one({"id": sale_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        # Preserve pelunasan (non-initial) payments SEBELUM rollback,
        # supaya edit item invoice tidak menghapus riwayat pelunasan kas.
        existing_payments = existing.get("payments") or []
        pelunasan_payments = [p for p in existing_payments if not p.get("is_initial")]
        # 1. Rollback dulu (stok + cash tx)
        await _rollback_sale_effects(existing)
        # 2. Recompute + apply — preserve sale_no, id, created_at, date
        try:
            result = await _build_and_persist_sale(
                payload, user,
                sale_no=existing.get("sale_no"),
                sale_id=existing.get("id"),
                created_at_iso=existing.get("created_at"),
                date_iso=existing.get("date"),
                is_update=True,
            )
        except HTTPException:
            # Reapply original stock deduction & cash tx supaya state konsisten
            try:
                # Deduct kembali stok
                rededuct: Dict[str, float] = {}
                for it in existing.get("items") or []:
                    comps = it.get("components") or []
                    if comps:
                        for c in comps:
                            mid = c.get("material_id")
                            if mid:
                                rededuct[mid] = rededuct.get(mid, 0) + float(c.get("consumption", 0))
                    else:
                        mid = it.get("material_id")
                        if mid:
                            rededuct[mid] = rededuct.get(mid, 0) + float(it.get("area_total", 0))
                now_iso = datetime.now(timezone.utc).isoformat()
                for mid, qty in rededuct.items():
                    mat = await db.materials.find_one({"id": mid})
                    if mat:
                        new_stock = round(float(mat.get("current_stock", 0)) - float(qty), 4)
                        await db.materials.update_one({"id": mid}, {"$set": {"current_stock": new_stock, "updated_at": now_iso}})
                # Reinsert cash tx — pakai payment_method untuk resolve akun (cash → 101, transfer → 301-BCA, dll)
                _ex_acc_code, _ex_acc_label = _resolve_payment_account(
                    existing.get("payment_method") or "cash",
                    existing.get("payment_bank"),
                )
                await _insert_cash_transaction(
                    account_code=_ex_acc_code,
                    description=f"Penjualan {existing.get('sale_no')} — {existing.get('customer_name')} · {_ex_acc_label}",
                    amount=float(existing.get("total", 0)),
                    reference=existing.get("sale_no"),
                    date_iso=existing.get("date"),
                    auto=True,
                    created_by=user.get("email"),
                )
                # Restore pelunasan cash tx & payments juga bila ada
                for p in pelunasan_payments:
                    try:
                        acc_code, acc_label = _resolve_payment_account(p.get("payment_method") or "cash", p.get("payment_bank"))
                        await _insert_cash_transaction(
                            account_code=acc_code,
                            description=f"Pelunasan {existing.get('sale_no')} — {existing.get('customer_name')} · {acc_label}",
                            amount=float(p.get("amount", 0)),
                            reference=existing.get("sale_no"),
                            date_iso=p.get("date"),
                            auto=True,
                            created_by=p.get("created_by") or user.get("email"),
                        )
                    except Exception:
                        pass
            except Exception as ex:
                logger.error(f"Rollback restore failed after update error: {ex}")
            raise

        # 3. Setelah update berhasil, RESTORE pelunasan payments & re-insert cash tx pelunasan
        if pelunasan_payments:
            total = float(result.get("total") or 0)
            initial_amt = float((result.get("payments") or [{}])[0].get("amount", 0))
            pelunasan_total = round(sum(float(p.get("amount", 0)) for p in pelunasan_payments), 2)
            # Cap agar total pembayaran tidak melebihi total invoice baru
            max_pelunasan_allowed = round(max(0.0, total - initial_amt), 2)
            if pelunasan_total > max_pelunasan_allowed + 0.01:
                logger.warning(
                    f"Pelunasan history ({pelunasan_total}) melebihi sisa dari total baru ({max_pelunasan_allowed}) "
                    f"utk sale {existing.get('sale_no')}. Menyimpan apa adanya tanpa cap; user perlu review."
                )
            new_cash_paid = round(initial_amt + pelunasan_total, 2)
            new_sisa = round(max(0.0, total - new_cash_paid), 2)
            new_status = "paid" if new_sisa <= 0.01 else "dp"
            now_iso = datetime.now(timezone.utc).isoformat()
            await db.sales.update_one(
                {"id": sale_id},
                {
                    "$set": {
                        "cash_paid": new_cash_paid,
                        "sisa_tagihan": new_sisa,
                        "status": new_status,
                    },
                    "$push": {"payments": {"$each": pelunasan_payments}},
                },
            )
            # Re-insert cash tx pelunasan (yang tadi dihapus rollback)
            for p in pelunasan_payments:
                try:
                    acc_code, acc_label = _resolve_payment_account(p.get("payment_method") or "cash", p.get("payment_bank"))
                    desc = f"Pelunasan {existing.get('sale_no')} — {existing.get('customer_name')} · {acc_label}"
                    if p.get("notes"):
                        desc += f" ({p['notes']})"
                    await _insert_cash_transaction(
                        account_code=acc_code,
                        description=desc,
                        amount=float(p.get("amount", 0)),
                        reference=existing.get("sale_no"),
                        date_iso=p.get("date") or now_iso[:10],
                        auto=True,
                        created_by=p.get("created_by") or user.get("email"),
                    )
                except Exception as ex:
                    logger.warning(f"Failed to re-insert pelunasan cash tx: {ex}")
            # Update result payload utk return
            result["payments"] = (result.get("payments") or []) + list(pelunasan_payments)
            result["cash_paid"] = new_cash_paid
            result["sisa_tagihan"] = new_sisa
            result["status"] = new_status
        return result


    @router.delete("/sales/{sale_id}")
    async def sales_delete(sale_id: str, user: dict = Depends(require_super_admin)):
        s = await db.sales.find_one({"id": sale_id})
        if not s:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        # Rollback stok + hapus semua auto cash tx (pakai helper konsisten dgn sales_update)
        await _rollback_sale_effects(s)
        await db.sales.delete_one({"id": sale_id})
        return {"ok": True}


    @router.get("/sales/{sale_id}/receipt", response_class=HTMLResponse)
    async def sales_receipt_html(sale_id: str, user: dict = Depends(require_super_admin)):
        s = await db.sales.find_one({"id": sale_id}, {"_id": 0})
        if not s:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        ci = _company_info()
        def _idr(n):
            return f"Rp {float(n or 0):,.0f}".replace(",", ".")
        def _num(n):
            return f"{float(n or 0):.4f}".rstrip("0").rstrip(".") or "0"
        items_rows = ""
        for it in s.get("items") or []:
            pricing_mode = it.get("product_pricing_mode")
            is_product_fixed = it.get("product_id") and pricing_mode == "fixed"
            # Detail dimensi/qty
            if is_product_fixed:
                dim_line = f"<span>{int(it.get('quantity', 1))} pcs</span><span>@ {_idr(it.get('unit_price'))}</span>"
            else:
                dim_line = (
                    f"<span>{_num(it.get('length_m'))}m × {_num(it.get('width_m'))}m × {int(it.get('quantity', 1))}</span>"
                    f"<span>= {_num(it.get('area_total'))}m²</span>"
                )
                price_line = f"<span>@ {_idr(it.get('unit_price'))}/m²</span><span class='strong'>{_idr(it.get('subtotal'))}</span>"
            # Bahan breakdown (untuk BOM)
            mat_line = ""
            comps = it.get("components") or []
            if len(comps) > 1:
                mat_bits = " + ".join(f"{c.get('material_name', '-')} {_num(c.get('consumption'))}{c.get('material_unit', '')}" for c in comps)
                mat_line = f'<div class="mat">Bahan: {mat_bits}</div>'
            elif it.get("material_name"):
                mat_line = f'<div class="mat">{it.get("material_name")}</div>'

            if is_product_fixed:
                items_rows += f"""
            <div class="item">
              <div class="prod">{it.get('product_name', '')}</div>
              {mat_line}
              <div class="row">{dim_line}</div>
              <div class="row"><span></span><span class="strong">{_idr(it.get('subtotal'))}</span></div>
            </div>
            """
            else:
                items_rows += f"""
            <div class="item">
              <div class="prod">{it.get('product_name', '')}</div>
              {mat_line}
              <div class="row">{dim_line}</div>
              <div class="row">{price_line}</div>
            </div>
            """
        discount_row = ""
        if float(s.get("discount", 0)) > 0:
            discount_row = f"""<div class="row"><span>Diskon</span><span>- {_idr(s['discount'])}</span></div>"""
        date_str = s.get("date", "")
        created = s.get("created_at", "")[:19].replace("T", " ")
        customer_phone_row = f'<div class="line-sm">Telp: {s.get("customer_phone")}</div>' if s.get("customer_phone") else ""
        notes_row = f'<div class="notes">Catatan: {s.get("notes")}</div>' if s.get("notes") else ""
        sisa = float(s.get("sisa_tagihan") or 0)
        status = s.get("status") or ("dp" if sisa > 0.01 else "paid")
        sisa_row = f'<div class="row" style="color:#E81123;font-weight:bold;"><span>SISA TAGIHAN</span><span>{_idr(sisa)}</span></div>' if sisa > 0.01 else ""
        if status == "dp":
            status_row = '<div class="row" style="background:#FEF3C7;padding:4px 6px;margin-top:2px;font-weight:bold;color:#92400E;text-align:center;justify-content:center;">DP · Belum Lunas</div>'
        else:
            status_row = '<div class="row" style="background:#DCFCE7;padding:4px 6px;margin-top:2px;font-weight:bold;color:#166534;text-align:center;justify-content:center;">LUNAS</div>'
        html = f"""<!DOCTYPE html>
    <html lang="id"><head><meta charset="UTF-8"><title>Nota {s.get('sale_no')}</title>
    <style>
      /* ===== Thermal 80mm (C80BT) - printable area ~72mm =====
         Semua teks WAJIB bold + hitam pekat agar tidak pudar
         saat dibakar oleh head printer thermal. Font sans-serif
         (Arial) lebih tebal & terbaca dibanding Courier. */
      @page {{ size: 80mm auto; margin: 0; }}
      * {{ box-sizing: border-box; }}
      html, body {{ margin: 0; padding: 0; }}
      body {{ font-family: Arial, Helvetica, 'Liberation Sans', sans-serif; font-size: 13px; font-weight: 700; line-height: 1.35; color: #000; background: #eee; -webkit-font-smoothing: none; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
      .receipt {{ width: 72mm; max-width: 72mm; padding: 3mm 3mm 4mm 3mm; background: white; margin: 8px auto; box-shadow: 0 1px 6px rgba(0,0,0,0.08); word-wrap: break-word; overflow-wrap: break-word; color: #000; font-weight: 700; }}
      .receipt, .receipt * {{ color: #000 !important; }}
      h1, h2, h3, p {{ margin: 0; padding: 0; }}
      .center {{ text-align: center; }}
      .strong {{ font-weight: 800; }}
      .sep {{ border-top: 1px dashed #000; margin: 4px 0; }}
      .header {{ text-align: center; padding-bottom: 5px; border-bottom: 1px dashed #000; }}
      .header .name {{ font-size: 16px; font-weight: 900; letter-spacing: 0.3px; word-break: break-word; }}
      .header .addr {{ font-size: 12px; font-weight: 700; margin-top: 3px; line-height: 1.35; word-break: break-word; }}
      .meta {{ padding: 4px 0; border-bottom: 1px dashed #000; font-size: 12px; font-weight: 700; }}
      .meta .row {{ display: flex; justify-content: space-between; gap: 4px; padding: 1px 0; }}
      .meta .row > span:last-child {{ text-align: right; word-break: break-all; }}
      .items {{ padding: 4px 0; border-bottom: 1px dashed #000; }}
      .item {{ padding: 4px 0; }}
      .item + .item {{ border-top: 1px dashed #000; }}
      .item .prod {{ font-weight: 900; font-size: 13px; word-break: break-word; }}
      .item .mat {{ font-size: 11px; font-weight: 700; word-break: break-word; }}
      .item .row {{ display: flex; justify-content: space-between; font-size: 12px; font-weight: 700; margin-top: 2px; gap: 4px; }}
      .item .row > span:last-child {{ text-align: right; white-space: nowrap; }}
      .totals {{ padding: 4px 0; border-bottom: 1px dashed #000; font-size: 13px; font-weight: 700; }}
      .totals .row {{ display: flex; justify-content: space-between; padding: 2px 0; }}
      .totals .grand {{ font-size: 16px; font-weight: 900; padding: 4px 0; border-top: 2px solid #000; margin-top: 3px; }}
      .pay {{ padding: 4px 0; border-bottom: 1px dashed #000; font-size: 13px; font-weight: 700; }}
      .pay .row {{ display: flex; justify-content: space-between; padding: 2px 0; }}
      .footer {{ padding-top: 8px; text-align: center; font-size: 11px; font-weight: 700; line-height: 1.4; }}
      .notes {{ font-size: 11px; font-weight: 700; padding: 4px 0; border-bottom: 1px dashed #000; font-style: italic; word-break: break-word; }}
      .toolbar {{ max-width: 72mm; margin: 0 auto 8px; text-align: center; padding-top: 10px; }}
      .toolbar button {{ background: #002FA7; color: white; border: 0; padding: 10px 22px; font-family: Arial, sans-serif; font-size: 12px; font-weight: 700; letter-spacing: 0.6px; cursor: pointer; text-transform: uppercase; }}
      .toolbar button:hover {{ background: #002080; }}
      .toolbar .hint {{ font-size: 11px; color: #333; margin-top: 6px; font-family: Arial, sans-serif; }}
      @media print {{
        html, body {{ background: white; width: 80mm; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
        .receipt {{ margin: 0 auto; box-shadow: none; padding: 2mm 3mm 3mm 3mm; width: 72mm; }}
        .toolbar {{ display: none; }}
      }}
    </style></head><body>
    <div class="toolbar">
      <button onclick="window.print()">🖨 Cetak Nota</button>
      <div class="hint">Thermal 80mm • Margin: None • Skala: 100% • Aktifkan "Background graphics"</div>
    </div>
    <div class="receipt">
      <div class="header">
        <div class="name">{ci['name'].upper()}</div>
        <div class="addr">{ci['address']}<br>HP : {ci['phone']}</div>
      </div>
      <div class="meta">
        <div class="row"><span>No. Nota</span><span class="strong">{s.get('sale_no', '')}</span></div>
        <div class="row"><span>Tanggal</span><span>{created}</span></div>
        <div class="row"><span>Kasir</span><span>{s.get('cashier_name', '')}</span></div>
        <div class="row"><span>Pelanggan</span><span>{s.get('customer_name', 'Umum')}</span></div>
        {customer_phone_row}
      </div>
      <div class="items">
        {items_rows}
      </div>
      <div class="totals">
        <div class="row"><span>Subtotal</span><span>{_idr(s.get('subtotal'))}</span></div>
        {discount_row}
        <div class="row grand"><span>TOTAL</span><span>{_idr(s.get('total'))}</span></div>
      </div>
      <div class="pay">
        <div class="row"><span>Metode</span><span class="strong">{(s.get('payment_method') or 'tunai').upper()}</span></div>
        <div class="row"><span>Bayar</span><span>{_idr(s.get('cash_paid'))}</span></div>
        {sisa_row}
        {status_row}
        <div class="row strong"><span>Kembali</span><span>{_idr(s.get('change'))}</span></div>
      </div>
      {notes_row}
      <div class="footer">
        Terima kasih atas kunjungan Anda<br>
        <span style="font-size:9px;">Simpan struk ini sebagai bukti pembayaran.</span>
      </div>
    </div>
    <script>
      // Auto-focus print dialog jika ada query ?auto=1
      if (new URLSearchParams(location.search).get('auto') === '1') {{
        setTimeout(() => window.print(), 400);
      }}
    </script>
    </body></html>"""
        return HTMLResponse(content=html)


    @router.get("/sales/{sale_id}/invoice-pdf")
    async def sales_invoice_pdf(sale_id: str, user: dict = Depends(require_super_admin)):
        """Cetak Nota A4 profesional (untuk customer korporat)."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

        s = await db.sales.find_one({"id": sale_id}, {"_id": 0})
        if not s:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        ci = _company_info()

        def _idr(n):
            return f"Rp {float(n or 0):,.0f}".replace(",", ".")

        def _num(n):
            return f"{float(n or 0):.4f}".rstrip("0").rstrip(".") or "0"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=18 * mm, rightMargin=18 * mm,
            topMargin=15 * mm, bottomMargin=15 * mm,
            title=f"Nota {s.get('sale_no')}",
        )
        styles = getSampleStyleSheet()
        story = []

        # Company header
        company_style = ParagraphStyle("company", parent=styles["Normal"], fontSize=16, textColor=colors.HexColor("#002FA7"),
                                       alignment=TA_LEFT, spaceAfter=2, leading=18, fontName="Helvetica-Bold")
        story.append(Paragraph(ci["name"].upper(), company_style))
        story.append(Paragraph(f"{ci['address']}<br/>HP: {ci['phone']}", ParagraphStyle("addr", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#333333"), leading=12)))
        story.append(Spacer(1, 8 * mm))

        # Title
        story.append(Paragraph("<b>NOTA PENJUALAN</b>", ParagraphStyle("title", parent=styles["Normal"], fontSize=14, alignment=TA_CENTER, spaceAfter=6 * mm, fontName="Helvetica-Bold")))

        # Meta table (2 kolom: kiri = No/Tgl, kanan = Pelanggan)
        created = s.get("created_at", "")[:19].replace("T", " ")
        meta_data = [
            [Paragraph("<b>No. Nota</b>", styles["Normal"]), Paragraph(str(s.get("sale_no", "")), styles["Normal"]),
             Paragraph("<b>Pelanggan</b>", styles["Normal"]), Paragraph(str(s.get("customer_name", "Umum")), styles["Normal"])],
            [Paragraph("<b>Tanggal</b>", styles["Normal"]), Paragraph(created, styles["Normal"]),
             Paragraph("<b>Telp</b>", styles["Normal"]), Paragraph(str(s.get("customer_phone", "-") or "-"), styles["Normal"])],
            [Paragraph("<b>Kasir</b>", styles["Normal"]), Paragraph(str(s.get("cashier_name", "")), styles["Normal"]),
             Paragraph("<b>Metode</b>", styles["Normal"]), Paragraph((s.get("payment_method") or "tunai").upper(), styles["Normal"])],
        ]
        meta_tbl = Table(meta_data, colWidths=[28 * mm, 55 * mm, 25 * mm, 66 * mm])
        meta_tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#eeeeee")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 6 * mm))

        # Items table
        right = ParagraphStyle("r", parent=styles["Normal"], alignment=TA_RIGHT)
        header_row = [
            Paragraph("<b>No</b>", ParagraphStyle("hn", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.white)),
            Paragraph("<b>Deskripsi</b>", ParagraphStyle("hd", parent=styles["Normal"], textColor=colors.white)),
            Paragraph("<b>Qty / Dim</b>", ParagraphStyle("hq", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.white)),
            Paragraph("<b>Harga</b>", ParagraphStyle("hp", parent=styles["Normal"], alignment=TA_RIGHT, textColor=colors.white)),
            Paragraph("<b>Subtotal</b>", ParagraphStyle("hs", parent=styles["Normal"], alignment=TA_RIGHT, textColor=colors.white)),
        ]
        rows = [header_row]
        for idx, it in enumerate(s.get("items") or [], 1):
            pricing_mode = it.get("product_pricing_mode")
            is_fixed = it.get("product_id") and pricing_mode == "fixed"
            name = it.get("product_name") or it.get("material_name") or "-"
            # Bahan breakdown
            comps = it.get("components") or []
            if len(comps) > 1:
                mat_bits = " + ".join(f"{c.get('material_name', '-')} {_num(c.get('consumption'))}{c.get('material_unit', '')}" for c in comps)
                desc = f"{name}<br/><font size=7 color='#666'>Bahan: {mat_bits}</font>"
            elif it.get("material_name") and not it.get("product_name"):
                desc = name
            elif comps:
                c = comps[0]
                desc = f"{name}<br/><font size=7 color='#666'>{c.get('material_name', '')}</font>"
            else:
                desc = name
            if is_fixed:
                qty_dim = f"{int(it.get('quantity', 1))} pcs"
                harga = f"{_idr(it.get('unit_price'))}"
            else:
                qty_dim = f"{_num(it.get('length_m'))}m × {_num(it.get('width_m'))}m × {int(it.get('quantity', 1))}<br/><font size=7 color='#666'>= {_num(it.get('area_total'))} m²</font>"
                harga = f"{_idr(it.get('unit_price'))}/m²"
            rows.append([
                Paragraph(str(idx), ParagraphStyle("n", parent=styles["Normal"], alignment=TA_CENTER)),
                Paragraph(desc, styles["Normal"]),
                Paragraph(qty_dim, ParagraphStyle("q", parent=styles["Normal"], alignment=TA_CENTER)),
                Paragraph(harga, right),
                Paragraph(_idr(it.get("subtotal")), right),
            ])
        items_tbl = Table(rows, colWidths=[10 * mm, 65 * mm, 34 * mm, 30 * mm, 35 * mm], repeatRows=1)
        items_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
        ]))
        story.append(items_tbl)
        story.append(Spacer(1, 4 * mm))

        # Totals (kanan)
        total_rows = [
            [Paragraph("Subtotal", right), Paragraph(_idr(s.get("subtotal")), right)],
        ]
        if float(s.get("discount", 0)) > 0:
            total_rows.append([Paragraph("Diskon", right), Paragraph(f"- {_idr(s.get('discount'))}", right)])
        total_rows.append([
            Paragraph("<b><font size=12>TOTAL</font></b>", right),
            Paragraph(f"<b><font size=12 color='#002FA7'>{_idr(s.get('total'))}</font></b>", right),
        ])
        total_rows.append([Paragraph("Bayar (Tunai)", right), Paragraph(_idr(s.get("cash_paid")), right)])
        _sisa = float(s.get("sisa_tagihan") or 0)
        _status = s.get("status") or ("dp" if _sisa > 0.01 else "paid")
        if _sisa > 0.01:
            total_rows.append([
                Paragraph("<b><font color='#E81123'>SISA TAGIHAN</font></b>", right),
                Paragraph(f"<b><font color='#E81123'>{_idr(_sisa)}</font></b>", right),
            ])
        total_rows.append([Paragraph("Kembali", right), Paragraph(_idr(s.get("change")), right)])
        # Status badge
        if _status == "dp":
            total_rows.append([
                Paragraph("", right),
                Paragraph("<b><font color='#92400E' backcolor='#FEF3C7'>&nbsp;&nbsp;DP · BELUM LUNAS&nbsp;&nbsp;</font></b>", right),
            ])
        else:
            total_rows.append([
                Paragraph("", right),
                Paragraph("<b><font color='#166534' backcolor='#DCFCE7'>&nbsp;&nbsp;LUNAS&nbsp;&nbsp;</font></b>", right),
            ])
        total_tbl = Table(total_rows, colWidths=[100 * mm, 55 * mm], hAlign="RIGHT")
        total_tbl.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEABOVE", (0, -3), (-1, -3), 1.5, colors.HexColor("#002FA7")),
            ("LINEBELOW", (0, -3), (-1, -3), 0.5, colors.HexColor("#cccccc")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(total_tbl)

        if s.get("notes"):
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph(f"<b>Catatan:</b> {s.get('notes')}", ParagraphStyle("notes", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#555"))))

        # Footer signatures
        story.append(Spacer(1, 15 * mm))
        sign_rows = [[
            Paragraph("<b>Pelanggan</b><br/><br/><br/><br/><br/>(_______________________)", ParagraphStyle("sc", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9)),
            Paragraph("", styles["Normal"]),
            Paragraph(f"<b>Hormat kami</b><br/>{ci['name']}<br/><br/><br/><br/>(_______________________)", ParagraphStyle("ss", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9)),
        ]]
        sign_tbl = Table(sign_rows, colWidths=[60 * mm, 30 * mm, 65 * mm])
        story.append(sign_tbl)

        doc.build(story)
        buf.seek(0)
        fname = f"Nota_{s.get('sale_no', 'penjualan').replace('/', '_')}.pdf"
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={fname}"},
        )


    @router.get("/sales/report/pdf")
    async def sales_report_pdf(
        user: dict = Depends(require_super_admin),
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        month: Optional[str] = None,  # YYYY-MM (opsional, override date_from/to)
    ):
        """Laporan Penjualan PDF landscape untuk periode tertentu."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER

        # Tentukan range
        if month:
            try:
                year, m = month.split("-")
                date_from = f"{year}-{int(m):02d}-01"
                if int(m) == 12:
                    date_to = f"{int(year)+1}-01-01"
                else:
                    date_to = f"{year}-{int(m)+1:02d}-01"
            except Exception:
                raise HTTPException(status_code=400, detail="Format month salah, gunakan YYYY-MM")

        q = {}
        if date_from or date_to:
            q["date"] = {}
            if date_from:
                q["date"]["$gte"] = date_from
            if date_to:
                q["date"]["$lt"] = date_to
        items = await db.sales.find(q, {"_id": 0}).sort("created_at", 1).to_list(length=20000)
        ci = _company_info()

        def _idr(n):
            return f"Rp {float(n or 0):,.0f}".replace(",", ".")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=12 * mm, rightMargin=12 * mm,
            topMargin=12 * mm, bottomMargin=12 * mm,
            title=f"Laporan Penjualan {date_from or ''} - {date_to or ''}",
        )
        styles = getSampleStyleSheet()
        story = []

        # Header
        story.append(Paragraph(ci["name"].upper(), ParagraphStyle("co", parent=styles["Normal"], fontSize=13, textColor=colors.HexColor("#002FA7"), fontName="Helvetica-Bold")))
        story.append(Paragraph(f"{ci['address']} · HP: {ci['phone']}", ParagraphStyle("addr", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#555"))))
        story.append(Spacer(1, 4 * mm))

        period_label = f"{date_from or '(awal)'} s/d {date_to or '(sekarang)'}"
        if month:
            period_label = f"Bulan {month}"
        story.append(Paragraph(f"<b>LAPORAN PENJUALAN</b>", ParagraphStyle("t", parent=styles["Normal"], fontSize=14, alignment=TA_CENTER, fontName="Helvetica-Bold")))
        story.append(Paragraph(f"Periode: <b>{period_label}</b> · Total transaksi: <b>{len(items)}</b>", ParagraphStyle("p", parent=styles["Normal"], fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor("#555"))))
        story.append(Spacer(1, 5 * mm))

        right = ParagraphStyle("r", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=9)
        center = ParagraphStyle("c", parent=styles["Normal"], alignment=TA_CENTER, fontSize=9)
        normal_sm = ParagraphStyle("ns", parent=styles["Normal"], fontSize=9)

        header_row = [
            Paragraph("<b>No</b>", ParagraphStyle("hn", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.white)),
            Paragraph("<b>Tanggal</b>", ParagraphStyle("hd", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.white)),
            Paragraph("<b>No. Nota</b>", ParagraphStyle("hno", parent=styles["Normal"], textColor=colors.white)),
            Paragraph("<b>Pelanggan</b>", ParagraphStyle("hc", parent=styles["Normal"], textColor=colors.white)),
            Paragraph("<b>Kasir</b>", ParagraphStyle("hk", parent=styles["Normal"], textColor=colors.white)),
            Paragraph("<b>Item</b>", ParagraphStyle("hi", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.white)),
            Paragraph("<b>Subtotal</b>", ParagraphStyle("hs", parent=styles["Normal"], alignment=TA_RIGHT, textColor=colors.white)),
            Paragraph("<b>Diskon</b>", ParagraphStyle("hd2", parent=styles["Normal"], alignment=TA_RIGHT, textColor=colors.white)),
            Paragraph("<b>Total</b>", ParagraphStyle("ht", parent=styles["Normal"], alignment=TA_RIGHT, textColor=colors.white)),
        ]
        rows = [header_row]
        total_subtotal = 0.0
        total_discount = 0.0
        total_grand = 0.0
        for idx, s in enumerate(items, 1):
            item_count = len(s.get("items") or [])
            rows.append([
                Paragraph(str(idx), center),
                Paragraph(str(s.get("date", "")), center),
                Paragraph(str(s.get("sale_no", "")), normal_sm),
                Paragraph(str(s.get("customer_name", "Umum") or "Umum")[:35], normal_sm),
                Paragraph(str(s.get("cashier_name", ""))[:20], normal_sm),
                Paragraph(f"{item_count}", center),
                Paragraph(_idr(s.get("subtotal")), right),
                Paragraph(_idr(s.get("discount")), right),
                Paragraph(_idr(s.get("total")), right),
            ])
            total_subtotal += float(s.get("subtotal", 0) or 0)
            total_discount += float(s.get("discount", 0) or 0)
            total_grand += float(s.get("total", 0) or 0)
        # Total row
        rows.append([
            Paragraph(""),
            Paragraph(""),
            Paragraph(""),
            Paragraph(""),
            Paragraph(""),
            Paragraph("<b>TOTAL</b>", ParagraphStyle("tt", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=10, fontName="Helvetica-Bold")),
            Paragraph(f"<b>{_idr(total_subtotal)}</b>", ParagraphStyle("ts", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=10, fontName="Helvetica-Bold")),
            Paragraph(f"<b>{_idr(total_discount)}</b>", ParagraphStyle("td", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=10, fontName="Helvetica-Bold")),
            Paragraph(f"<b>{_idr(total_grand)}</b>", ParagraphStyle("tg", parent=styles["Normal"], alignment=TA_RIGHT, fontSize=10, fontName="Helvetica-Bold", textColor=colors.HexColor("#002FA7"))),
        ])

        tbl = Table(rows, colWidths=[10 * mm, 22 * mm, 30 * mm, 55 * mm, 30 * mm, 14 * mm, 32 * mm, 30 * mm, 35 * mm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8f8f8")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e8ecf7")),
            ("LINEABOVE", (0, -1), (-1, -1), 1.5, colors.HexColor("#002FA7")),
        ]))
        story.append(tbl)

        if not items:
            story.append(Spacer(1, 10 * mm))
            story.append(Paragraph("<i>Belum ada transaksi pada periode ini.</i>", ParagraphStyle("empty", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.HexColor("#999"))))

        doc.build(story)
        buf.seek(0)
        fname_period = month or (f"{date_from}_sd_{date_to}" if (date_from or date_to) else "semua")
        fname = f"Laporan_Penjualan_{fname_period}.pdf"
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"inline; filename={fname}"},
        )


    @router.get("/sales/report/excel")
    async def sales_report_excel(
        user: dict = Depends(require_super_admin),
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        customer: Optional[str] = None,
        month: Optional[str] = None,
        hidden_pay_cols: Optional[str] = None,
    ):
        """Laporan Penjualan Excel — format persis seperti tabel Excel-style di UI.
        12 kolom utama + 8 grup pembayaran (Cash/BCA/Mandiri × Plaza/Kastem + Shopee × Plaza/Kastem)
        masing-masing Nominal + Tanggal.

        Kolom "Total (Uang Diterima)" = jumlah uang aktual diterima pada baris tsb
        (Initial DP untuk baris pertama sale, atau amount pelunasan untuk baris pelunasan).
        SUM kolom ini otomatis SAMA dengan angka Omzet (Uang Diterima) di dashboard.

        `hidden_pay_cols`: comma-separated list dari key kolom pembayaran yang di-hide di UI.
        Baris (sale/pelunasan) yang payment_col-nya termasuk hidden AKAN DILEWATI—tidak dihitung
        ke total apa pun, agar Excel sinkron dengan dashboard yang di-filter.
        """
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if month:
            try:
                year, m = month.split("-")
                date_from = f"{year}-{int(m):02d}-01"
                if int(m) == 12:
                    date_to = f"{int(year)+1}-01-01"
                else:
                    date_to = f"{year}-{int(m)+1:02d}-01"
            except Exception:
                raise HTTPException(status_code=400, detail="Format month salah, gunakan YYYY-MM")

        q: Dict[str, Any] = {"status": {"$nin": ["cancelled", "void", "voided", "canceled"]}}
        if date_from or date_to:
            q["date"] = {}
            if date_from:
                q["date"]["$gte"] = date_from
            if date_to:
                q["date"]["$lte"] = date_to
        if customer:
            safe = re.escape(customer.strip())
            q["customer_name"] = {"$regex": safe, "$options": "i"}
        sales = await db.sales.find(q, {"_id": 0}).sort("created_at", 1).to_list(length=20000)
        ci = _company_info()

        # Preload maps (sama seperti analytics endpoint)
        customers = await db.customers.find({}, {"_id": 0, "name": 1, "address": 1}).to_list(length=5000)
        cust_addr_map = {(c.get("name") or "").strip().lower(): (c.get("address") or "").strip() for c in customers}
        products_p = await db.products.find({}, {"_id": 0, "name": 1, "length_meter": 1}).to_list(length=5000)
        prod_length_map = {(p.get("name") or "").strip().lower(): float(p.get("length_meter") or 0) for p in products_p}

        PAY_COLS = [
            ("cash_plaza", "Cash Plaza"),
            ("cash_kastem", "Cash Kastem"),
            ("bca_plaza", "BCA Plaza"),
            ("bca_kastem", "BCA Kastem"),
            ("mandiri_plaza", "Mandiri Plaza"),
            ("mandiri_kastem", "Mandiri Kastem"),
            ("shopee_plaza", "Shopee Plaza"),
            ("shopee_kastem", "Shopee Kastem"),
        ]
        _all_pay_keys = {k for k, _ in PAY_COLS}
        hidden_set = set(x.strip() for x in (hidden_pay_cols or "").split(",") if x.strip())
        hidden_set &= _all_pay_keys  # sanitize, ignore unknown keys
        is_pay_filtered = bool(hidden_set)
        # Ringkasan per tab pembayaran: {key: {"count": int_sales, "total": Rp}}
        pay_summary: Dict[str, Dict[str, Any]] = {k: {"count": 0, "total": 0.0, "shopee_fee": 0.0} for k, _ in PAY_COLS}
        # Ringkasan per cabang × method group (Plaza / Kastem × Cash / BCA / Mandiri / Shopee)
        METHOD_GROUPS = ["cash", "bca", "mandiri", "shopee"]
        branch_summary: Dict[str, Dict[str, Any]] = {
            b: {**{m: 0.0 for m in METHOD_GROUPS}, "count": 0, "shopee_fee": 0.0}
            for b in ("plaza", "kastem")
        }
        daily_by_branch: Dict[str, Dict[str, float]] = {}  # {date: {"plaza": ..., "kastem": ...}}

        # Flatten rows (mirror analytics endpoint)
        excel_rows = []
        row_no = 0
        for s in sales:
            s_date = s.get("date") or ""
            s_customer = s.get("customer_name") or "Umum"
            s_method = s.get("payment_method") or "cash"
            s_bank = s.get("payment_bank") or ""
            s_pnotes = s.get("payment_notes") or ""
            s_notes = s.get("notes") or ""
            s_total_after_disc = float(s.get("total") or 0)
            s_discount = float(s.get("discount") or 0)
            s_branch = _sanitize_branch(s.get("branch"))
            pay_col = _resolve_report_payment_col(s_method, s_bank, s_branch)
            s_alamat = cust_addr_map.get(s_customer.strip().lower(), "")
            # Payment history: initial + pelunasan
            _payments = _get_sale_payments(s)
            _initial_p = next((p for p in _payments if p.get("is_initial")), None)
            s_paid_amount = float(_initial_p.get("amount") or 0) if _initial_p else min(float(s.get("cash_paid") or 0), s_total_after_disc)
            # Shopee netto adjustment: HANYA saat tidak ada filter tab pembayaran aktif
            # (mengikuti perilaku frontend `displayOmzet` — filtered pakai gross Shopee).
            _shopee_fee = float(s.get("shopee_admin_fee") or 0)
            _is_shopee = s_method in ("shopee_plaza", "shopee_kastem")
            if _is_shopee and _shopee_fee > 0 and not is_pay_filtered:
                first_row_received = max(0.0, s_paid_amount - _shopee_fee)
            else:
                first_row_received = s_paid_amount
            # Skip semua product rows sale ini bila:
            # - payment_col-nya di-hide, ATAU
            # - filter tab aktif TAPI sale ini tidak punya payment_col (data lama tanpa cabang).
            # Ini mencegah tumpang tindih data Plaza/Kastem saat filter tab aktif.
            skip_product_rows = (
                (bool(pay_col) and pay_col in hidden_set)
                or (is_pay_filtered and not pay_col)
            )
            # Kontribusi ke Ringkasan (hanya bila TIDAK di-hide)
            if pay_col and pay_col in pay_summary and not skip_product_rows:
                pay_summary[pay_col]["count"] += 1
                pay_summary[pay_col]["total"] += first_row_received
                if _is_shopee and _shopee_fee > 0 and not is_pay_filtered:
                    pay_summary[pay_col]["shopee_fee"] += _shopee_fee
                # Sheet Cabang: pay_col format = "{method}_{branch}"
                try:
                    _mg, _br = pay_col.rsplit("_", 1)
                    if _br in branch_summary and _mg in branch_summary[_br]:
                        branch_summary[_br][_mg] += first_row_received
                        branch_summary[_br]["count"] += 1
                        if _is_shopee and _shopee_fee > 0 and not is_pay_filtered:
                            branch_summary[_br]["shopee_fee"] += _shopee_fee
                        if s_date:
                            daily_by_branch.setdefault(s_date, {"plaza": 0.0, "kastem": 0.0})
                            daily_by_branch[s_date][_br] = daily_by_branch[s_date].get(_br, 0.0) + first_row_received
                except ValueError:
                    pass
            first_item = True
            for it in (s.get("items") or []):
                if skip_product_rows:
                    continue
                row_no += 1
                name = it.get("product_name") or it.get("material_name") or "-"
                qty = int(it.get("quantity") or 0)
                unit_price = float(it.get("unit_price") or 0)
                length_m = prod_length_map.get(name.strip().lower(), 0.0)
                meter = round(qty * length_m, 4) if length_m > 0 else 0
                row = {
                    "No": row_no,
                    "Tanggal": s_date,
                    "No. Nota": s.get("sale_no", ""),
                    "Alamat": s_alamat,
                    "Nama Barang": name,
                    "Pcs": qty,
                    "Meter": meter,
                    "Harga": unit_price,
                    "Disc": s_discount if first_item else 0,
                    "Jumlah": round(unit_price * qty, 2),
                    # Kolom "Total" sekarang = UANG DITERIMA pada baris ini (bukan invoice total)
                    # → SUM(kolom) sinkron dengan Omzet (Uang Diterima) di dashboard.
                    "Total": first_row_received if first_item else 0,
                    "Keterangan": s_pnotes or s_notes or "",
                }
                # Payment columns: 8 pairs × (Nominal + Tanggal)
                for k, _ in PAY_COLS:
                    row[f"{k}__n"] = 0
                    row[f"{k}__d"] = ""
                if first_item and pay_col:
                    row[f"{pay_col}__n"] = s_paid_amount
                    row[f"{pay_col}__d"] = s_date
                excel_rows.append(row)
                first_item = False
            # Pelunasan rows — satu baris per entry non-initial (juga hormati hidden_set)
            for p_ in [p for p in _payments if not p.get("is_initial")]:
                p_method = p_.get("payment_method") or "cash"
                p_bank = p_.get("payment_bank")
                p_amount = float(p_.get("amount") or 0)
                p_date = p_.get("date") or s_date
                p_col = _resolve_report_payment_col(p_method, p_bank, s_branch)
                # Sama: skip pelunasan yang metode-nya tidak match visible cols (strict mode)
                if p_col and p_col in hidden_set:
                    continue
                if is_pay_filtered and not p_col:
                    continue
                if p_col and p_col in pay_summary:
                    pay_summary[p_col]["total"] += p_amount
                    try:
                        _mg, _br = p_col.rsplit("_", 1)
                        if _br in branch_summary and _mg in branch_summary[_br]:
                            branch_summary[_br][_mg] += p_amount
                            _pd = (p_.get("date") or s_date)
                            if _pd:
                                daily_by_branch.setdefault(_pd, {"plaza": 0.0, "kastem": 0.0})
                                daily_by_branch[_pd][_br] = daily_by_branch[_pd].get(_br, 0.0) + p_amount
                    except ValueError:
                        pass
                row_no += 1
                p_label = _payment_label(p_method, p_bank)
                row = {
                    "No": row_no,
                    "Tanggal": p_date,
                    "No. Nota": s.get("sale_no", ""),
                    "Alamat": s_alamat,
                    "Nama Barang": f"(Pelunasan · {p_label})",
                    "Pcs": 0,
                    "Meter": 0,
                    "Harga": 0,
                    "Disc": 0,
                    "Jumlah": 0,
                    # Total (Uang Diterima) untuk baris pelunasan = amount pelunasan
                    "Total": p_amount,
                    "Keterangan": (p_.get("notes") or "Pelunasan sisa tagihan"),
                }
                for k, _ in PAY_COLS:
                    row[f"{k}__n"] = 0
                    row[f"{k}__d"] = ""
                if p_col:
                    row[f"{p_col}__n"] = p_amount
                    row[f"{p_col}__d"] = p_date
                excel_rows.append(row)

        # Build DataFrame with grouped columns
        if not excel_rows:
            excel_rows.append({
                "No": 1, "Tanggal": "", "No. Nota": "(Tidak ada transaksi)", "Alamat": "",
                "Nama Barang": "", "Pcs": 0, "Meter": 0, "Harga": 0, "Disc": 0,
                "Jumlah": 0, "Total": 0, "Keterangan": "",
                **{f"{k}__n": 0 for k, _ in PAY_COLS},
                **{f"{k}__d": "" for k, _ in PAY_COLS},
            })
        df = pd.DataFrame(excel_rows)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            sheet_name = f"Penjualan {month or 'periode'}"[:31]
            # Write starting from Excel row 6 (0-indexed=5) leaving rows 1-5 for company + column-group + sub-headers
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=5, header=False)
            ws = writer.sheets[sheet_name]

            # Row 1: Company name
            ws["A1"] = ci["name"].upper()
            ws["A1"].font = Font(bold=True, size=14, color="002FA7")
            # Row 2: Address
            ws["A2"] = f"{ci['address']} · HP: {ci['phone']}"
            ws["A2"].font = Font(size=9, italic=True, color="666666")
            # Row 3: Period info (+ filter status bila ada tab yg di-hide)
            period_label = f"Bulan {month}" if month else f"{date_from or '(awal)'} s/d {date_to or '(sekarang)'}"
            _filter_suffix = f" · FILTER TAB AKTIF (hide: {', '.join(sorted(hidden_set))})" if is_pay_filtered else ""
            ws["A3"] = f"Laporan Penjualan · Periode: {period_label} · {len(sales)} transaksi · {len(df)} item{_filter_suffix}"
            ws["A3"].font = Font(bold=True, size=10, color=("F97316" if is_pay_filtered else "000000"))

            # Row 4: Grouped headers (12 main + 6 payment groups)
            MAIN_HEADERS = ["No", "Tanggal", "No. Nota", "Alamat", "Nama Barang", "Pcs", "Meter", "Harga", "Disc", "Jumlah", "Total (Uang Diterima)", "Keterangan"]
            thin = Side(border_style="thin", color="333333")
            border = Border(top=thin, bottom=thin, left=thin, right=thin)
            header_fill = PatternFill("solid", fgColor="1F2937")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            pay_fills = {
                "cash_plaza": "008A00",
                "cash_kastem": "34C759",
                "bca_plaza": "002FA7",
                "bca_kastem": "4A6FE0",
                "mandiri_plaza": "E81123",
                "mandiri_kastem": "FF6B6B",
                "shopee_plaza": "F97316",  # orange 500
                "shopee_kastem": "FDBA74", # orange 300
            }
            # Row 4 main headers span both header rows 4-5 (merged)
            for i, h in enumerate(MAIN_HEADERS, start=1):
                ws.cell(row=4, column=i, value=h)
                ws.merge_cells(start_row=4, start_column=i, end_row=5, end_column=i)
                c = ws.cell(row=4, column=i)
                c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
            # Row 4 payment group headers (merged over 2 columns), Row 5 sub-headers Nominal/Tanggal
            col_cursor = len(MAIN_HEADERS) + 1
            for k, label in PAY_COLS:
                ws.merge_cells(start_row=4, start_column=col_cursor, end_row=4, end_column=col_cursor + 1)
                gh = ws.cell(row=4, column=col_cursor, value=label)
                gh.font = header_font
                gh.fill = PatternFill("solid", fgColor=pay_fills[k])
                gh.alignment = Alignment(horizontal="center", vertical="center")
                gh.border = border
                for j, sub in enumerate(("Nominal", "Tanggal")):
                    sc = ws.cell(row=5, column=col_cursor + j, value=sub)
                    sc.font = Font(bold=True, color="FFFFFF", size=9)
                    sc.fill = header_fill
                    sc.alignment = Alignment(horizontal="center", vertical="center")
                    sc.border = border
                col_cursor += 2

            # Format data rows: number columns as currency for nominal cols
            n_rows_data = len(df)
            start_data_row = 6
            end_data_row = start_data_row + n_rows_data - 1
            currency_fmt = "#,##0"
            # Currency columns: Harga (8), Disc (9), Jumlah (10), Total (11) + all pay __n columns
            currency_cols = [8, 9, 10, 11]
            col_cursor = len(MAIN_HEADERS) + 1
            for _ in PAY_COLS:
                currency_cols.append(col_cursor)      # nominal
                col_cursor += 2
            for r in range(start_data_row, end_data_row + 1):
                for cc in currency_cols:
                    cell = ws.cell(row=r, column=cc)
                    cell.number_format = currency_fmt

            # Auto-width per column
            total_cols = len(MAIN_HEADERS) + len(PAY_COLS) * 2
            default_widths = {1: 6, 2: 12, 3: 16, 4: 24, 5: 28, 6: 6, 7: 8, 8: 12, 9: 10, 10: 12, 11: 14, 12: 22}
            for i in range(1, total_cols + 1):
                ws.column_dimensions[get_column_letter(i)].width = default_widths.get(i, 12)

            # Total footer row
            total_row = end_data_row + 2
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=11, color="002FA7")
            # Sum Pcs (col 6), Meter (col 7), Disc (col 9), Jumlah (col 10), Total (col 11)
            if n_rows_data > 0:
                ws.cell(row=total_row, column=6, value=f"=SUM(F{start_data_row}:F{end_data_row})")
                ws.cell(row=total_row, column=7, value=f"=SUM(G{start_data_row}:G{end_data_row})")
                ws.cell(row=total_row, column=9, value=f"=SUM(I{start_data_row}:I{end_data_row})")
                ws.cell(row=total_row, column=10, value=f"=SUM(J{start_data_row}:J{end_data_row})")
                ws.cell(row=total_row, column=11, value=f"=SUM(K{start_data_row}:K{end_data_row})")
                # Sum each pay-nominal column
                col_cursor = len(MAIN_HEADERS) + 1
                for _ in PAY_COLS:
                    col_letter = get_column_letter(col_cursor)
                    ws.cell(row=total_row, column=col_cursor, value=f"=SUM({col_letter}{start_data_row}:{col_letter}{end_data_row})")
                    col_cursor += 2
                # Bold + currency format
                for cc in currency_cols + [6, 7]:
                    cell = ws.cell(row=total_row, column=cc)
                    cell.font = Font(bold=True, size=11, color="1F2937")
                    if cc in currency_cols:
                        cell.number_format = currency_fmt

            # Freeze panes below header row 5
            ws.freeze_panes = ws["A6"]

            # ============= SHEET 2: RINGKASAN PER TAB PEMBAYARAN =============
            summary_sheet = "Ringkasan"
            ws2 = writer.book.create_sheet(summary_sheet)
            # Header
            ws2["A1"] = ci["name"].upper()
            ws2["A1"].font = Font(bold=True, size=14, color="002FA7")
            ws2["A2"] = f"Ringkasan Penjualan · Periode: {period_label}"
            ws2["A2"].font = Font(bold=True, size=10)
            _filter_note = f"Filter tab aktif — hide: {', '.join(sorted(hidden_set))}" if is_pay_filtered else "Semua tab pembayaran ditampilkan"
            ws2["A3"] = _filter_note
            ws2["A3"].font = Font(italic=True, size=9, color=("F97316" if is_pay_filtered else "666666"))

            # Table headers
            headers2 = ["Tab Pembayaran", "Jumlah Transaksi", "Total (Uang Diterima)", "Biaya Admin", "Kontribusi %", "Status"]
            for i, h in enumerate(headers2, start=1):
                c2 = ws2.cell(row=5, column=i, value=h)
                c2.font = header_font
                c2.fill = header_fill
                c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c2.border = border

            grand_total = sum(v["total"] for v in pay_summary.values())
            r2 = 6
            for k, label in PAY_COLS:
                v = pay_summary[k]
                is_hidden = k in hidden_set
                contribution = (v["total"] / grand_total * 100) if grand_total > 0 else 0
                cells = [
                    (1, label, False),
                    (2, v["count"] if v["count"] > 0 else 0, False),
                    (3, round(v["total"], 2), True),
                    (4, round(v["shopee_fee"], 2) if v["shopee_fee"] > 0 else 0, True),
                    (5, round(contribution, 2), False),
                    (6, "HIDDEN" if is_hidden else ("aktif" if v["total"] > 0 else "—"), False),
                ]
                for col, val, is_currency in cells:
                    c2 = ws2.cell(row=r2, column=col, value=val)
                    c2.border = border
                    c2.alignment = Alignment(horizontal="left" if col == 1 else ("right" if col in (2, 3, 4) else "center"), vertical="center")
                    if is_currency:
                        c2.number_format = currency_fmt
                    elif col == 2:
                        c2.number_format = "#,##0"
                    elif col == 5:
                        c2.number_format = '0.00"%"'
                    if is_hidden:
                        c2.font = Font(color="9CA3AF", italic=True)
                        c2.fill = PatternFill("solid", fgColor="F3F4F6")
                    elif v["total"] > 0:
                        # Highlight kolom label pakai warna tab
                        if col == 1:
                            c2.fill = PatternFill("solid", fgColor=pay_fills.get(k, "1F2937"))
                            c2.font = Font(bold=True, color="FFFFFF", size=10)
                r2 += 1

            # Grand total row
            total_row2 = r2
            ws2.cell(row=total_row2, column=1, value="TOTAL OMZET (Uang Diterima)").font = Font(bold=True, size=11, color="002FA7")
            ws2.cell(row=total_row2, column=2, value=sum(v["count"] for v in pay_summary.values())).font = Font(bold=True, size=11)
            ws2.cell(row=total_row2, column=2).number_format = "#,##0"
            tc = ws2.cell(row=total_row2, column=3, value=round(grand_total, 2))
            tc.font = Font(bold=True, size=12, color="002FA7")
            tc.number_format = currency_fmt
            fc = ws2.cell(row=total_row2, column=4, value=round(sum(v["shopee_fee"] for v in pay_summary.values()), 2))
            fc.font = Font(bold=True, size=11, color="EE4D2D")
            fc.number_format = currency_fmt
            pc = ws2.cell(row=total_row2, column=5, value=100.0 if grand_total > 0 else 0)
            pc.font = Font(bold=True, size=11)
            pc.number_format = '0.00"%"'
            for col in range(1, 7):
                ws2.cell(row=total_row2, column=col).border = Border(top=Side(border_style="medium", color="002FA7"), bottom=thin, left=thin, right=thin)
                ws2.cell(row=total_row2, column=col).fill = PatternFill("solid", fgColor="EFF2FA")

            # Note if filter aktif — angka harus SAMA dengan Sheet 1 SUM(K)
            ws2.cell(row=total_row2 + 2, column=1,
                     value="Catatan: Angka TOTAL OMZET ini sama persis dengan SUM Kolom 'Total (Uang Diterima)' di Sheet 'Penjualan …' dan dengan kartu Omzet di dashboard.").font = Font(italic=True, size=8, color="6B7280")

            # Column widths for Sheet 2
            widths2 = {1: 22, 2: 16, 3: 22, 4: 16, 5: 14, 6: 12}
            for i in range(1, 7):
                ws2.column_dimensions[get_column_letter(i)].width = widths2.get(i, 14)
            ws2.freeze_panes = ws2["A6"]

            # ============= SHEET 3: RINGKASAN PER CABANG (Plaza vs Kastem) =============
            ws3 = writer.book.create_sheet("Per Cabang")
            ws3["A1"] = ci["name"].upper()
            ws3["A1"].font = Font(bold=True, size=14, color="002FA7")
            ws3["A2"] = f"Ringkasan Per Cabang · Periode: {period_label}"
            ws3["A2"].font = Font(bold=True, size=10)
            ws3["A3"] = _filter_note
            ws3["A3"].font = Font(italic=True, size=9, color=("F97316" if is_pay_filtered else "666666"))

            # --- Section A: Matrix Cabang × Metode ---
            ws3["A5"] = "A. OMZET PER CABANG × METODE PEMBAYARAN"
            ws3["A5"].font = Font(bold=True, size=11, color="002FA7")
            method_headers = ["Cabang", "Cash", "BCA", "Mandiri", "Shopee", "Biaya Admin", "Total Cabang", "Jumlah Tx", "Kontribusi %"]
            for i, h in enumerate(method_headers, start=1):
                c3 = ws3.cell(row=6, column=i, value=h)
                c3.font = header_font
                c3.fill = header_fill
                c3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c3.border = border

            _branch_totals = {b: sum(branch_summary[b][m] for m in METHOD_GROUPS) for b in ("plaza", "kastem")}
            _grand_branch = sum(_branch_totals.values())
            branch_colors = {"plaza": "002FA7", "kastem": "008A00"}
            r3 = 7
            for b in ("plaza", "kastem"):
                v = branch_summary[b]
                btotal = _branch_totals[b]
                contribution = (btotal / _grand_branch * 100) if _grand_branch > 0 else 0
                values = [
                    ("Plaza" if b == "plaza" else "Kastem", "label"),
                    (round(v["cash"], 2), "money"),
                    (round(v["bca"], 2), "money"),
                    (round(v["mandiri"], 2), "money"),
                    (round(v["shopee"], 2), "money"),
                    (round(v["shopee_fee"], 2), "money"),
                    (round(btotal, 2), "money-bold"),
                    (v["count"], "int"),
                    (round(contribution, 2), "pct"),
                ]
                for col, (val, kind) in enumerate(values, start=1):
                    c3 = ws3.cell(row=r3, column=col, value=val)
                    c3.border = border
                    if col == 1:
                        c3.fill = PatternFill("solid", fgColor=branch_colors[b])
                        c3.font = Font(bold=True, color="FFFFFF", size=11)
                        c3.alignment = Alignment(horizontal="center", vertical="center")
                    elif kind == "money" or kind == "money-bold":
                        c3.number_format = currency_fmt
                        c3.alignment = Alignment(horizontal="right", vertical="center")
                        if kind == "money-bold":
                            c3.font = Font(bold=True, size=11, color="002FA7")
                    elif kind == "int":
                        c3.number_format = "#,##0"
                        c3.alignment = Alignment(horizontal="right", vertical="center")
                    elif kind == "pct":
                        c3.number_format = '0.00"%"'
                        c3.alignment = Alignment(horizontal="center", vertical="center")
                r3 += 1

            # Grand total row (branch matrix)
            gt_row = r3
            ws3.cell(row=gt_row, column=1, value="TOTAL").font = Font(bold=True, size=11, color="002FA7")
            ws3.cell(row=gt_row, column=1).fill = PatternFill("solid", fgColor="EFF2FA")
            ws3.cell(row=gt_row, column=1).alignment = Alignment(horizontal="center")
            for mi, m in enumerate(METHOD_GROUPS, start=2):
                tv = branch_summary["plaza"][m] + branch_summary["kastem"][m]
                c3 = ws3.cell(row=gt_row, column=mi, value=round(tv, 2))
                c3.font = Font(bold=True, size=11)
                c3.number_format = currency_fmt
                c3.alignment = Alignment(horizontal="right")
                c3.fill = PatternFill("solid", fgColor="EFF2FA")
            _sum_fee = sum(branch_summary[b]["shopee_fee"] for b in ("plaza", "kastem"))
            c3 = ws3.cell(row=gt_row, column=6, value=round(_sum_fee, 2))
            c3.font = Font(bold=True, size=11, color="EE4D2D")
            c3.number_format = currency_fmt
            c3.alignment = Alignment(horizontal="right")
            c3.fill = PatternFill("solid", fgColor="EFF2FA")
            c3 = ws3.cell(row=gt_row, column=7, value=round(_grand_branch, 2))
            c3.font = Font(bold=True, size=12, color="002FA7")
            c3.number_format = currency_fmt
            c3.alignment = Alignment(horizontal="right")
            c3.fill = PatternFill("solid", fgColor="EFF2FA")
            c3 = ws3.cell(row=gt_row, column=8, value=sum(branch_summary[b]["count"] for b in ("plaza", "kastem")))
            c3.font = Font(bold=True, size=11)
            c3.number_format = "#,##0"
            c3.alignment = Alignment(horizontal="right")
            c3.fill = PatternFill("solid", fgColor="EFF2FA")
            c3 = ws3.cell(row=gt_row, column=9, value=100.0 if _grand_branch > 0 else 0)
            c3.font = Font(bold=True, size=11)
            c3.number_format = '0.00"%"'
            c3.alignment = Alignment(horizontal="center")
            c3.fill = PatternFill("solid", fgColor="EFF2FA")

            # --- Section B: Daily Omzet per Branch ---
            section_b_row = gt_row + 3
            ws3.cell(row=section_b_row, column=1, value="B. OMZET HARIAN PER CABANG").font = Font(bold=True, size=11, color="002FA7")
            daily_headers = ["Tanggal", "Plaza", "Kastem", "Total Harian", "Selisih (Plaza − Kastem)"]
            for i, h in enumerate(daily_headers, start=1):
                c3 = ws3.cell(row=section_b_row + 1, column=i, value=h)
                c3.font = header_font
                c3.fill = header_fill
                c3.alignment = Alignment(horizontal="center", vertical="center")
                c3.border = border

            dr = section_b_row + 2
            _sum_p = _sum_k = 0.0
            for d_iso in sorted(daily_by_branch.keys()):
                plaza_v = daily_by_branch[d_iso].get("plaza", 0.0)
                kastem_v = daily_by_branch[d_iso].get("kastem", 0.0)
                _sum_p += plaza_v
                _sum_k += kastem_v
                values = [
                    (d_iso, "text"),
                    (round(plaza_v, 2), "money"),
                    (round(kastem_v, 2), "money"),
                    (round(plaza_v + kastem_v, 2), "money-bold"),
                    (round(plaza_v - kastem_v, 2), "money-signed"),
                ]
                for col, (val, kind) in enumerate(values, start=1):
                    c3 = ws3.cell(row=dr, column=col, value=val)
                    c3.border = border
                    if kind == "money" or kind == "money-bold" or kind == "money-signed":
                        c3.number_format = currency_fmt
                        c3.alignment = Alignment(horizontal="right", vertical="center")
                        if kind == "money-bold":
                            c3.font = Font(bold=True, color="002FA7")
                        elif kind == "money-signed":
                            c3.font = Font(color=("008A00" if val >= 0 else "E81123"))
                    elif kind == "text":
                        c3.font = Font(name="Consolas", size=10)
                        c3.alignment = Alignment(horizontal="center", vertical="center")
                dr += 1
            # Daily grand total
            if daily_by_branch:
                ws3.cell(row=dr, column=1, value="TOTAL").font = Font(bold=True, size=11, color="002FA7")
                ws3.cell(row=dr, column=1).fill = PatternFill("solid", fgColor="EFF2FA")
                ws3.cell(row=dr, column=1).alignment = Alignment(horizontal="center")
                for col, val in enumerate([round(_sum_p, 2), round(_sum_k, 2), round(_sum_p + _sum_k, 2), round(_sum_p - _sum_k, 2)], start=2):
                    c3 = ws3.cell(row=dr, column=col, value=val)
                    c3.font = Font(bold=True, size=11, color="002FA7" if col in (2, 3, 4) else ("008A00" if val >= 0 else "E81123"))
                    c3.number_format = currency_fmt
                    c3.alignment = Alignment(horizontal="right")
                    c3.fill = PatternFill("solid", fgColor="EFF2FA")
                    c3.border = Border(top=Side(border_style="medium", color="002FA7"), bottom=thin, left=thin, right=thin)

            # Column widths & freeze
            widths3 = {1: 14, 2: 18, 3: 18, 4: 18, 5: 20, 6: 16, 7: 20, 8: 12, 9: 14}
            for i in range(1, 10):
                ws3.column_dimensions[get_column_letter(i)].width = widths3.get(i, 14)
            ws3.freeze_panes = ws3["A7"]

        buf.seek(0)
        fname_period = month or (f"{date_from}_sd_{date_to}" if (date_from or date_to) else "semua")
        _f_tag = "_filtered" if is_pay_filtered else ""
        fname = f"Laporan_Penjualan_{fname_period}{_f_tag}.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )


    # ---------------- Laporan Rincian Penjualan Online Shopee ----------------
    @router.get("/sales/report/shopee-rincian")
    async def sales_shopee_rincian(
        user: dict = Depends(require_super_admin),
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ):
        """Split per outlet (Plaza / Kastem) untuk transaksi Shopee."""
        q: Dict[str, Any] = {
            "status": {"$nin": ["cancelled", "void", "voided", "canceled"]},
            "payment_method": {"$in": ["shopee_plaza", "shopee_kastem"]},
        }
        if date_from or date_to:
            q["date"] = {}
            if date_from:
                q["date"]["$gte"] = date_from
            if date_to:
                q["date"]["$lte"] = date_to
        sales = await db.sales.find(q, {"_id": 0}).sort("created_at", 1).to_list(length=20000)

        products_p = await db.products.find({}, {"_id": 0, "name": 1, "length_meter": 1}).to_list(length=5000)
        prod_length_map = {(p.get("name") or "").strip().lower(): float(p.get("length_meter") or 0) for p in products_p}

        def _row_from_sale(s: Dict[str, Any]) -> Dict[str, Any]:
            items = s.get("items") or []
            pesanan_parts = []
            pcs = 0
            meter = 0.0
            harga_satuan = 0.0
            for it in items:
                name = (it.get("product_name") or it.get("material_name") or "-").strip()
                qty = int(it.get("quantity") or 0)
                up = float(it.get("unit_price") or 0)
                pesanan_parts.append(name)
                pcs += qty
                if harga_satuan == 0:
                    harga_satuan = up
                length_m = prod_length_map.get(name.lower(), 0.0)
                if length_m > 0:
                    meter += qty * length_m
            pesanan = " · ".join(pesanan_parts) if pesanan_parts else "-"
            jumlah = float(s.get("subtotal") or 0)
            saldo_masuk = s.get("saldo_masuk")
            if saldo_masuk is None or saldo_masuk == "":
                saldo_val = None
                potongan = None
                persentase = None
            else:
                saldo_val = float(saldo_masuk)
                potongan = round(jumlah - saldo_val, 2)
                persentase = round((potongan / jumlah * 100), 2) if jumlah > 0 else 0
            return {
                "id": s.get("id"),
                "sale_id": s.get("id"),
                "sale_no": s.get("sale_no"),
                "date": s.get("date"),
                "nama": s.get("customer_name") or "Umum",
                "pesanan": pesanan,
                "pcs": pcs,
                "meter": round(meter, 4),
                "harga_satuan": round(harga_satuan, 2),
                "jumlah": round(jumlah, 2),
                "saldo_masuk": saldo_val,
                "potongan": potongan,
                "persentase": persentase,
            }

        plaza_rows = []
        kastem_rows = []
        plaza_totals = {"jumlah": 0.0, "saldo_masuk": 0.0, "potongan": 0.0}
        kastem_totals = {"jumlah": 0.0, "saldo_masuk": 0.0, "potongan": 0.0}
        for s in sales:
            row = _row_from_sale(s)
            target = plaza_rows if s.get("payment_method") == "shopee_plaza" else kastem_rows
            totals = plaza_totals if s.get("payment_method") == "shopee_plaza" else kastem_totals
            target.append(row)
            totals["jumlah"] += float(row["jumlah"] or 0)
            totals["saldo_masuk"] += float(row["saldo_masuk"] or 0)
            totals["potongan"] += float(row["potongan"] or 0)
        for t in (plaza_totals, kastem_totals):
            for k in t:
                t[k] = round(t[k], 2)

        return {
            "plaza": {"rows": plaza_rows, "totals": plaza_totals, "count": len(plaza_rows)},
            "kastem": {"rows": kastem_rows, "totals": kastem_totals, "count": len(kastem_rows)},
        }



    @router.patch("/sales/{sale_id}/saldo-masuk")
    async def sales_update_saldo_masuk(sale_id: str, payload: SaldoMasukIn, user: dict = Depends(require_super_admin)):
        existing = await db.sales.find_one({"id": sale_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        val = payload.saldo_masuk
        if val is not None and float(val) < 0:
            raise HTTPException(status_code=400, detail="Saldo Masuk tidak boleh negatif")
        await db.sales.update_one(
            {"id": sale_id},
            {"$set": {
                "saldo_masuk": float(val) if val is not None else None,
                "saldo_masuk_updated_at": datetime.now(timezone.utc).isoformat(),
                "saldo_masuk_updated_by": user.get("email"),
            }},
        )
        return {"ok": True, "sale_id": sale_id, "saldo_masuk": val}



    @router.post("/sales/{sale_id}/pay-remaining")
    async def sales_pay_remaining(sale_id: str, payload: PayRemainingIn, user: dict = Depends(require_super_admin)):
        """Pelunasan sisa tagihan (DP → LUNAS). Otomatis catat ke Jurnal Kas."""
        existing = await db.sales.find_one({"id": sale_id})
        if not existing:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        current_sisa = round(float(existing.get("sisa_tagihan") or 0), 2)
        current_status = existing.get("status") or ("dp" if current_sisa > 0.01 else "paid")
        if current_status != "dp" or current_sisa <= 0.01:
            raise HTTPException(status_code=400, detail="Transaksi ini sudah LUNAS")
        amount = round(float(payload.amount or 0), 2)
        if amount <= 0:
            raise HTTPException(status_code=400, detail="Nominal pembayaran harus > 0")
        # Toleransi kecil - allow overpay up to 1 rupiah untuk rounding
        if amount > current_sisa + 0.01:
            raise HTTPException(status_code=400, detail=f"Nominal melebihi sisa tagihan (Rp {current_sisa:,.0f})")
        pay_date = (payload.date or datetime.now(timezone.utc).date().isoformat())[:10]
        try:
            datetime.strptime(pay_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Format tanggal harus YYYY-MM-DD")

        new_sisa = round(max(0.0, current_sisa - amount), 2)
        new_status = "paid" if new_sisa <= 0.01 else "dp"
        new_cash_paid = round(float(existing.get("cash_paid") or 0) + amount, 2)
        now_iso = datetime.now(timezone.utc).isoformat()

        # Log entry
        payment_entry = {
            "id": str(uuid.uuid4()),
            "amount": amount,
            "payment_method": payload.payment_method or "cash",
            "payment_bank": (payload.payment_bank or "").strip() or None,
            "date": pay_date,
            "notes": (payload.notes or "").strip() or None,
            "is_initial": False,
            "created_at": now_iso,
            "created_by": user.get("email"),
        }

        await db.sales.update_one(
            {"id": sale_id},
            {
                "$set": {
                    "sisa_tagihan": new_sisa,
                    "status": new_status,
                    "cash_paid": new_cash_paid,
                    "last_payment_at": now_iso,
                    "last_payment_by": user.get("email"),
                },
                "$push": {"payments": payment_entry},
            },
        )

        # Auto insert Jurnal Kas
        try:
            acc_code, acc_label = _resolve_payment_account(payload.payment_method, payload.payment_bank)
            desc = f"Pelunasan {existing.get('sale_no')} — {existing.get('customer_name')} · {acc_label}"
            if new_status == "paid":
                desc += " · LUNAS"
            else:
                desc += f" · sisa Rp {new_sisa:,.0f}"
            if payment_entry["notes"]:
                desc += f" ({payment_entry['notes']})"
            await _insert_cash_transaction(
                account_code=acc_code,
                description=desc,
                amount=amount,
                reference=existing.get("sale_no"),
                date_iso=pay_date,
                auto=True,
                created_by=user.get("email"),
            )
        except Exception as ex:
            logger.warning(f"Cashbook auto-insert (pay-remaining) failed: {ex}")

        return {
            "ok": True,
            "sale_id": sale_id,
            "amount_paid": amount,
            "sisa_tagihan": new_sisa,
            "status": new_status,
            "cash_paid_total": new_cash_paid,
        }


    def _payment_label(method: str, bank: Optional[str]) -> str:
        """Return human label for payment method."""
        pm = (method or "").lower()
        if pm == "cash" or pm == "tunai":
            return "Cash / Tunai"
        if pm == "transfer":
            b = (bank or "").lower()
            if b == "mandiri":
                return "Transfer Mandiri"
            if b == "bca":
                return "Transfer BCA"
            return "Transfer"
        if pm == "shopee_plaza":
            return "Shopee Plaza"
        if pm == "shopee_kastem":
            return "Shopee Kastem"
        return method or "-"


    def _get_sale_payments(sale: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Return unified payment history for a sale.
        - Untuk sale baru: gunakan sale.payments[] langsung
        - Untuk sale lama (tanpa payments[]): synthesize entry pertama dari sale-level info
        """
        payments = sale.get("payments") or []
        if payments:
            # Ensure semua entry punya is_initial & label
            out = []
            for p in payments:
                e = dict(p)
                e.setdefault("is_initial", False)
                e["label"] = _payment_label(e.get("payment_method"), e.get("payment_bank"))
                out.append(e)
            return out
        # Backward-compat: synthesize dari sale-level
        total = float(sale.get("total") or 0)
        cash_paid = float(sale.get("cash_paid") or 0)
        initial_amount = round(min(cash_paid, total), 2)
        entries = []
        if initial_amount > 0 or (sale.get("status") == "paid" and total == 0):
            entries.append({
                "id": f"legacy-initial-{sale.get('id')}",
                "amount": initial_amount,
                "payment_method": sale.get("payment_method") or "cash",
                "payment_bank": sale.get("payment_bank"),
                "date": sale.get("date"),
                "notes": sale.get("payment_notes"),
                "is_initial": True,
                "created_at": sale.get("created_at"),
                "created_by": sale.get("cashier"),
                "label": _payment_label(sale.get("payment_method"), sale.get("payment_bank")),
            })
        return entries


    @router.get("/sales/{sale_id}/payments")
    async def sales_payments_get(sale_id: str, user: dict = Depends(require_super_admin)):
        """Riwayat pembayaran (DP + Pelunasan) untuk satu transaksi."""
        sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
        if not sale:
            raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
        payments = _get_sale_payments(sale)
        # Sort by date ascending, initial first if same date
        payments.sort(key=lambda p: (str(p.get("date") or ""), 0 if p.get("is_initial") else 1, str(p.get("created_at") or "")))
        total_paid = round(sum(float(p.get("amount") or 0) for p in payments), 2)
        total_amount = float(sale.get("total") or 0)
        sisa = round(max(0.0, total_amount - total_paid), 2)
        return {
            "sale_id": sale_id,
            "sale_no": sale.get("sale_no"),
            "customer_name": sale.get("customer_name"),
            "total": total_amount,
            "total_paid": total_paid,
            "sisa_tagihan": sisa,
            "status": sale.get("status") or ("dp" if sisa > 0.01 else "paid"),
            "payments": payments,
        }



    @router.get("/sales/report/analytics")
    async def sales_analytics(
        user: dict = Depends(require_super_admin),
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        customer: Optional[str] = None,
    ):
        """Analytics untuk Laporan Penjualan.

        OMZET = UANG DITERIMA (DP + Pelunasan) yang tanggal PEMBAYARAN-nya berada
        dalam periode filter. Bukan total invoice. Hutang / sisa tagihan pelanggan
        DIKELUARKAN dari angka Omzet Utama sehingga sinkron dengan Buku Kas.
        Untuk Shopee, dipakai NETTO (payment amount − biaya admin Shopee).
        """
        q: Dict[str, Any] = {"status": {"$nin": ["cancelled", "void", "voided", "canceled"]}}
        # Broaden query: include sale bila sale.date ATAU payments.date ada di dalam periode.
        if date_from or date_to:
            date_conds: Dict[str, Any] = {}
            if date_from:
                date_conds["$gte"] = date_from
            if date_to:
                date_conds["$lte"] = date_to
            q["$or"] = [
                {"date": date_conds},
                {"payments.date": date_conds},
            ]
        if customer:
            safe = re.escape(customer.strip())
            q["customer_name"] = {"$regex": safe, "$options": "i"}
        sales = await db.sales.find(q, {"_id": 0}).sort("created_at", 1).to_list(length=20000)

        def _in_period(d: Optional[str]) -> bool:
            d10 = (d or "")[:10]
            if not d10:
                return False
            if date_from and d10 < date_from:
                return False
            if date_to and d10 > date_to:
                return False
            return True

        # Preload customer address map (by name, case-insensitive) & product length_meter map (by name)
        customers = await db.customers.find({}, {"_id": 0, "name": 1, "address": 1}).to_list(length=5000)
        cust_addr_map = {(c.get("name") or "").strip().lower(): (c.get("address") or "").strip() for c in customers}
        products_p = await db.products.find({}, {"_id": 0, "name": 1, "length_meter": 1}).to_list(length=5000)
        prod_length_map = {(p.get("name") or "").strip().lower(): float(p.get("length_meter") or 0) for p in products_p}

        # Flatten per-item rows
        rows: List[Dict[str, Any]] = []
        product_totals: Dict[str, Dict[str, float]] = {}
        daily_series: Dict[str, float] = {}
        method_totals: Dict[str, float] = {}
        weekly_total = 0.0
        period_total = 0.0
        total_shopee_fees = 0.0
        total_shopee_gross = 0.0

        today = datetime.now(timezone.utc).date()
        week_start = today - timedelta(days=today.weekday())
        week_start_iso = week_start.isoformat()

        for s in sales:
            s_date = s.get("date") or ""
            sale_in_period = _in_period(s_date)
            s_customer = s.get("customer_name") or "Umum"
            s_method = s.get("payment_method") or "cash"
            s_bank = s.get("payment_bank") or ""
            s_pnotes = s.get("payment_notes") or ""
            s_notes = s.get("notes") or ""
            s_total_after_disc = float(s.get("total") or 0)
            s_discount = float(s.get("discount") or 0)
            s_subtotal = float(s.get("subtotal") or 0)
            s_branch = _sanitize_branch(s.get("branch"))  # "plaza" | "kastem" | None
            # DP: cash_paid = jumlah aktual diterima; sisa_tagihan = piutang
            s_cash_paid = float(s.get("cash_paid") or 0)
            s_sisa = float(s.get("sisa_tagihan") if s.get("sisa_tagihan") is not None else max(0, s_total_after_disc - s_cash_paid))
            # kolom pembayaran (row pertama = INITIAL DP amount saja, bukan cumulative).
            # payments[0] adalah initial (DP awal). Fallback ke cash_paid utk data legacy.
            _all_payments = _get_sale_payments(s)
            _initial_p = next((p for p in _all_payments if p.get("is_initial")), None)
            if _initial_p is not None:
                s_paid_amount = float(_initial_p.get("amount") or 0)
            else:
                s_paid_amount = min(s_cash_paid, s_total_after_disc)
            s_status = s.get("status") or ("dp" if s_sisa > 0.01 else "paid")
            # Payment column key (Cash/BCA/Mandiri + Plaza/Kastem) — derived from method+bank+branch
            pay_col = _resolve_report_payment_col(s_method, s_bank, s_branch)
            # Alamat: prefer customer master lookup by name (case-insensitive)
            s_alamat = cust_addr_map.get(s_customer.strip().lower(), "")
            s_items = s.get("items") or []
            # -------- Product rows: hanya jika sale.date berada di periode --------
            if sale_in_period:
                first_item = True
                for it in s_items:
                    name = it.get("product_name") or it.get("material_name") or "-"
                    qty = int(it.get("quantity") or 0)
                    unit_price = float(it.get("unit_price") or 0)
                    subtotal_item = float(it.get("subtotal") or (unit_price * qty))
                    size = it.get("size") or "-"
                    length_m = prod_length_map.get(name.strip().lower(), 0.0)
                    meter = round(qty * length_m, 4) if length_m > 0 else 0.0
                    # payment nominal awal (initial DP) hanya muncul di baris pertama
                    # DAN hanya jika tanggal DP initial berada di periode (agar tidak menghitung DP
                    # di luar periode saat sale.date OoP tapi ada pelunasan in-period).
                    _init_date = ((_initial_p or {}).get("date") or s_date)[:10]
                    _pay_nominal_row = s_paid_amount if (first_item and _in_period(_init_date)) else 0
                    _pay_date_row = _init_date if (first_item and _in_period(_init_date)) else ""
                    rows.append({
                        "date": s_date,
                        "customer_name": s_customer,
                        "alamat": s_alamat,
                        "sale_no": s.get("sale_no"),
                        "product_name": name,
                        "size": size,
                        "pcs": qty,
                        "meter": meter,
                        "quantity": qty,  # legacy alias
                        "unit_price": unit_price,
                        "total": subtotal_item,
                        "sale_total": s_total_after_disc,
                        "sale_subtotal": s_subtotal,
                        "sale_discount": s_discount,
                        "sale_cash_paid": s_cash_paid,
                        "sale_sisa_tagihan": s_sisa,
                        "sale_status": s_status,  # "paid" (LUNAS) atau "dp"
                        "keterangan": s_pnotes or s_notes or "",
                        "branch": s_branch,
                        "payment_method": s_method,
                        "payment_bank": s_bank,
                        "payment_notes": s_pnotes,
                        "payment_column": pay_col,
                        "payment_nominal_on_row": _pay_nominal_row,
                        "payment_date_on_row": _pay_date_row,
                        "is_first_item_of_sale": first_item,
                    })
                    first_item = False
                    pk = name
                    if pk not in product_totals:
                        product_totals[pk] = {"qty": 0, "total": 0.0}
                    product_totals[pk]["qty"] += qty
                    product_totals[pk]["total"] += subtotal_item
            # --- Pelunasan rows (per-payment entries selain initial DP) ---
            # Tampilkan HANYA jika tanggal pelunasan berada dalam periode filter
            pelunasan_entries = [p for p in _all_payments if not p.get("is_initial")]
            for p_ in pelunasan_entries:
                p_date = (p_.get("date") or s_date)[:10]
                if not _in_period(p_date):
                    continue
                p_method = p_.get("payment_method") or "cash"
                p_bank = p_.get("payment_bank")
                p_amount = float(p_.get("amount") or 0)
                p_col = _resolve_report_payment_col(p_method, p_bank, s_branch)
                p_label = _payment_label(p_method, p_bank)
                rows.append({
                    "date": p_date,
                    "customer_name": s_customer,
                    "alamat": s_alamat,
                    "sale_no": s.get("sale_no"),
                    "product_name": f"(Pelunasan · {p_label})",
                    "size": "-",
                    "pcs": 0,
                    "meter": 0.0,
                    "quantity": 0,
                    "unit_price": 0.0,
                    "total": 0.0,
                    "sale_total": s_total_after_disc,
                    "sale_subtotal": s_subtotal,
                    "sale_discount": s_discount,
                    "sale_cash_paid": s_cash_paid,
                    "sale_sisa_tagihan": s_sisa,
                    "sale_status": s_status,
                    "keterangan": (p_.get("notes") or "").strip() or "Pelunasan sisa tagihan",
                    "branch": s_branch,
                    "payment_method": p_method,
                    "payment_bank": p_bank,
                    "payment_notes": p_.get("notes") or "",
                    "payment_column": p_col,
                    "payment_nominal_on_row": p_amount,
                    "payment_date_on_row": p_date,
                    "is_first_item_of_sale": False,
                    "is_pelunasan_row": True,
                    "pelunasan_id": p_.get("id"),
                })
            # -------- OMZET AKUMULASI: hanya uang diterima (payment.date in period) --------
            _shopee_fee = float(s.get("shopee_admin_fee") or 0)
            _is_shopee = s_method in ("shopee_plaza", "shopee_kastem")
            for p_ in _all_payments:
                p_date = (p_.get("date") or s_date)[:10]
                if not _in_period(p_date):
                    continue
                p_amount = float(p_.get("amount") or 0)
                if p_amount <= 0:
                    continue
                # Shopee: potong admin fee dari pembayaran initial (model NETTO)
                netto_amount = p_amount
                if _is_shopee and p_.get("is_initial") and _shopee_fee > 0:
                    netto_amount = max(0.0, p_amount - _shopee_fee)
                    total_shopee_fees += _shopee_fee
                    total_shopee_gross += p_amount
                period_total += netto_amount
                daily_series[p_date] = daily_series.get(p_date, 0) + netto_amount
                if p_date >= week_start_iso:
                    weekly_total += netto_amount
                # method breakdown
                m_key = p_.get("payment_method") or "cash"
                if m_key == "transfer" and p_.get("payment_bank"):
                    m_key = f"transfer_{str(p_.get('payment_bank')).lower()}"
                method_totals[m_key] = method_totals.get(m_key, 0) + netto_amount

        top_products = sorted(
            [{"name": k, "qty": int(v["qty"]), "total": round(v["total"], 2)} for k, v in product_totals.items()],
            key=lambda x: x["total"], reverse=True,
        )
        top_product = top_products[0]["name"] if top_products else None
        daily_data = [{"date": d, "total": round(v, 2)} for d, v in sorted(daily_series.items())]

        return {
            "rows": rows,
            "summary": {
                "period_total": round(period_total, 2),  # NETTO (setelah dikurangi biaya admin Shopee)
                "period_total_gross": round(period_total + total_shopee_fees, 2),  # sebelum dikurangi admin fee
                "shopee_gross": round(total_shopee_gross, 2),
                "shopee_admin_fee": round(total_shopee_fees, 2),
                "shopee_netto": round(total_shopee_gross - total_shopee_fees, 2),
                "weekly_total": round(weekly_total, 2),
                "week_start": week_start_iso,
                "transaction_count": sum(1 for s in sales if _in_period(s.get("date") or "")),
                "item_count": len(rows),
                "top_product": top_product,
            },
            "top_products": top_products[:10],
            "daily_series": daily_data,
            "method_breakdown": [
                {"method": k, "total": round(v, 2)} for k, v in sorted(method_totals.items(), key=lambda x: x[1], reverse=True)
            ],
        }


    def _resolve_report_payment_col(payment_method: str, payment_bank: Optional[str], branch: Optional[str]) -> Optional[str]:
        """Map (payment_method, payment_bank, branch) -> report column key.
        Returns one of:
          cash_plaza, cash_kastem, bca_plaza, bca_kastem, mandiri_plaza, mandiri_kastem,
          shopee_plaza, shopee_kastem, or None.
        - Shopee: branch di-ambil dari payment_method (shopee_plaza / shopee_kastem).
        - Kolom lain (Cash/BCA/Mandiri): butuh sale.branch. Jika sale.branch kosong
          (transaksi lama pre-fitur cabang), default ke "plaza" agar nominal tetap muncul.
        """
        pm = (payment_method or "").lower()
        # Shopee columns — plaza/kastem sudah baked in payment_method
        if pm == "shopee_plaza":
            return "shopee_plaza"
        if pm == "shopee_kastem":
            return "shopee_kastem"
        # Kolom lain: default branch = plaza (fallback untuk data lama)
        b = (branch or "plaza").lower()
        if b not in ("plaza", "kastem"):
            b = "plaza"
        bank = (payment_bank or "").lower()
        if pm in ("cash", "tunai"):
            return f"cash_{b}"
        if pm == "transfer":
            if bank == "bca":
                return f"bca_{b}"
            if bank in ("mandiri", "mdr"):
                return f"mandiri_{b}"
            return None
        return None


    @router.get("/sales/stats/today")
    async def sales_stats_today(user: dict = Depends(require_super_admin)):
        today = datetime.now(timezone.utc).date().isoformat()
        items = await db.sales.find({"date": today}, {"_id": 0}).to_list(length=5000)
        total_today = sum(float(s.get("total", 0)) for s in items)
        # This month
        month_start = datetime.now(timezone.utc).date().replace(day=1).isoformat()
        month_items = await db.sales.find({"date": {"$gte": month_start}}, {"_id": 0, "total": 1}).to_list(length=20000)
        total_month = sum(float(s.get("total", 0)) for s in month_items)
        return {
            "date": today,
            "count_today": len(items),
            "total_today": round(total_today, 2),
            "count_month": len(month_items),
            "total_month": round(total_month, 2),
        }



    # ---- Shopee admin fee endpoint & helper (extracted with sales) ----
    @router.post("/sales/shopee/bulk-set-admin-fee")
    async def sales_bulk_set_shopee_fee(
        payload: Dict[str, Any],
        user: dict = Depends(require_super_admin),
    ):
        """Bulk-set biaya admin Shopee untuk transaksi Shopee dalam periode.

        Body:
          {"date_from": "2026-07-01", "date_to": "2026-07-31",
           "mode": "flat" | "percent" | "per_sale",
           "value": 5000,               # nominal utk flat, atau persentase utk percent (e.g., 5 = 5%)
           "sales": [{"sale_id": "...", "amount": 12345}, ...]  # utk mode per_sale
          }

        Response: {updated_count, total_fee, sample:[...]}
        """
        date_from = payload.get("date_from")
        date_to = payload.get("date_to")
        mode = payload.get("mode") or "flat"
        q: Dict[str, Any] = {"payment_method": {"$in": ["shopee_plaza", "shopee_kastem"]}}
        if date_from and date_to:
            q["date"] = {"$gte": date_from, "$lte": date_to}
        elif date_from:
            q["date"] = {"$gte": date_from}
        elif date_to:
            q["date"] = {"$lte": date_to}

        sales = await db.sales.find(q, {"_id": 0}).to_list(length=10000)
        updated = 0
        total_fee = 0.0
        sample: List[Dict[str, Any]] = []

        if mode == "per_sale":
            per_sale_map = {row["sale_id"]: float(row.get("amount") or 0) for row in payload.get("sales", []) if row.get("sale_id")}
            for s in sales:
                fee = per_sale_map.get(s["id"])
                if fee is None:
                    continue
                fee = round(fee, 2)
                if fee < 0:
                    continue
                await _apply_shopee_admin_fee_update(s, fee, user)
                updated += 1
                total_fee += fee
                if len(sample) < 10:
                    sample.append({"sale_no": s.get("sale_no"), "date": s.get("date"), "fee": fee})
        else:
            value = float(payload.get("value") or 0)
            if value < 0:
                raise HTTPException(status_code=400, detail="Value harus >= 0")
            for s in sales:
                gross = float(s.get("total") or 0)
                if mode == "percent":
                    fee = round(gross * value / 100.0, 2)
                else:  # flat
                    fee = round(value, 2)
                await _apply_shopee_admin_fee_update(s, fee, user)
                updated += 1
                total_fee += fee
                if len(sample) < 10:
                    sample.append({"sale_no": s.get("sale_no"), "date": s.get("date"), "gross": gross, "fee": fee})

        return {
            "ok": True,
            "mode": mode,
            "updated_count": updated,
            "total_fee": round(total_fee, 2),
            "sample": sample,
        }


    async def _apply_shopee_admin_fee_update(sale: Dict[str, Any], new_fee: float, user: dict):
        """Update field shopee_admin_fee di sale + rekonsiliasi baris kas 301-SPP/SPK (model NETTO).

        Strategi (single-entry NETTO):
        - Update sale.shopee_admin_fee = new_fee
        - Delete SEMUA cash_tx untuk sale ini (301-SPP / 301-SPK / legacy 502-SHP)
        - Re-insert 1 baris cash tx pemasukan dgn amount = gross_paid − new_fee (NETTO)
        - Amount pemasukan Shopee di Buku Kas otomatis netto, tidak ada baris pengeluaran fee terpisah
        """
        sale_no = sale.get("sale_no")
        if not sale_no:
            return
        await db.sales.update_one(
            {"id": sale["id"]},
            {"$set": {"shopee_admin_fee": round(new_fee, 2), "updated_at": datetime.now(timezone.utc).isoformat(), "updated_by": user.get("email")}},
        )
        # Clear existing cash tx untuk sale ini di akun terkait Shopee (rekonsiliasi netto)
        method = sale.get("payment_method") or "shopee_plaza"
        acc_code, acc_label = _resolve_payment_account(method, None)
        await db.cash_transactions.delete_many({"reference": sale_no, "account_code": {"$in": [acc_code, "502-SHP"]}})
        # Hitung netto: gross_paid - fee
        total = float(sale.get("total") or 0)
        cash_paid = float(sale.get("cash_paid") or 0)
        gross_recorded = min(cash_paid, total)  # sama dgn POST /sales logic
        if gross_recorded <= 0:
            return
        netto = round(gross_recorded - float(new_fee or 0), 2)
        if netto <= 0.01:
            return  # netto 0 atau negatif → tidak insert baris
        # Determine status label
        sisa = round(float(sale.get("sisa_tagihan") or 0), 2)
        status_tag = "LUNAS" if sisa <= 0.01 else f"DP (sisa Rp {sisa:,.0f})"
        fee_tag = f" · − Admin Rp {new_fee:,.0f}" if new_fee > 0 else ""
        desc = f"Penjualan {sale_no} — {sale.get('customer_name') or 'Umum'} · {acc_label} · {status_tag}{fee_tag} [NETTO RESYNC]"
        try:
            await _insert_cash_transaction(
                account_code=acc_code,
                description=desc,
                amount=netto,
                reference=sale_no,
                date_iso=sale.get("date"),
                auto=True,
                created_by=user.get("email"),
            )
        except Exception as ex:
            logger.warning(f"Insert netto cash tx for {sale_no} failed: {ex}")


    return router
